#!/usr/bin/env python3
"""Oracle fidelity revision: Evidence/FA consistency, LR explanations, PO caps, metadata, caches."""
from __future__ import annotations

import csv
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
MERIDIAN = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"

HEADER_DARK = "3B2F2A"
HEADER_MID = "5C4A3A"
HEADER_LIGHT = PatternFill(start_color=HEADER_MID, end_color=HEADER_MID, fill_type="solid")
HEADER_DARK_FILL = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
OLD_NAVY = {"001F4E79", "000B2E4F", "1F4E79", "0B2E4F", "FF1F4E79", "FF0B2E4F"}

CREATOR = "Yanou IT Asset Specialist"
APPLICATION = "Microsoft Excel"

MISSING_FA = {"MD-00130", "MD-00131", "MD-00132"}
UNDER_THRESHOLD = {"MD-00130", "MD-00131"}


def load_csv(name: str) -> list[dict]:
    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def gid(pattern: str, text: str | None) -> str | None:
    if not text:
        return None
    m = re.search(pattern, str(text))
    return m.group(1) if m else None


def slot(*parts: object, mod: int) -> int:
    h = 0
    for p in parts:
        for c in str(p):
            h = (h * 131 + ord(c)) % max(mod, 1)
    return h


def source_ledger_map(wb) -> dict[str, str]:
    sl = wb["Source Ledger"]
    out: dict[str, str] = {}
    for r in range(5, sl.max_row + 1):
        tag = sl.cell(r, 2).value
        fa = sl.cell(r, 1).value
        if tag and fa:
            out[str(tag)] = str(fa)
    return out


def po_by_tag() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for row in load_csv("hardware_purchase_orders.csv"):
        po = row["purchase_order"]
        for tag in re.split(r"[;\s]+", row["asset_tags"]):
            tag = tag.strip()
            if tag.startswith("MD-"):
                mapping[tag] = po
    return mapping


def money_clauses(tag: str, po: str | None, fa: str | None) -> list[str]:
    if tag in UNDER_THRESHOLD:
        return [
            f"PO {po} on file; under-threshold asset {tag} - no FAR row expected.",
            f"{po} covers {tag} as an expense-class buy; FAR silence is expected under ITAM-001 §7.",
            f"Purchasing {po} lists {tag}; capitalization not required - no FA line.",
            f"{tag}: under $2,500 gate. {po} present; missing FA expected.",
        ]
    if tag == "MD-00132":
        return [
            f"PO {po} on file; no matching FAR row for {tag} - capital gap.",
            f"{po} in hardware_purchase_orders.csv; FAR silent on capital-qualifying {tag}.",
            f"Purchasing {po} without a capital line for {tag} ($6,470).",
            f"{tag}: {po} approved but FAR extract has no row.",
        ]
    if po and fa:
        return [
            f"Procurement {po} maps to ledger {fa}.",
            f"{po} / {fa} support the cost basis.",
            f"Purchase order {po} and FAR row {fa} both cite {tag}.",
            f"Cost path: {po} into {fa}.",
            f"{po} and {fa} agree on {tag}.",
            f"Ledger {fa} backs {po} for {tag}.",
            f"FAR extract includes {fa} for {tag} via {po}.",
            f"Capital trail {tag}: {po} -> {fa}.",
        ]
    if fa:
        return [
            f"Fixed-asset extract includes {fa} for {tag}.",
            f"FAR row {fa} covers {tag}.",
            f"Ledger tag {fa} present for {tag}.",
        ]
    if po:
        return [f"PO {po} on file for {tag}."]
    return []


def rewrite_evidence(wb, fa_by_tag: dict[str, str], po_map: dict[str, str]) -> int:
    """Fix Evidence Source money clauses using Source Ledger FA map."""
    cr = wb["Corrected Register"]
    n = 0
    false_absence = re.compile(
        r"no matching FAR|FAR silent|without a capital line|no FA row|no ledger|missing FA(?! expected)",
        re.I,
    )
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        ev = str(cr.cell(r, 11).value or "")
        fa = fa_by_tag.get(tag)
        po = po_map.get(tag) or gid(r"(PO-[\d-]+)", ev)

        # Strip bad money sentences
        sentences = [s.strip() for s in re.split(r"(?<=[.])\s+", ev) if s.strip()]
        kept = []
        for s in sentences:
            if false_absence.search(s) and tag not in MISSING_FA:
                continue
            if tag in UNDER_THRESHOLD and re.search(r"maps to ledger|FAR row|Cost path|backs", s, re.I):
                continue
            if tag == "MD-00132" and re.search(r"maps to ledger FA-|FAR row FA-", s):
                continue
            kept.append(s)

        # Remove trailing extras that duplicate money story; rebuild money clause
        money = money_clauses(tag, po, fa)
        clause = money[slot(tag, po or "", fa or "", r, mod=len(money))] if money else ""

        # Drop RN/under-threshold extras that conflict if wrongly present on FA assets
        extras_keep = []
        for s in kept:
            if tag not in MISSING_FA and re.search(r"RN-0132|under-threshold claim|below \$2,500", s, re.I):
                if "duplicate serial" in s.lower() or "MISMATCH" in s or "receiving_exception" in s:
                    extras_keep.append(s)
                continue
            extras_keep.append(s)

        # Ensure money clause present once
        body = " ".join(extras_keep)
        if clause and clause.rstrip(".") not in body:
            # Insert money clause before "Also:" extras if any
            if " Also:" in body or " Note -" in body or " Note —" in body:
                body = body.replace(" Also:", f" {clause} Also:", 1)
                if clause not in body:
                    body = body.rstrip(".") + ". " + clause
            else:
                body = (body.rstrip(".") + ". " + clause).strip()

        if tag == "MD-00132" and "RN-0132" not in body:
            body = body.rstrip(".") + ". RN-0132 under-threshold claim rejected; $6,470 exceeds capitalization gate with no FA row."
        if tag in UNDER_THRESHOLD and "missing FA expected" not in body.lower() and "FAR absence expected" not in body and "no FAR row expected" not in body and "capitalization not required" not in body:
            body = body.rstrip(".") + f". ${cr.cell(r, 13).value} below $2,500 - missing FA expected."

        body = body.replace("—", "-").replace("–", "-")
        body = re.sub(r"\s+", " ", body).strip()
        cr.cell(r, 11).value = body
        n += 1
    return n


def materialize_ledger_asset_ids(wb, fa_by_tag: dict[str, str]) -> int:
    """Replace Column O formulas with cached FA IDs (or blank) so they open correctly."""
    cr = wb["Corrected Register"]
    n = 0
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        fa = fa_by_tag.get(str(tag), "")
        # Keep formula-compatible by writing the value directly (readable without Excel).
        cr.cell(r, 15).value = fa if fa else None
        n += 1
    return n


def fix_ledger_reconciliation(wb) -> None:
    lr = wb["Ledger Reconciliation"]
    cr = wb["Corrected Register"]
    costs = {}
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if tag:
            costs[str(tag)] = float(cr.cell(r, 13).value or 0)
    c130 = costs.get("MD-00130", 700)
    c131 = costs.get("MD-00131", 1175)
    c132 = costs.get("MD-00132", 6470)
    combined = c130 + c131 + c132

    lr["A17"] = (
        f"1) Three assets (MD-00130, MD-00131, MD-00132) appear on the operational register and purchase orders "
        f"but have no fixed-asset ledger row. Combined register acquisition cost = ${combined:,.0f}. "
        f"MD-00130 (${c130:,.0f}) and MD-00131 (${c131:,.0f}) are under the $2,500 capitalization gate "
        f"(expected FA absence). MD-00132 (${c132:,.0f}) is capital-qualifying and is a Critical gap."
    )
    lr["A18"] = (
        "2) Seven assets have ledger-over-register acquisition cost variances "
        "(MD-00017 $85, MD-00034 $140, MD-00051 $210, MD-00068 $55, MD-00085 $195, "
        "MD-00102 $120, MD-00119 $165). Combined ledger cost excess = $970. "
        "PO approved amounts tie to the inventory lines; Finance should align the FAR. "
        "Per §9 these are High remediation items, not sole Critical blockers."
    )
    lr["A19"] = (
        "3) Eight disposal certificates clear operational status to Disposed with $0 NBV on ledger. "
        "Two recycler transfers (MD-00114, MD-00118) lack verified certificates; ledger already shows "
        "Disposed/$0 while operational status remains Retired/Pending Disposal. Active ledger NBV otherwise "
        "ties to corrected register remaining book value for assets present on both files."
    )
    # Clear old A20 duplicate if present
    if lr["A20"].value and str(lr["A20"].value).startswith("4)"):
        lr["A20"] = None
    print("LR A17-A19 rewritten as static 1/2/3 explanations")


def fix_purchase_orders() -> None:
    path = ROOT / "hardware_purchase_orders.csv"
    rows = load_csv("hardware_purchase_orders.csv")
    # Build asset cost map from inventory workbook corrected register if available
    wb = load_workbook(WORKBOOK, data_only=True)
    cr = wb["Corrected Register"]
    costs = {}
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if tag:
            costs[str(tag)] = float(cr.cell(r, 13).value or 0)

    new_rows = []
    for row in rows:
        tags = [t.strip() for t in re.split(r"[;]", row["asset_tags"]) if t.strip()]
        if row["purchase_order"] == "PO-2023-0022":
            capital_tags = [t for t in tags if t not in UNDER_THRESHOLD]
            under_tags = [t for t in tags if t in UNDER_THRESHOLD]
            cap_amt = sum(costs.get(t, 0) for t in capital_tags)
            under_amt = sum(costs.get(t, 0) for t in under_tags)
            row = dict(row)
            row["asset_tags"] = "; ".join(capital_tags)
            row["approved_amount"] = str(int(cap_amt)) if cap_amt == int(cap_amt) else str(cap_amt)
            row["capitalization_approved"] = "Yes"
            new_rows.append(row)
            if under_tags:
                new_rows.append(
                    {
                        "purchase_order": "PO-2023-0023",
                        "vendor": row.get("vendor", "CDW"),
                        "order_date": row.get("order_date", "2023-10-15"),
                        "asset_tags": "; ".join(under_tags),
                        "approved_amount": str(int(under_amt)) if under_amt == int(under_amt) else str(under_amt),
                        "capitalization_approved": "No",
                        "approver_role": "IT Procurement Manager",
                    }
                )
            continue

        # Spot-check: all assets under $2500 -> No
        tag_costs = [costs.get(t, 9999) for t in tags]
        if tag_costs and all(c < 2500 for c in tag_costs):
            row = dict(row)
            row["capitalization_approved"] = "No"
        new_rows.append(row)

    fields = list(rows[0].keys())
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(new_rows)
    yes = sum(1 for r in new_rows if r["capitalization_approved"] == "Yes")
    no = sum(1 for r in new_rows if r["capitalization_approved"] == "No")
    print(f"POs rewritten: {len(new_rows)} rows (Yes={yes}, No={no})")


def strip_em_dashes(wb) -> int:
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and ("—" in cell.value or "–" in cell.value):
                    cell.value = cell.value.replace("—", "-").replace("–", "-")
                    n += 1
    print("cells with em/en dashes rewritten:", n)
    return n


def replace_header_fills(wb) -> int:
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row or 1), max_col=min(30, ws.max_column or 1)):
            for cell in row:
                fill = cell.fill
                if not fill or not fill.fgColor:
                    continue
                rgb = str(fill.fgColor.rgb or "")
                hex6 = rgb[-6:].upper() if len(rgb) >= 6 else ""
                if hex6 in {"1F4E79", "0B2E4F"} or rgb in OLD_NAVY:
                    cell.fill = HEADER_DARK_FILL if hex6 == "0B2E4F" else HEADER_LIGHT
                    changed += 1
    print("navy header cells replaced:", changed)
    return changed


def patch_xlsx_metadata(path: Path) -> None:
    tmp = path.with_suffix(".meta.tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                txt = data.decode("utf-8")
                if "<dc:creator" not in txt:
                    txt = txt.replace(
                        "</cp:coreProperties>",
                        f"<dc:creator xmlns:dc=\"http://purl.org/dc/elements/1.1/\">{CREATOR}</dc:creator></cp:coreProperties>",
                    )
                else:
                    txt = re.sub(
                        r"(<dc:creator(?:\s[^>]*)?>)[^<]*(</dc:creator>)",
                        rf"\1{CREATOR}\2",
                        txt,
                        count=1,
                    )
                data = txt.encode("utf-8")
            elif item.filename == "docProps/app.xml":
                txt = data.decode("utf-8")
                if "<Application>" in txt:
                    txt = re.sub(
                        r"<Application>[^<]*</Application>",
                        f"<Application>{APPLICATION}</Application>",
                        txt,
                        count=1,
                    )
                else:
                    txt = txt.replace(
                        "</Properties>",
                        f"<Application>{APPLICATION}</Application></Properties>",
                    )
                # Drop AppVersion that screams openpyxl if present
                txt = re.sub(r"<AppVersion>[^<]*</AppVersion>", "", txt)
                data = txt.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)
    print(f"patched metadata on {path.name}")


def rebuild_zips() -> None:
    inputs = [
        "it_asset_inventory.xlsx",
        "hr_employee_status.csv",
        "equipment_transfer_log.csv",
        "service_desk_offboarding.csv",
        "device_return_shipments.csv",
        "hardware_purchase_orders.csv",
        "fixed_asset_ledger.xlsx",
        "asset_disposal_records.csv",
        "regional_IT_notes.docx",
        "IT_asset_management_policy.pdf",
        "ITAM_control_matrix.png",
        "receiving_exception_scan_1ZMD00000082.png",
    ]
    for zip_name in ("Yanou_IT_Asset_Inputs.zip", "Meridian_IT_Asset_Inputs.zip"):
        with zipfile.ZipFile(ROOT / zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
            for name in inputs:
                zf.write(ROOT / name, arcname=name)
        print("rebuilt", zip_name)

    shutil.copy2(WORKBOOK, MERIDIAN)
    patch_xlsx_metadata(MERIDIAN)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(MERIDIAN, "Meridian_IT_Asset_Reconciliation.xlsx")
    print("rebuilt deliverable zips")


def verify() -> None:
    wb = load_workbook(WORKBOOK, data_only=True)
    wb_f = load_workbook(WORKBOOK, data_only=False)
    cr = wb["Corrected Register"]
    cr_f = wb_f["Corrected Register"]
    sl = wb["Source Ledger"]
    fa_by_tag = {}
    for r in range(5, sl.max_row + 1):
        if sl.cell(r, 2).value:
            fa_by_tag[str(sl.cell(r, 2).value)] = str(sl.cell(r, 1).value)

    false_abs = 0
    blank_o_present = 0
    o_ok = 0
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        ev = str(cr.cell(r, 11).value or "")
        o = cr.cell(r, 15).value
        if tag in fa_by_tag:
            if re.search(r"no matching FAR|FAR silent|without a capital line for", ev, re.I):
                false_abs += 1
            if not o:
                blank_o_present += 1
            else:
                o_ok += 1
        else:
            assert not o, tag

    lr = wb["Ledger Reconciliation"]
    print("false no-FA claims on present assets:", false_abs)
    print("O cached for present:", o_ok, "blank among present:", blank_o_present)
    print("LR A17:", str(lr["A17"].value)[:80])
    print("LR A18:", str(lr["A18"].value)[:80])
    print("LR A19:", str(lr["A19"].value)[:80])
    print("B5/B6/B8/B12/B13:", lr["B5"].value, lr["B6"].value, lr["B8"].value, lr["B12"].value, lr["B13"].value)
    assert "175" not in str(lr["A18"].value)
    assert "970" in str(lr["A18"].value)
    assert str(lr["A17"].value).startswith("1)")
    assert str(lr["A18"].value).startswith("2)")
    assert str(lr["A19"].value).startswith("3)")

    em = 0
    navy = 0
    for ws in wb_f.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    em += cell.value.count("—") + cell.value.count("–")
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb).upper()
                    if "1F4E79" in rgb or "0B2E4F" in rgb:
                        navy += 1
    print("em/en dashes:", em, "navy headers:", navy)

    with zipfile.ZipFile(WORKBOOK) as z:
        core = z.read("docProps/core.xml").decode()
        app = z.read("docProps/app.xml").decode()
    print("creator ok:", CREATOR in core, "openpyxl in app:", "openpyxl" in app.lower())

    pos = load_csv("hardware_purchase_orders.csv")
    print("PO count", len(pos), "No caps", sum(1 for p in pos if p["capitalization_approved"] == "No"))
    for p in pos:
        if p["purchase_order"] in ("PO-2023-0022", "PO-2023-0023"):
            print(p)


def main() -> None:
    fix_purchase_orders()
    # Patch input xlsx metadata early
    for name in ("it_asset_inventory.xlsx", "fixed_asset_ledger.xlsx"):
        patch_xlsx_metadata(ROOT / name)

    wb = load_workbook(WORKBOOK)
    fa_by_tag = source_ledger_map(wb)
    print(f"Source Ledger FA map: {len(fa_by_tag)} tags")
    po_map = po_by_tag()

    n_ev = rewrite_evidence(wb, fa_by_tag, po_map)
    print("rewrote evidence rows:", n_ev)
    n_o = materialize_ledger_asset_ids(wb, fa_by_tag)
    print("materialized Ledger Asset ID values:", n_o)
    fix_ledger_reconciliation(wb)
    strip_em_dashes(wb)
    replace_header_fills(wb)

    wb.save(WORKBOOK)
    patch_xlsx_metadata(WORKBOOK)
    rebuild_zips()
    verify()


if __name__ == "__main__":
    main()
