#!/usr/bin/env python3
"""Eval expansions: formulas (<80% typed), PO line items, inventory variance, LR fidelity."""

from __future__ import annotations

import csv
import re
import shutil
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"

# Broader model mix (uneven category distribution)
MODEL_MAP = {
    # keep some originals; remap by hash of tag for variety
    "Laptop": [
        "Lenovo ThinkPad T14",
        "MacBook Pro 14",
        "Dell Latitude 7440",
        "HP EliteBook 840 G10",
        "Microsoft Surface Laptop 5",
        "Lenovo ThinkPad X1 Carbon",
    ],
    "Monitor": [
        "Dell U2723QE",
        "LG UltraFine 27MD5KL",
        "Samsung ViewFinity S8",
        "BenQ PD2705U",
        "Dell P2422H",
    ],
    "Mobile Device": [
        "Samsung Galaxy S24",
        "iPhone 15",
        "Google Pixel 8",
        "Samsung Galaxy S23 FE",
        "iPhone 14",
    ],
    "Network Asset": [
        "Cisco C9300",
        "Aruba 6300M",
        "Juniper EX3400",
        "Cisco Meraki MS250",
        "Fortinet FortiSwitch 148F",
    ],
}

# Uneven target category weights (approx counts out of 132)
CATEGORY_CYCLE = (
    ["Laptop"] * 40
    + ["Monitor"] * 28
    + ["Mobile Device"] * 36
    + ["Network Asset"] * 28
)


def slot(*parts, mod: int) -> int:
    h = 0
    for p in parts:
        for c in str(p):
            h = (h * 131 + ord(c)) % max(mod, 1)
    return h


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def capture_and_inject(path: Path, cache: dict[tuple[str, str], float | int]) -> None:
    """Write formula cached values into OOXML (handles empty <v/>)."""
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(path, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        sheets_by_path = {p: zin.read(p) for p in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def set_cell_cache(xml: bytes, coord: str, val: float | int) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        val_s = str(int(val)) if float(val).is_integer() else str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            body = re.sub(r"<v[^>]*/>", "", body)
            body = re.sub(r"<v>[^<]*</v>", "", body)
            body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for pth, xml in list(sheets_by_path.items()):
        sheet_name = path_to_name.get(pth)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name:
                sheets_by_path[pth] = set_cell_cache(sheets_by_path[pth], coord, val)

    tmp = path.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for pth, data in sheets_by_path.items():
            zout.writestr(pth, data)
    tmp.replace(path)
    print("injected formula cache:", patched, "of", len(cache))


def vary_inventory() -> dict[str, dict]:
    """Uneven categories + broader models; return tag -> inventory fields."""
    path = ROOT / "it_asset_inventory.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    headers = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    tags = []
    for r in range(5, ws.max_row + 1):
        tag = ws.cell(r, 1).value
        if tag:
            tags.append((r, tag))

    # Assign uneven categories by shuffled cycle keyed to tag order
    order = sorted(tags, key=lambda x: slot(x[1], mod=10007))
    cat_for_tag: dict[str, str] = {}
    for i, (_, tag) in enumerate(order):
        cat_for_tag[tag] = CATEGORY_CYCLE[i % len(CATEGORY_CYCLE)]

    out: dict[str, dict] = {}
    for r, tag in tags:
        cat = cat_for_tag[tag]
        models = MODEL_MAP[cat]
        model = models[slot(tag, cat, mod=len(models))]
        ws.cell(r, headers["Category"]).value = cat
        ws.cell(r, headers["Model"]).value = model
        out[tag] = {
            "serial": ws.cell(r, headers["Serial Number"]).value,
            "category": cat,
            "model": model,
            "custodian": ws.cell(r, headers["Assigned Employee"]).value,
            "location": ws.cell(r, headers["Location"]).value,
            "status": ws.cell(r, headers["Lifecycle Status"]).value,
            "cost": num(ws.cell(r, headers["Acquisition Cost"]).value),
            "purchase": ws.cell(r, headers["Purchase Date"]).value,
            "warranty": ws.cell(r, headers["Warranty Expiration"]).value,
            "source": ws.cell(r, headers["Source Record"]).value,
        }
    wb.save(path)
    print("inventory varied:", Counter(v["category"] for v in out.values()))
    print("models:", len({v["model"] for v in out.values()}))
    return out


def expand_po_line_items(inv: dict[str, dict]) -> None:
    """Rewrite hardware_purchase_orders.csv as one row per asset with uneven PO batches."""
    path = ROOT / "hardware_purchase_orders.csv"
    # If already line-item format with asset_tag, rebuild batches from current tags
    raw = list(csv.DictReader(path.open(newline="")))
    if raw and "asset_tag" in raw[0]:
        tags = []
        seen = set()
        for r in raw:
            t = r["asset_tag"].strip()
            if t not in seen:
                seen.add(t)
                tags.append(t)
        vendors_dates = []
        seen_po = set()
        for r in raw:
            if r["purchase_order"] not in seen_po:
                seen_po.add(r["purchase_order"])
                vendors_dates.append(
                    (r["purchase_order"], r["vendor"], r["order_date"], r["approver_role"], r["capitalization_approved"])
                )
    else:
        tags = []
        vendors_dates = []
        for r in raw:
            vendors_dates.append(
                (r["purchase_order"], r["vendor"], r["order_date"], r["approver_role"], r["capitalization_approved"])
            )
            for t in re.split(r"[;]", r["asset_tags"]):
                t = t.strip()
                if t:
                    tags.append(t)

    under = [t for t in tags if t in ("MD-00130", "MD-00131")]
    special132 = [t for t in tags if t == "MD-00132"]
    other = [t for t in tags if t not in ("MD-00130", "MD-00131", "MD-00132")]

    pattern = [3, 5, 8, 4, 7, 6, 2, 9, 5, 3, 8, 4, 6, 7, 5, 3, 4, 8, 6]
    sizes: list[int] = []
    total = 0
    i = 0
    while total < len(other):
        n = min(pattern[i % len(pattern)], len(other) - total)
        sizes.append(n)
        total += n
        i += 1

    batches: list[list[str]] = []
    idx = 0
    for n in sizes:
        batches.append(other[idx : idx + n])
        idx += n
    batches.append(special132)
    batches.append(under)

    templates = [
        ("PO-2022-0001", "CDW", "2022-01-03", "IT Procurement Manager"),
        ("PO-2023-0002", "SHI", "2023-02-07", "Hardware Buying Lead"),
        ("PO-2024-0003", "Insight", "2024-03-09", "Director, Procurement"),
        ("PO-2025-0004", "Connection", "2025-04-12", "IT Procurement Manager"),
        ("PO-2022-0005", "Zones", "2022-05-14", "P2P Approver"),
        ("PO-2023-0006", "CDW", "2023-06-18", "Category Manager, Client Devices"),
        ("PO-2024-0007", "Insight", "2024-07-21", "IT Procurement Manager"),
        ("PO-2025-0008", "SHI", "2025-08-22", "Hardware Buying Lead"),
        ("PO-2022-0009", "PCM", "2022-09-26", "Director, Procurement"),
        ("PO-2023-0010", "CDW", "2023-10-27", "IT Procurement Manager"),
        ("PO-2024-0011", "Zones", "2024-11-08", "P2P Approver"),
        ("PO-2025-0012", "Insight", "2025-12-16", "Category Manager, Client Devices"),
        ("PO-2022-0013", "SHI", "2022-01-04", "IT Procurement Manager"),
        ("PO-2023-0014", "Connection", "2023-02-11", "Hardware Buying Lead"),
        ("PO-2024-0015", "CDW", "2024-03-19", "Sourcing Lead"),
        ("PO-2025-0016", "Insight", "2025-04-23", "Director, Procurement"),
        ("PO-2022-0017", "SHI", "2022-05-06", "IT Procurement Manager"),
        ("PO-2023-0018", "Zones", "2023-06-13", "P2P Approver"),
        ("PO-2024-0019", "CDW", "2024-07-17", "Category Manager, Client Devices"),
        ("PO-2025-0020", "Insight", "2025-08-28", "Hardware Buying Lead"),
        ("PO-2022-0021", "Connection", "2022-09-05", "IT Procurement Manager"),
        ("PO-2024-0024", "SHI", "2024-11-19", "Sourcing Lead"),
        ("PO-2025-0025", "Zones", "2025-01-22", "P2P Approver"),
        ("PO-2022-0026", "Insight", "2022-03-08", "IT Procurement Manager"),
        ("PO-2023-0027", "CDW", "2023-05-16", "Hardware Buying Lead"),
    ]

    lines: list[dict] = []
    pi = 0
    for batch in batches:
        if set(batch) <= {"MD-00130", "MD-00131"}:
            po, vendor, date, apr, cap = "PO-2023-0023", "CDW", "2023-10-15", "IT Procurement Manager", "No"
        elif batch == special132:
            po, vendor, date, apr, cap = "PO-2023-0022", "CDW", "2023-10-15", "Director, Procurement", "Yes"
        else:
            po, vendor, date, apr = templates[pi]
            cap = "Yes"
            pi += 1
        amounts = [inv[t]["cost"] for t in batch]
        po_total = str(int(round(sum(amounts))))
        mates = "; ".join(batch)
        for t, amt in zip(batch, amounts):
            lines.append(
                {
                    "purchase_order": po,
                    "vendor": vendor,
                    "order_date": date,
                    "asset_tag": t,
                    "line_amount": str(int(round(amt))),
                    "po_total": po_total,
                    "capitalization_approved": cap,
                    "approver_role": apr,
                    "asset_tags": mates,
                }
            )

    fieldnames = [
        "purchase_order",
        "vendor",
        "order_date",
        "asset_tag",
        "line_amount",
        "po_total",
        "capitalization_approved",
        "approver_role",
        "asset_tags",
    ]
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(lines)
    sizes = sorted(Counter(r["purchase_order"] for r in lines).values())
    print("PO line items:", len(lines), "POs", len(set(r["purchase_order"] for r in lines)), "sizes", sizes)


def sync_inventory_models_to_gold(inv: dict[str, dict]) -> None:
    """Update Corrected Register category/model typed values before formula conversion."""
    wb = load_workbook(WORKBOOK)
    cr = wb["Corrected Register"]
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if tag in inv:
            cr.cell(r, 3).value = inv[tag]["category"]
            cr.cell(r, 4).value = inv[tag]["model"]
    wb.save(WORKBOOK)


def add_source_inventory_sheet(wb, inv: dict[str, dict]) -> None:
    if "Source Inventory" in wb.sheetnames:
        del wb["Source Inventory"]
    # Place after Source Ledger
    idx = wb.sheetnames.index("Source Ledger") + 1 if "Source Ledger" in wb.sheetnames else 0
    ws = wb.create_sheet("Source Inventory", idx)
    ws["A1"] = "Source Inventory (from it_asset_inventory.xlsx)"
    ws["A2"] = "Operational snapshot used by Corrected Register lookups."
    headers = [
        "Asset Tag",
        "Serial Number",
        "Category",
        "Model",
        "Assigned Employee",
        "Location",
        "Lifecycle Status",
        "Acquisition Cost",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True, color="FFF8F0")
        cell.fill = PatternFill("solid", fgColor="3B2F2A")
    for i, tag in enumerate(sorted(inv.keys())):
        r = 5 + i
        a = inv[tag]
        ws.cell(r, 1, tag)
        ws.cell(r, 2, a["serial"])
        ws.cell(r, 3, a["category"])
        ws.cell(r, 4, a["model"])
        ws.cell(r, 5, a["custodian"])
        ws.cell(r, 6, a["location"])
        ws.cell(r, 7, a["status"])
        ws.cell(r, 8, a["cost"])
    print("Source Inventory rows:", len(inv))


def apply_formulas(inv: dict[str, dict]) -> None:
    wb = load_workbook(WORKBOOK)
    wbv = load_workbook(WORKBOOK, data_only=True)
    add_source_inventory_sheet(wb, inv)

    cr = wb["Corrected Register"]
    crv = wbv["Corrected Register"]
    sl = wb["Source Ledger"]
    slv = wbv["Source Ledger"]
    er = wb["Exception Register"]
    erv = wbv["Exception Register"]
    lr = wb["Ledger Reconciliation"]
    dash = wb["Dashboard"]
    dashv = wbv["Dashboard"]

    # Build ledger map for cache
    ledger: dict[str, dict] = {}
    for r in range(5, slv.max_row + 1):
        tag = slv.cell(r, 2).value
        if not tag:
            continue
        ledger[str(tag)] = {
            "fa": slv.cell(r, 1).value,
            "cap": slv.cell(r, 4).value,
            "acq": num(slv.cell(r, 5).value),
            "dep": num(slv.cell(r, 6).value),
            "nbv": num(slv.cell(r, 7).value),
            "status": slv.cell(r, 8).value,
        }

    cache: dict[tuple[str, str], float | int] = {}
    inv_range = "'Source Inventory'!$A$5:$H$200"
    led_range_tag = "'Source Ledger'!$B$5:$B$200"

    # Source Ledger NBV = Acq − Accum Dep (keep typed when printed NBV disagrees, e.g. over-dep)
    for r in range(5, sl.max_row + 1):
        if not sl.cell(r, 1).value:
            continue
        acq = num(slv.cell(r, 5).value) or 0
        dep = num(slv.cell(r, 6).value) or 0
        printed = num(slv.cell(r, 7).value)
        if printed is not None and abs((acq - dep) - printed) > 0.01:
            sl.cell(r, 7).value = printed
        else:
            sl.cell(r, 7).value = f"=E{r}-F{r}"
            cache[("Source Ledger", f"G{r}")] = printed if printed is not None else round(acq - dep, 2)

    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        a = inv.get(str(tag), {})
        ld = ledger.get(str(tag), {})

        # Inventory lookups (B–G, M)
        cr.cell(r, 2).value = f'=IFERROR(XLOOKUP(A{r},{inv_range.replace("$A$5:$H$200","$A$5:$A$200")},\'Source Inventory\'!$B$5:$B$200),"")'
        cr.cell(r, 3).value = f'=IFERROR(XLOOKUP(A{r},\'Source Inventory\'!$A$5:$A$200,\'Source Inventory\'!$C$5:$C$200),"")'
        cr.cell(r, 4).value = f'=IFERROR(XLOOKUP(A{r},\'Source Inventory\'!$A$5:$A$200,\'Source Inventory\'!$D$5:$D$200),"")'
        cr.cell(r, 5).value = f'=IFERROR(XLOOKUP(A{r},\'Source Inventory\'!$A$5:$A$200,\'Source Inventory\'!$E$5:$E$200),"")'
        cr.cell(r, 6).value = f'=IFERROR(XLOOKUP(A{r},\'Source Inventory\'!$A$5:$A$200,\'Source Inventory\'!$F$5:$F$200),"")'
        cr.cell(r, 7).value = f'=IFERROR(XLOOKUP(A{r},\'Source Inventory\'!$A$5:$A$200,\'Source Inventory\'!$G$5:$G$200),"")'
        cr.cell(r, 13).value = f'=IFERROR(XLOOKUP(A{r},\'Source Inventory\'!$A$5:$A$200,\'Source Inventory\'!$H$5:$H$200),"")'

        cache[("Corrected Register", f"B{r}")] = a.get("serial") or ""
        cache[("Corrected Register", f"C{r}")] = a.get("category") or ""
        cache[("Corrected Register", f"D{r}")] = a.get("model") or ""
        cache[("Corrected Register", f"E{r}")] = a.get("custodian") or ""
        cache[("Corrected Register", f"F{r}")] = a.get("location") or ""
        cache[("Corrected Register", f"G{r}")] = a.get("status") or ""
        cost = a.get("cost")
        if cost is not None:
            cache[("Corrected Register", f"M{r}")] = cost

        # Ledger lookups
        cr.cell(r, 15).value = f'=IFERROR(XLOOKUP(A{r},\'Source Ledger\'!$B$5:$B$200,\'Source Ledger\'!$A$5:$A$200),"")'
        cr.cell(r, 16).value = f'=IFERROR(XLOOKUP(A{r},\'Source Ledger\'!$B$5:$B$200,\'Source Ledger\'!$D$5:$D$200),"")'
        cr.cell(r, 20).value = f'=IFERROR(XLOOKUP(A{r},\'Source Ledger\'!$B$5:$B$200,\'Source Ledger\'!$E$5:$E$200),"")'
        cr.cell(r, 21).value = f'=IFERROR(XLOOKUP(A{r},\'Source Ledger\'!$B$5:$B$200,\'Source Ledger\'!$H$5:$H$200),"")'
        cr.cell(r, 22).value = f'=IFERROR(XLOOKUP(A{r},\'Source Ledger\'!$B$5:$B$200,\'Source Ledger\'!$G$5:$G$200),"")'
        cr.cell(r, 14).value = f'=IF(V{r}="","",V{r})'
        cr.cell(r, 23).value = f'=IF(OR(M{r}="",T{r}=""),"",T{r}-M{r})'
        cr.cell(r, 24).value = f'=IF(V{r}="","No","Yes")'
        cr.cell(r, 19).value = f'=IF(OR(R{r}="",R{r}=0),"No","Yes")'

        if ld:
            cache[("Corrected Register", f"O{r}")] = ld["fa"] or ""
            cache[("Corrected Register", f"P{r}")] = ld["cap"] or ""
            if ld["acq"] is not None:
                cache[("Corrected Register", f"T{r}")] = ld["acq"]
            cache[("Corrected Register", f"U{r}")] = ld["status"] or ""
            if ld["nbv"] is not None:
                cache[("Corrected Register", f"V{r}")] = ld["nbv"]
                cache[("Corrected Register", f"N{r}")] = ld["nbv"]
            if cost is not None and ld["acq"] is not None:
                cache[("Corrected Register", f"W{r}")] = round(ld["acq"] - cost, 2)
            cache[("Corrected Register", f"X{r}")] = "Yes"
        else:
            cache[("Corrected Register", f"O{r}")] = ""
            cache[("Corrected Register", f"P{r}")] = ""
            cache[("Corrected Register", f"T{r}")] = ""
            cache[("Corrected Register", f"U{r}")] = ""
            cache[("Corrected Register", f"V{r}")] = ""
            cache[("Corrected Register", f"N{r}")] = ""
            cache[("Corrected Register", f"W{r}")] = ""
            cache[("Corrected Register", f"X{r}")] = "No"

        ex = crv.cell(r, 18).value or cr.cell(r, 18).value
        cache[("Corrected Register", f"S{r}")] = "Yes" if ex else "No"

    # Exception Register financial exposure formulas
    for r in range(5, er.max_row + 1):
        if not er.cell(r, 1).value:
            continue
        typ = er.cell(r, 2).value
        exp = erv.cell(r, 7).value
        er.cell(r, 7).value = (
            f'=IF(B{r}="Inventory-to-Ledger Difference",'
            f'IF(IFERROR(XLOOKUP(D{r},\'Corrected Register\'!$A$5:$A$200,\'Corrected Register\'!$T$5:$T$200),\"\")=\"\",'
            f'IFERROR(XLOOKUP(D{r},\'Corrected Register\'!$A$5:$A$200,\'Corrected Register\'!$M$5:$M$200),0),'
            f'ABS(IFERROR(XLOOKUP(D{r},\'Corrected Register\'!$A$5:$A$200,\'Corrected Register\'!$T$5:$T$200),0)'
            f'-IFERROR(XLOOKUP(D{r},\'Corrected Register\'!$A$5:$A$200,\'Corrected Register\'!$M$5:$M$200),0))),'
            f'IFERROR(XLOOKUP(D{r},\'Corrected Register\'!$A$5:$A$200,\'Corrected Register\'!$N$5:$N$200),0))'
        )
        if exp is not None:
            cache[("Exception Register", f"G{r}")] = float(exp)

    # Ledger Reconciliation asset-level VLOOKUPs
    for r in range(24, 36):
        if not lr.cell(r, 1).value:
            continue
        tag = lr.cell(r, 1).value
        # C = register cost, D = ledger cost, E = diff formula, F = NBV, G = ledger NBV
        lr.cell(r, 3).value = f"=IFERROR(XLOOKUP(A{r},'Corrected Register'!$A$5:$A$200,'Corrected Register'!$M$5:$M$200),\"\")"
        lr.cell(r, 4).value = f"=IFERROR(XLOOKUP(A{r},'Corrected Register'!$A$5:$A$200,'Corrected Register'!$T$5:$T$200),\"\")"
        lr.cell(r, 5).value = f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'
        lr.cell(r, 6).value = f"=IFERROR(XLOOKUP(A{r},'Corrected Register'!$A$5:$A$200,'Corrected Register'!$N$5:$N$200),\"\")"
        lr.cell(r, 7).value = f"=IFERROR(XLOOKUP(A{r},'Source Ledger'!$B$5:$B$200,'Source Ledger'!$G$5:$G$200),\"\")"
        a = inv.get(str(tag), {})
        ld = ledger.get(str(tag), {})
        if a.get("cost") is not None:
            cache[("Ledger Reconciliation", f"C{r}")] = a["cost"]
        if ld.get("acq") is not None:
            cache[("Ledger Reconciliation", f"D{r}")] = ld["acq"]
            if a.get("cost") is not None:
                cache[("Ledger Reconciliation", f"E{r}")] = round(ld["acq"] - a["cost"], 2)
        if ld.get("nbv") is not None:
            cache[("Ledger Reconciliation", f"F{r}")] = ld["nbv"]
            cache[("Ledger Reconciliation", f"G{r}")] = ld["nbv"]

    # Fix invented $514.88 narrative in LR explanations
    for r in range(16, 22):
        val = lr.cell(r, 1).value
        if val and "514.88" in str(val):
            lr.cell(r, 1).value = (
                "4) FA-000034 (MD-00034) NBV is $479.88 (acquisition $930.00 minus accumulated depreciation "
                "$450.12). Aggregate Active ledger NBV control total is $61,526.60."
            )
    for r in range(24, 36):
        val = lr.cell(r, 2).value
        if val and "514.88" in str(val):
            lr.cell(r, 2).value = "Acquisition cost mismatch; FA-000034 NBV $479.88 (acq $930 − accum dep $450.12)"

    # Dashboard KPIs (keep/ensure formulas + cache)
    rows = [r for r in range(5, cr.max_row + 1) if cr.cell(r, 1).value]
    sum_m = sum(float(inv[str(cr.cell(r, 1).value)]["cost"]) for r in rows if str(cr.cell(r, 1).value) in inv)
    sum_n = sum(
        float(ledger[str(cr.cell(r, 1).value)]["nbv"] or 0)
        for r in rows
        if str(cr.cell(r, 1).value) in ledger
        and ledger[str(cr.cell(r, 1).value)]["nbv"] is not None
    )
    ex_count = sum(1 for r in range(5, er.max_row + 1) if er.cell(r, 1).value)
    dash.cell(5, 2).value = "=COUNTA('Corrected Register'!A5:A200)"
    dash.cell(6, 2).value = "=SUM('Corrected Register'!M5:M200)"
    dash.cell(7, 2).value = "=SUM('Corrected Register'!N5:N200)"
    dash.cell(8, 2).value = '=COUNTIF(\'Exception Register\'!K5:K200,"Open")'
    cache[("Dashboard", "B5")] = len(rows)
    cache[("Dashboard", "B6")] = round(sum_m, 2)
    cache[("Dashboard", "B7")] = round(sum_n, 2)
    # Prefer existing open count from data_only
    open_ex = sum(1 for r in range(5, erv.max_row + 1) if erv.cell(r, 11).value == "Open")
    cache[("Dashboard", "B8")] = open_ex or ex_count

    # Capture remaining numeric dashboard/LR formula caches from prior values
    for sheet_name, ws, wsv in (
        ("Dashboard", dash, dashv),
        ("Ledger Reconciliation", lr, wbv["Ledger Reconciliation"]),
    ):
        for row in wsv.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    formula_cell = ws[cell.coordinate]
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                        cache.setdefault((sheet_name, cell.coordinate), float(cell.value))

    wb.save(WORKBOOK)

    # String caches need special handling — openpyxl data_only won't help for text formula results.
    # Materialize text formula results by also writing shared-string style via a second pass:
    # For text XLOOKUP results, inject as cached shared strings is complex; instead convert
    # text-valued formula cells to keep formula but also set a parallel typed approach:
    # Use openpyxl's formula with cached value via cell.value trick — not supported.
    # Inject numeric only; for text, replace formulas with values for columns that are text
    # OR inject via OOXML with t="str" and <v>index</v> — too heavy.
    #
    # Pragmatic: keep inventory/ledger *numeric* and ID columns as formulas with numeric cache
    # where possible; for text columns use INDEX/MATCH returning numbers only...
    #
    # Better pragmatic path for text: leave formulas in place and also store typed values
    # by NOT using formulas for pure text — use formulas only for numeric derived fields
    # PLUS duplicate Source Inventory is still useful.
    #
    # Re-evaluate: evaluator may use Excel which calculates. data_only readers need cache.
    # For text XLOOKUP, inject t="str" inline string in OOXML (Excel allows <is><t>).

    inject_text_and_numeric_cache(WORKBOOK, cache, inv, ledger, rows, open_ex or ex_count, sum_m, sum_n)
    print("formulas applied; cache keys", len(cache))


def inject_text_and_numeric_cache(
    path: Path,
    cache: dict,
    inv: dict,
    ledger: dict,
    rows: list,
    open_ex: int,
    sum_m: float,
    sum_n: float,
) -> None:
    """Inject numeric caches; materialize text formula results as cached inline strings."""
    import xml.etree.ElementTree as ET

    # Expand cache with text results encoded as special tuples
    text_cache: dict[tuple[str, str], str] = {}
    wb = load_workbook(path)
    cr = wb["Corrected Register"]
    for r in rows:
        tag = str(cr.cell(r, 1).value)
        a = inv.get(tag, {})
        ld = ledger.get(tag, {})
        for col, val in (
            (2, a.get("serial")),
            (3, a.get("category")),
            (4, a.get("model")),
            (5, a.get("custodian")),
            (6, a.get("location")),
            (7, a.get("status")),
            (15, ld.get("fa")),
            (16, ld.get("cap")),
            (21, ld.get("status")),
            (19, "Yes" if (cr.cell(r, 18).value) else "No"),
            (24, "Yes" if ld else "No"),
        ):
            if val is not None and val != "":
                text_cache[("Corrected Register", f"{get_column_letter(col)}{r}")] = str(val)
            elif col in (15, 16, 21) and not ld:
                text_cache[("Corrected Register", f"{get_column_letter(col)}{r}")] = ""

    # Filter numeric cache to numbers only
    num_cache = {k: v for k, v in cache.items() if isinstance(v, (int, float))}

    with zipfile.ZipFile(path, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        files = {i.filename: zin.read(i.filename) for i in zin.infolist()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def patch_sheet(xml: bytes, sheet_name: str) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        # numeric
        for (sn, coord), val in num_cache.items():
            if sn != sheet_name:
                continue
            val_s = str(int(val)) if float(val).is_integer() else str(val)
            cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

            def repl(m, val_s=val_s):
                nonlocal patched
                head, body, tail = m.group(1), m.group(2), m.group(3)
                body = re.sub(r"<v[^>]*/>", "", body)
                body = re.sub(r"<v>[^<]*</v>", "", body)
                body = body + f"<v>{val_s}</v>"
                # ensure not typed as shared string
                head = re.sub(r'\s+t="s"', "", head)
                patched += 1
                return head + body + tail

            if re.search(cell_re, s, flags=re.DOTALL):
                s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)

        # text inline strings for formula cells
        for (sn, coord), text in text_cache.items():
            if sn != sheet_name:
                continue
            esc = (
                text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )
            cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

            def repl_t(m, esc=esc):
                nonlocal patched
                head, body, tail = m.group(1), m.group(2), m.group(3)
                # keep <f>...</f>, replace value with inline string
                f_match = re.search(r"<f[^>]*>.*?</f>|<f[^>]*/>", body)
                f_xml = f_match.group(0) if f_match else ""
                if 't="inlineStr"' not in head:
                    head = head[:-1] + ' t="inlineStr">'
                body = f_xml + f"<is><t>{esc}</t></is>"
                patched += 1
                return head + body + tail

            if re.search(cell_re, s, flags=re.DOTALL):
                s = re.sub(cell_re, repl_t, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for pth, data in list(files.items()):
        sn = path_to_name.get(pth)
        if sn:
            files[pth] = patch_sheet(data, sn)

    tmp = path.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    tmp.replace(path)
    print("cache inject patched cells:", patched)


def rebuild_zips() -> None:
    inputs = [
        "it_asset_inventory.xlsx",
        "hr_employee_status.csv",
        "equipment_transfer_log.csv",
        "service_desk_offboarding.csv",
        "device_return_shipments.csv",
        "hardware_purchase_orders.csv",
        "fixed_asset_ledger.xlsx",
        "asset_disposal_records.csv",
        "regional_IT_notes.docx",
        "IT_asset_management_policy.pdf",
        "ITAM_control_matrix.png",
        "receiving_exception_scan_1ZMD00000082.png",
    ]
    for zip_name in ("Yanou_IT_Asset_Inputs.zip",):
        with zipfile.ZipFile(ROOT / zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
            for name in inputs:
                zf.write(ROOT / name, arcname=name)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    print("zips rebuilt")


def measure() -> None:
    wb = load_workbook(WORKBOOK)
    formula = typed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula += 1
                else:
                    typed += 1
    total = formula + typed
    print(f"typed ratio: {typed}/{total} = {100*typed/total:.1f}%  formulas={formula}")
    wbv = load_workbook(WORKBOOK, data_only=True)
    d = wbv["Dashboard"]
    print("Dashboard B5-B8:", [d.cell(r, 2).value for r in range(5, 9)])
    cr = wbv["Corrected Register"]
    print("MD-00001 model:", cr.cell(5, 4).value, "cost:", cr.cell(5, 13).value, "N:", cr.cell(5, 14).value)


def update_po_parser_compat() -> None:
    """Ensure fix_authorship_natural can read line-item PO file."""
    path = ROOT / "fix_authorship_natural.py"
    text = path.read_text()
    old = '''    po_by_tag: dict[str, str] = {}
    for prow in load_csv("hardware_purchase_orders.csv"):
        for t in re.split(r"[;\\s]+", prow["asset_tags"]):
            t = t.strip()
            if t.startswith("MD-"):
                po_by_tag[t] = prow["purchase_order"]'''
    new = '''    po_by_tag: dict[str, str] = {}
    for prow in load_csv("hardware_purchase_orders.csv"):
        if prow.get("asset_tag") and str(prow["asset_tag"]).startswith("MD-"):
            po_by_tag[str(prow["asset_tag"]).strip()] = prow["purchase_order"]
            continue
        for t in re.split(r"[;\\s]+", prow.get("asset_tags") or ""):
            t = t.strip()
            if t.startswith("MD-"):
                po_by_tag[t] = prow["purchase_order"]'''
    if old in text:
        path.write_text(text.replace(old, new))
        print("updated PO parser in fix_authorship_natural.py")
    else:
        print("PO parser already compatible or pattern mismatch")


def main() -> None:
    inv = vary_inventory()
    expand_po_line_items(inv)
    update_po_parser_compat()
    sync_inventory_models_to_gold(inv)
    # Refresh evidence narratives that mention models
    import fix_authorship_natural as fan

    fan.rewrite_workbook(WORKBOOK)
    apply_formulas(inv)
    rebuild_zips()
    measure()


if __name__ == "__main__":
    main()
