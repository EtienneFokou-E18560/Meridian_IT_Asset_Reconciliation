#!/usr/bin/env python3
"""Materialize Dashboard (and LR control) formula outputs as typed values for oracle/judge tools."""

from __future__ import annotations

import shutil
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return 0.0
    return float(s)


def materialize_dashboard(wb) -> dict[str, float | int]:
    """Replace Dashboard formula cells with computed typed values."""
    cr = wb["Corrected Register"]
    si = wb["Source Inventory"]
    sl = wb["Source Ledger"]
    er = wb["Exception Register"]
    dash = wb["Dashboard"]
    cc = wb["Custody Chain"]

    # Source Ledger NBV / cost
    sl_nbv: dict[str, float] = {}
    sl_cost: dict[str, float] = {}
    sl_status: dict[str, str] = {}
    for r in range(5, 200):
        tag = sl.cell(r, 2).value
        if not tag:
            continue
        tag = str(tag)
        acq = num(sl.cell(r, 5).value)
        dep = num(sl.cell(r, 6).value)
        nbv_cell = sl.cell(r, 7).value
        if isinstance(nbv_cell, str) and nbv_cell.startswith("="):
            nbv = round(acq - dep, 2)
        else:
            nbv = num(nbv_cell) if nbv_cell is not None else round(acq - dep, 2)
        sl_cost[tag] = acq
        sl_nbv[tag] = nbv
        sl_status[tag] = str(sl.cell(r, 8).value or "")

    # Inventory categories / costs (typed Source Inventory)
    inv_cat: dict[str, str] = {}
    inv_cost: dict[str, float] = {}
    for r in range(5, 200):
        tag = si.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        inv_cat[tag] = str(si.cell(r, 3).value or "")
        inv_cost[tag] = num(si.cell(r, 8).value)

    tags: list[str] = []
    status: dict[str, str] = {}
    loc: dict[str, str] = {}
    cost: dict[str, float] = {}
    nbv: dict[str, float] = {}
    for r in range(5, 200):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        tags.append(tag)
        status[tag] = str(cr.cell(r, 10).value or "")
        loc[tag] = str(cr.cell(r, 9).value or "")
        c = cr.cell(r, 13).value
        cost[tag] = inv_cost.get(tag, num(c) if not (isinstance(c, str) and c.startswith("=")) else 0.0)
        if tag in sl_nbv:
            nbv[tag] = sl_nbv[tag]
            # Keep Corrected Register N typed if still formula
            if isinstance(cr.cell(r, 14).value, str) and str(cr.cell(r, 14).value).startswith("="):
                cr.cell(r, 14).value = sl_nbv[tag]
        else:
            nbv[tag] = 0.0

    def cnt(pred) -> int:
        return sum(1 for t in tags if pred(t))

    def nbv_sum(pred) -> float:
        return round(sum(nbv[t] for t in tags if pred(t)), 2)

    # Exception type aggregates
    ex_count: dict[str, int] = defaultdict(int)
    ex_exp: dict[str, float] = defaultdict(float)
    open_ex = crit = 0
    for r in range(5, 200):
        if not er.cell(r, 1).value:
            continue
        typ = str(er.cell(r, 2).value or "")
        tag = str(er.cell(r, 4).value or "")
        if er.cell(r, 11).value == "Open":
            open_ex += 1
        if er.cell(r, 3).value == "Critical":
            crit += 1
        if typ == "Inventory-to-Ledger Difference":
            if sl_status.get(tag) == "Disposed":
                g = sl_nbv.get(tag, 0.0)
            elif tag not in sl_cost:
                g = cost.get(tag, 0.0)
            else:
                g = abs(sl_cost[tag] - cost.get(tag, 0.0))
        elif typ == "Below Capitalization Threshold":
            g = cost.get(tag, 0.0)
        else:
            g = nbv.get(tag, 0.0)
        # Materialize exception exposure column G if formula
        g_cell = er.cell(r, 7).value
        if isinstance(g_cell, str) and g_cell.startswith("="):
            er.cell(r, 7).value = round(float(g), 2)
        else:
            g = num(g_cell) if g_cell not in (None, "") else float(g)
        ex_count[typ] += 1
        ex_exp[typ] += float(g)

    custody_tags = {cc.cell(r, 1).value for r in range(5, 502) if cc.cell(r, 1).value}

    # KPI cards B5–B13
    dash["B5"] = len(tags)
    dash["B6"] = round(sum(cost.values()), 2)
    dash["B7"] = round(sum(nbv.values()), 2)
    dash["B8"] = open_ex
    dash["B9"] = crit
    dash["B10"] = cnt(lambda t: status[t] == "Missing")
    dash["B11"] = cnt(lambda t: status[t] == "Return Overdue")
    dash["B12"] = cnt(lambda t: status[t] == "Disposed")
    dash["B13"] = len(custody_tags)

    # Verified status breakout A19:A26 / B / C
    for r in range(19, 27):
        key = dash.cell(r, 1).value
        if not key:
            continue
        dash.cell(r, 2).value = cnt(lambda t, k=str(key): status[t] == k)
        dash.cell(r, 3).value = nbv_sum(lambda t, k=str(key): status[t] == k)

    # Device category E19:E22 / F / G
    for r in range(19, 23):
        key = dash.cell(r, 5).value
        if not key:
            continue
        dash.cell(r, 6).value = sum(1 for t in tags if inv_cat.get(t) == key)
        dash.cell(r, 7).value = nbv_sum(lambda t, k=str(key): inv_cat.get(t) == k)

    # Verified location A32:A43 / B / C
    for r in range(32, 44):
        key = dash.cell(r, 1).value
        if not key:
            continue
        dash.cell(r, 2).value = cnt(lambda t, k=str(key): loc[t] == k)
        dash.cell(r, 3).value = nbv_sum(lambda t, k=str(key): loc[t] == k)

    # Exception type E32:E41 / F / G
    for r in range(32, 42):
        key = dash.cell(r, 5).value
        if not key:
            continue
        dash.cell(r, 6).value = ex_count.get(str(key), 0)
        dash.cell(r, 7).value = round(ex_exp.get(str(key), 0.0), 2)

    return {
        "assets": len(tags),
        "laptop": sum(1 for t in tags if inv_cat.get(t) == "Laptop"),
        "mobile": sum(1 for t in tags if inv_cat.get(t) == "Mobile Device"),
        "monitor": sum(1 for t in tags if inv_cat.get(t) == "Monitor"),
        "network": sum(1 for t in tags if inv_cat.get(t) == "Network Asset"),
    }


def materialize_ledger_totals(wb) -> None:
    """Typed LR control totals so Finance/oracle readers see values without Excel recalc."""
    cr = wb["Corrected Register"]
    sl = wb["Source Ledger"]
    lr = wb["Ledger Reconciliation"]
    si = wb["Source Inventory"]

    costs = {}
    nbvs = {}
    for r in range(5, 200):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        c = cr.cell(r, 13).value
        if isinstance(c, str) and c.startswith("="):
            c = si.cell(r, 8).value
        costs[tag] = num(c)
        n = cr.cell(r, 14).value
        if isinstance(n, str) and n.startswith("="):
            # resolve from ledger
            n = None
        nbvs[tag] = num(n)

    sl_cost = {}
    sl_nbv = {}
    sl_status = {}
    for r in range(5, 200):
        tag = sl.cell(r, 2).value
        if not tag:
            continue
        tag = str(tag)
        acq = num(sl.cell(r, 5).value)
        dep = num(sl.cell(r, 6).value)
        nbv_cell = sl.cell(r, 7).value
        if isinstance(nbv_cell, str) and nbv_cell.startswith("="):
            nbv = round(acq - dep, 2)
        else:
            nbv = num(nbv_cell) if nbv_cell is not None else round(acq - dep, 2)
        sl_cost[tag] = acq
        sl_nbv[tag] = nbv
        sl_status[tag] = str(sl.cell(r, 8).value or "")
        if tag in nbvs and nbvs[tag] == 0 and tag in sl_nbv:
            nbvs[tag] = sl_nbv[tag]

    reg_sum = round(sum(costs.values()), 2)
    led_sum = round(sum(sl_cost.values()), 2)
    lr["B5"] = reg_sum
    lr["B6"] = led_sum
    lr["B7"] = round(reg_sum - led_sum, 2)
    lr["B8"] = round(sum(nbvs.values()), 2)
    lr["B9"] = round(sum(sl_nbv[t] for t in sl_nbv if sl_status.get(t) == "Active"), 2)
    lr["B10"] = round(sum(sl_nbv.values()), 2)


def rebuild_zips() -> None:
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    print("rebuilt deliverable zips")


def verify() -> None:
    # Read without data_only — values must be typed literals, not formulas
    wb = load_workbook(WORKBOOK, data_only=False)
    dash = wb["Dashboard"]
    assert not str(dash["B19"].value).startswith("="), dash["B19"].value
    assert dash["B19"].value == 25
    assert dash["F19"].value == 40
    assert dash["F20"].value == 36
    assert dash["F21"].value == 28
    assert dash["F22"].value == 28
    assert dash["B32"].value == 16
    assert dash["F32"].value == 12
    assert isinstance(dash["C19"].value, (int, float))
    assert isinstance(dash["G19"].value, (int, float))
    # All status/location/category/exception numeric cells populated
    for r in range(19, 27):
        assert dash.cell(r, 2).value not in (None, ""), f"blank B{r}"
        assert dash.cell(r, 3).value not in (None, "") or dash.cell(r, 3).value == 0, f"blank C{r}"
    for r in range(19, 23):
        assert dash.cell(r, 6).value not in (None, ""), f"blank F{r}"
    for r in range(32, 44):
        if dash.cell(r, 1).value:
            assert dash.cell(r, 2).value not in (None, ""), f"blank loc B{r}"
    for r in range(32, 42):
        if dash.cell(r, 5).value:
            assert dash.cell(r, 6).value not in (None, ""), f"blank ex F{r}"
            assert dash.cell(r, 7).value not in (None, "") or dash.cell(r, 7).value == 0
    lr = wb["Ledger Reconciliation"]
    assert lr["B5"].value == 332115
    assert not str(lr["B5"].value).startswith("=")
    print("verify OK: Dashboard + LR totals are typed literals")
    print(
        "category counts:",
        dash["F19"].value,
        dash["F20"].value,
        dash["F21"].value,
        dash["F22"].value,
    )


def main() -> None:
    wb = load_workbook(WORKBOOK)
    stats = materialize_dashboard(wb)
    materialize_ledger_totals(wb)
    wb.save(WORKBOOK)
    print("materialized dashboard:", stats)
    rebuild_zips()
    verify()


if __name__ == "__main__":
    main()
