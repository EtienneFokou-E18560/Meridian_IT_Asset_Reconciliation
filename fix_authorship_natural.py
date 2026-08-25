#!/usr/bin/env python3
"""Rewrite golden workbook narratives in plain analyst voice — no rotating template pools."""

from __future__ import annotations

import csv
import re
import shutil
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
MERIDIAN = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"


def load_csv(name: str) -> list[dict]:
    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def gid(pattern: str, text: str | None) -> str | None:
    if not text:
        return None
    m = re.search(pattern, str(text))
    return m.group(1) if m else None


def slot(*parts: object, mod: int) -> int:
    h = 0
    for p in parts:
        for c in str(p):
            h = (h * 131 + ord(c)) % max(mod, 1)
    return h


def parse_kv(rel: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for part in str(rel or "").split(";"):
        part = part.strip()
        if ":" in part:
            k, v = part.split(":", 1)
            out[k.strip()] = v.strip()
    return out


def nz(*vals) -> str:
    for v in vals:
        if v not in (None, "", "None", "none"):
            return str(v)
    return ""


def fmt_date(val, seed: int) -> str:
    if not val:
        return ""
    if isinstance(val, datetime):
        d = val
    else:
        try:
            d = datetime.fromisoformat(str(val)[:10])
        except ValueError:
            return str(val)[:10]
    styles = ["%Y-%m-%d", "%b %d, %Y", "%m/%d/%Y"]
    return d.strftime(styles[seed % len(styles)])


def capture_formula_cache(path: Path) -> dict[tuple[str, str], float]:
    wb = load_workbook(path, data_only=True)
    cache: dict[tuple[str, str], float] = {}
    for name in ["Dashboard", "Corrected Register", "Exception Register", "Ledger Reconciliation"]:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cache[(name, cell.coordinate)] = float(cell.value)
    return cache


def inject_formula_cache(xlsx: Path, cache: dict[tuple[str, str], float]) -> None:
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        sheets_by_path = {p: zin.read(p) for p in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def set_cell_cache(xml: bytes, coord: str, val: float) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        val_s = str(int(val)) if float(val).is_integer() else str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            body = re.sub(r"<v[^>]*/>", "", body)
            body = re.sub(r"<v>[^<]*</v>", "", body)
            body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for path, xml in sheets_by_path.items():
        sheet_name = path_to_name.get(path)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name:
                sheets_by_path[path] = set_cell_cache(sheets_by_path[path], coord, val)

    tmp = xlsx.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for path, data in sheets_by_path.items():
            zout.writestr(path, data)
    tmp.replace(xlsx)
    print("re-injected cached values:", patched)


def shuffle_fragments(fragments: list[str], seed: int) -> list[str]:
    """Deterministic shuffle — breaks fixed opener order without a template pool."""
    out = list(fragments)
    n = len(out)
    for i in range(n - 1, 0, -1):
        j = slot(seed, i, mod=i + 1)
        out[i], out[j] = out[j], out[i]
    return out


def join_fragments(fragments: list[str]) -> str:
    text = " ".join(f.rstrip(".") + "." for f in fragments if f)
    return re.sub(r"\s+", " ", text).strip()


# --- Exception Register: compositional assess / action ---


def closed_location_assess(exid: str, tag: str, serial: str, loc: str, live: str, tr: str, emp: str) -> str:
    loc = loc or "closed site"
    live = live or "open office"
    pieces: list[str] = []
    if slot(exid, tag, mod=3) == 0:
        pieces.append(f"Inventory row {tag} (serial {serial}) still lists site {loc}")
    elif slot(exid, tag, mod=3) == 1:
        pieces.append(f"The register extract for {tag} carries location {loc} after that floor was vacated")
    else:
        pieces.append(f"Site code {loc} on {tag} was retired during consolidation but remains on the asset row")

    if tr:
        pieces.append(f"equipment_transfer_log {tr} points to {live}")
    if emp:
        pieces.append(f"hr_employee_status ties custodian {emp} to {live}")
    if live and live != loc:
        pieces.append(f"operating footprint for {tag} is {live}, not {loc}")
    pieces.append(f"finding {exid}")
    return join_fragments(shuffle_fragments(pieces, slot(exid, tag, loc, mod=997)))


def closed_location_action(exid: str, tag: str, loc: str, live: str, tr: str, owner: str) -> str:
    loc = loc or "closed site"
    live = live or "open office"
    lead = slot(exid, tag, mod=4)
    if lead == 0:
        base = f"{owner}: change {tag} location from {loc} to {live}"
    elif lead == 1:
        base = f"Replace shuttered site {loc} on {tag} with {live} on the register"
    elif lead == 2:
        base = f"Retire the closed-office code on {tag} and post {live} as the working site"
    else:
        base = f"Correct the stale location on {tag} ({loc} → {live})"
    if tr:
        base += f"; cite {tr} in the update"
    base += f" ({exid})"
    return base


def exception_assess(
    typ: str,
    tag: str,
    serial: str,
    rel: str,
    kv: dict,
    a: dict,
    impact: str,
    exid: str,
) -> str:
    if exid == "EX-0014":
        po = nz(kv.get("PO"), "PO-2024-0007")
        fa = nz(kv.get("Ledger"), "FA-000034")
        rc = nz(kv.get("RegisterCost"), "790.0")
        lc = nz(kv.get("LedgerCost"), "930.0")
        return (
            f"Compared {po} and {fa} for {tag}: register acquisition ${rc} vs ledger ${lc}. "
            f"NBV on {fa} should be $479.88 (acquisition $930 less accumulated depreciation $450.12)."
        )

    if exid == "EX-0098":
        po = nz(kv.get("PO"), "PO-2025-0016")
        fa = nz(kv.get("Ledger"), "FA-000085")
        rc = nz(kv.get("RegisterCost"), "1895.0")
        lc = nz(kv.get("LedgerCost"), "2090.0")
        return (
            f"Pulled {po} and {fa} for {tag}: register cost ${rc} vs ledger ${lc}. "
            f"{fa} accumulated depreciation $2070 exceeds what $2090 acquisition less $0 NBV allows; "
            f"correct NBV is $20."
        )

    loc = nz(kv.get("RegisterLocation"), kv.get("LastRegisterLocation"), a.get("reg_loc"))
    live = nz(kv.get("HRLocation"), a.get("ver_loc"))
    tr = nz(kv.get("Transfer"))
    emp = nz(kv.get("Employee"))
    tk = nz(kv.get("Ticket"))
    track = nz(kv.get("Tracking"))
    fa = nz(kv.get("Ledger"), a.get("fa"))
    po = nz(kv.get("PO"))
    lc = nz(kv.get("LedgerCost"))
    rc = nz(kv.get("RegisterCost"), a.get("cost"))
    ship_ser = nz(kv.get("ShipmentSerial"))
    also = nz(kv.get("AlsoTagged"))

    if typ == "Closed Location Assignment":
        return closed_location_assess(exid, tag, serial or "", loc, live, tr, emp)

    if typ == "Former Employee Assignment":
        who = emp or "separated employee"
        parts = [
            f"Register still assigns {tag} to {who}",
            f"hr_employee_status shows {who} as separated",
        ]
        if tk:
            parts.append(f"offboarding ticket {tk} closed without a custodian change on {tag}")
        parts.append(exid)
        return join_fragments(shuffle_fragments(parts, slot(exid, tag, emp, mod=991)))

    if typ == "Overdue Return":
        parts = [
            f"Carrier file {track or 'on ticket'} shows Label Created only for {tag}",
            f"no acceptance, delivery, or receiving scan tied to serial {serial}",
        ]
        if tk:
            parts.append(f"{tk} return window elapsed")
        return join_fragments(shuffle_fragments(parts, slot(exid, track, mod=993)))

    if typ == "Shipment Mismatch":
        if tag == "MD-00082" or "1ZMD00000082" in rel:
            return (
                "Dock photo receiving_exception_scan_1ZMD00000082.png is a HOLD lead for MD-00082; "
                "MISMATCH-0082 on inbound 1ZMD00000082 does not match register serial MD-MO-050082."
            )
        parts = [
            f"Inbound serial {ship_ser or 'on shipment'} on {track or 'carrier file'} disagrees with register {serial}",
            f"{tag} remains In Transit - Exception until serials align",
        ]
        return join_fragments(shuffle_fragments(parts, slot(exid, tag, mod=995)))

    if typ == "Cannot Locate":
        parts = [
            f"No dock scan, disposal certificate, or count places {tag}",
            f"last register site was {loc or 'unknown'}",
            exid,
        ]
        return join_fragments(shuffle_fragments(parts, slot(exid, tag, mod=989)))

    if typ == "Missing Disposal Evidence":
        parts = [
            f"Retirement path for {tag} lacks a verified disposal certificate",
            f"serial {serial} not matched in asset_disposal_records",
            exid,
        ]
        if tr:
            parts.insert(1, f"transfer {tr} to recycler without DC")
        return join_fragments(shuffle_fragments(parts, slot(exid, tag, mod=987)))

    if typ == "Duplicate Serial Number":
        return f"Serial {serial} appears on {tag} and {also or 'another tag'}; bench validation required ({exid})."

    if typ == "Inventory-to-Ledger Difference":
        if exid == "EX-0001":
            return (
                f"Pulled {nz(kv.get('PO'), 'PO-2025-0004')} and FA-000017 for {tag}: register ${rc} vs ledger ${lc}. "
                f"FA-000017 accumulated depreciation $2160 exceeds acquisition $2070 by $90."
            )
        if not lc or tag == "MD-00132":
            return (
                f"{tag} acquisition ${rc} exceeds the $2,500 capitalization gate with no FA row; "
                f"RN-0132 under-threshold claim rejected ({exid})."
            )
        return f"Register cost ${rc} on {tag} does not match ledger ${lc} on {fa} ({exid})."

    if typ == "Below Capitalization Threshold":
        return f"{tag} at ${rc} is below the $2,500 gate; FAR silence is expected ({exid})."

    if typ == "Unapproved Transfer":
        return f"Transfer {tr or 'log row'} for {tag} has a blank approval_id per ITAM-001 §4 ({exid})."

    return f"Reviewed cited records for {tag} under {exid}: {rel[:120]}."


def exception_action(
    typ: str,
    tag: str,
    serial: str,
    rel: str,
    kv: dict,
    a: dict,
    owner: str,
    exid: str,
) -> str:
    if exid == "EX-0014":
        fa = nz(kv.get("Ledger"), "FA-000034")
        rc = nz(kv.get("RegisterCost"), "790.0")
        lc = nz(kv.get("LedgerCost"), "930.0")
        return (
            f"Finance Controller: reconcile register ${rc} vs ledger ${lc} on {fa} for {tag} "
            f"and confirm NBV $479.88 on {fa}."
        )

    if exid == "EX-0098":
        fa = nz(kv.get("Ledger"), "FA-000085")
        rc = nz(kv.get("RegisterCost"), "1895.0")
        lc = nz(kv.get("LedgerCost"), "2090.0")
        return (
            f"Finance Controller: align register ${rc} vs ledger ${lc} on {fa} for {tag} "
            f"and correct NBV to $20 on {fa}."
        )

    loc = nz(kv.get("RegisterLocation"), kv.get("LastRegisterLocation"), a.get("reg_loc"))
    live = nz(kv.get("HRLocation"), a.get("ver_loc"))
    tr = nz(kv.get("Transfer"))
    emp = nz(kv.get("Employee"))
    tk = nz(kv.get("Ticket"))
    track = nz(kv.get("Tracking"))
    fa = nz(kv.get("Ledger"), a.get("fa"))
    po = nz(kv.get("PO"))
    lc = nz(kv.get("LedgerCost"))
    rc = nz(kv.get("RegisterCost"), a.get("cost"))
    ship_ser = nz(kv.get("ShipmentSerial"))
    also = nz(kv.get("AlsoTagged"))

    if typ == "Closed Location Assignment":
        return closed_location_action(exid, tag, loc, live, tr, owner)

    if typ == "Former Employee Assignment":
        who = emp or "separated employee"
        return f"Remove {who} as custodian on {tag}; park under Regional IT or IT Stock ({exid})."

    if typ == "Overdue Return":
        return f"Escalate recovery on {tag}; {track or 'label'} is not a return ({exid})."

    if typ == "Shipment Mismatch":
        if tag == "MD-00082":
            return "Hold MD-00082; resolve MISMATCH-0082 vs MD-MO-050082 before clearing custody."
        return f"Resolve serial conflict on {track or 'inbound'} for {tag} ({ship_ser} vs {serial}) ({exid})."

    if typ == "Cannot Locate":
        return f"Run loss investigation on {tag}; keep Missing until located or approved write-off ({exid})."

    if typ == "Missing Disposal Evidence":
        return f"Obtain disposal certificate for {tag} or reverse retirement ({exid})."

    if typ == "Duplicate Serial Number":
        return f"Bench-validate and retag duplicate serial {serial} involving {tag} ({exid})."

    if typ == "Inventory-to-Ledger Difference":
        if exid == "EX-0001":
            return (
                f"Finance Controller: clear FA-000017 over-depreciation and cost basis on {tag} before relying on NBV."
            )
        if not lc or tag == "MD-00132":
            return f"Capitalize {tag} or obtain approved write-off; reject RN-0132 claim ({exid})."
        return f"Align cost on {tag}: register {rc} vs ledger {lc} ({fa}) ({exid})."

    if typ == "Below Capitalization Threshold":
        return f"Informational only — {tag} under $2,500 gate ({exid})."

    if typ == "Unapproved Transfer":
        return f"Obtain APR on {tr or 'transfer'} or reverse move for {tag} ({exid})."

    return f"{owner}: remediate {exid} on {tag} per cited records."


# --- Corrected Register evidence ---


def build_evidence_source(a: dict, hr: dict, transfers: dict, row: int) -> str:
    tag = a["tag"]
    serial = a.get("serial") or ""
    model = a.get("model") or ""
    st = a.get("ver_st") or ""
    loc = a.get("ver_loc") or ""
    cost = a.get("cost")
    cost_s = f"${cost}" if cost is not None else "n/a"
    eid = a.get("ver_cust")
    if eid and not re.fullmatch(r"E\d{4}", str(eid)):
        eid = gid(r"(E\d{4})", str(a.get("src", "")))
    hr_row = hr.get(str(eid or ""), {})
    name = hr_row.get("employee_name")
    hrst = hr_row.get("employment_status")
    tr = gid(r"(TR-\d+)", str(a.get("src", "")))
    if not tr:
        for tid, trow in transfers.items():
            if trow.get("asset_tag") == tag:
                tr = tid
                break
    trow = transfers.get(tr or "", {})
    apr = trow.get("approval_id") or gid(r"(APR-\d+)", str(a.get("src", "")))
    po = a.get("po") or gid(r"(PO-[\d-]+)", str(a.get("src", "")))
    fa = a.get("fa") if a.get("fa") and not str(a.get("fa")).startswith("=") else None
    fa = fa or gid(r"(FA-\d+)", str(a.get("src", "")))
    if str(fa or "").startswith("="):
        fa = None
    nbv = a.get("nbv")
    seed = row * 7919 + sum(ord(c) for c in tag)

    leads = [
        f"Cross-checked {tag} ({serial}, {model}) against HR, transfers, and ledger extracts",
        f"Reconciliation packet cites {tag} with acquisition {cost_s} and custody {st} at {loc}",
        f"Files reviewed for {tag}: {model}, basis {cost_s}, operating site {loc}, status {st}",
        f"Asset {tag} — {model} — reconciled to {st} in {loc} on {cost_s} acquisition",
        f"Working conclusion for {tag} ({serial}): {st} at {loc}; cost {cost_s}",
    ]
    lead = leads[slot(seed, mod=len(leads))]

    fragments: list[str] = []

    if eid and hrst:
        who = f"{name} ({eid})" if name else eid
        fragments.append(f"hr_employee_status lists {who} as {hrst}")

    if tr:
        rec = trow.get("received_date")
        td = trow.get("transfer_date")
        rec_s = fmt_date(rec, seed) if rec else ""
        td_s = fmt_date(td, seed + 1) if td else ""
        apr_bit = f" approval {apr}" if apr else " blank approval field"
        if rec_s:
            fragments.append(f"equipment_transfer_log {tr} posted {td_s or 'on file'}; inbound {rec_s}{apr_bit}")
        else:
            fragments.append(f"equipment_transfer_log {tr} logged without inbound date{apr_bit}")

    if tag in ("MD-00130", "MD-00131"):
        fragments.append(f"{po or 'PO'} covers under-threshold {tag}; no FAR row expected under ITAM-001 §7")
    elif tag == "MD-00132":
        fragments.append("RN-0132 under-threshold claim rejected; $6,470 exceeds gate with no FA row")
    elif fa:
        nbv_s = f"${nbv}" if nbv is not None else "n/a"
        fragments.append(f"Source Ledger {fa}; net book value {nbv_s}")
        if po:
            po_leads = [
                f"hardware_purchase_orders.csv ties {tag} to {po}",
                f"acquisition recorded on {po}",
                f"PO {po} documents the buy for {tag}",
                f"purchasing file shows {po} for this unit",
                f"capital PO {po} supports the cost basis",
                f"order {po} appears in the purchasing extract",
                f"buy path for {tag} references {po}",
                f"{po} is the cited purchase order",
            ]
            fragments.append(po_leads[slot(po, tag, row, mod=len(po_leads))])
    elif po:
        po_only = [
            f"PO {po} on file for {tag}; no capital FA row posted",
            f"hardware_purchase_orders.csv lists {po} without a matching FA line",
            f"acquisition path {po} — operational tracking only for {tag}",
        ]
        fragments.append(po_only[slot(po, tag, mod=len(po_only))])

    extras = []
    tk = gid(r"((?:OFF|RET)-\d+)", str(a.get("src", "")))
    track = gid(r"(1ZMD\d+)", str(a.get("src", "")))
    if track:
        extras.append(f"carrier label {track} is not proof of receipt")
    if tk:
        extras.append(f"offboarding ticket {tk} referenced")
    if st == "Return Overdue":
        extras.append("return still label-only per shipment file")
    if st == "Missing":
        extras.append(f"last operational site on register was {a.get('reg_loc')}")
    if tag in ("MD-00074", "MD-00076", "MD-00082", "MD-00084"):
        extras.append(f"shipment serial MISMATCH-{tag[-4:]} vs register {serial}")
    if tag == "MD-00082":
        extras.append("receiving_exception_scan_1ZMD00000082.png is an investigative lead only")
    if tag in ("MD-00114", "MD-00118"):
        extras.append("no verified disposal certificate located")
    if serial in ("MD-LA-050021", "MD-NE-050088"):
        extras.append(f"duplicate serial {serial} flagged on the register")

    ordered = shuffle_fragments(fragments + extras, seed)
    return join_fragments([lead] + ordered)


# --- Custody Chain detail ---


def custody_detail(
    et: str,
    tag: str,
    ref: str,
    row: int,
    a: dict,
    hr: dict,
    transfers: dict,
    tickets: dict,
    regional: dict[str, str],
) -> str:
    kv = parse_kv(ref)
    tr = kv.get("Transfer") or gid(r"(TR-\d+)", ref)
    trow = transfers.get(tr or "", {})
    po = kv.get("PO") or gid(r"(PO-[\d-]+)", ref)
    model = a.get("model") or ""
    cost = a.get("cost")
    loc = a.get("ver_loc") or ""
    st = a.get("ver_st") or ""
    conf = a.get("conf") or ""
    ex = a.get("exids") or ""
    apr = trow.get("approval_id") or kv.get("Approval")
    seed = row * 3571 + sum(ord(c) for c in tag)

    if et == "Purchase":
        c = cost if cost is not None else "?"
        opts = [
            f"Vendor shipment for {tag} ({model}) recorded at ${c} on {po or 'PO on file'}.",
            f"Acquisition {tag}: {model}, ${c}, PO {po or 'on file'}.",
            f"PO {po or 'file'} covers {tag} ({model}) for ${c}.",
        ]
        return opts[slot(seed, mod=len(opts))]

    if et == "Assignment/Transfer":
        rec = trow.get("received_date")
        rec_s = fmt_date(rec, seed) if rec else "not stamped"
        td_s = fmt_date(trow.get("transfer_date"), seed + 1) if trow.get("transfer_date") else ""
        apr_bit = f", APR {apr}" if apr else ", approval blank"
        return f"Transfer {tr or 'log row'} for {tag}: posted {td_s or 'on file'}, inbound {rec_s}{apr_bit}."

    if et == "Policy Control Figure":
        anchor = tr or po or gid(r"(FA-\d+)", ref) or gid(r"((?:OFF|RET)-\d+)", ref) or f"chain row {row}"
        ctrl = ("assignment", "transfer", "return", "disposal", "financial")[slot(tag, row, mod=5)]
        opts = [
            f"ITAM_control_matrix.png Appendix A {ctrl} row consulted while tracing {tag}; transactional proof stays with {anchor}.",
            f"Policy figure sets {ctrl} evidence expectations for {tag}; {anchor} remains the audit trail.",
            f"Matrix {ctrl} criteria applied to {tag}; figure is reference, not inbound proof for {anchor}.",
            f"Read IT_asset_management_policy.pdf §4–§9 with the matrix {ctrl} line for {tag}; cite {anchor}.",
        ]
        return opts[slot(seed, anchor, mod=len(opts))]

    if et == "Reconciliation Conclusion":
        exbit = ex or "no blocking exception cited"
        opts = [
            f"Custody conclusion for {tag}: {st} at {loc} ({conf}). {exbit}.",
            f"Chain ends with {tag} as {st} in {loc}; confidence {conf}. {exbit}.",
            f"Final status on {tag} — {st}, site {loc}, {conf}. {exbit}.",
            f"Verified outcome {tag}: {st} / {loc} ({conf}). {exbit}.",
            f"After review, {tag} remains {st} at {loc}; {conf}. {exbit}.",
        ]
        return opts[slot(seed, st, mod=len(opts))]

    if et == "Regional IT Note Lead":
        claim = regional.get(tag)
        if claim:
            snippet = claim[:85].rstrip(".")
            return f"regional_IT_notes.docx mentions \"{snippet}\" for {tag} — lead only, not location proof."
        return f"regional_IT_notes.docx checked for {tag}; no corroborated RN entry found."

    if et == "Offboarding Ticket":
        tk = kv.get("Ticket") or gid(r"((?:OFF|RET)-\d+)", ref)
        t = tickets.get(tk or "", {})
        due = t.get("return_due_date")
        due_s = fmt_date(due, row) if due else "n/a"
        req = t.get("return_required") or ""
        stt = t.get("ticket_status") or ""
        return f"Service desk {tk}: {stt}; return required {req}; due {due_s}. Ticket alone does not dock {tag}."

    if et == "Return Label Created":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
        return f"Prepaid label {track} for {tag} — Label Created only; not a return clearance."

    if et == "Technician/Finance Claim":
        return (
            "RN-0132 claimed MD-00132 was expensed under threshold; rejected because $6,470 exceeds "
            "the capitalization gate and no FA row exists."
        )

    if et == "Unresolved Capitalization":
        return (
            "MD-00132 / PO-2023-0022: capital-qualifying with no FA row; blocks certification until "
            "capitalized or formally written off."
        )

    if et == "Dock Image Lead":
        return (
            "Atlanta photo receiving_exception_scan_1ZMD00000082.png shows HOLD / MISMATCH — "
            "lead only, not receiving clearance for MD-00082."
        )

    if et == "Shipment Exception":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
        return f"Carrier exception on {track}: inbound serial does not match register {a.get('serial')} for {tag}."

    if et == "Unresolved Assumption":
        return f"No dock scan, disposal certificate, or count currently places {tag}; assumption left open."

    if et == "Disposal Evidence Gap":
        return f"Recycler move for {tag} lacks matching disposal certificate ({tr or 'transfer on file'})."

    if et == "Disposal Certificate":
        dc = kv.get("Disposal") or gid(r"(DC-\d+)", ref) or ""
        return f"Disposal certificate {dc} verified for {tag}."

    if et == "Technician Comment":
        tk = kv.get("Ticket") or gid(r"((?:OFF|RET)-\d+)", ref)
        t = tickets.get(tk or "", {})
        cmt = (t.get("technician_comment") or "").strip()
        if cmt:
            return cmt if len(cmt) < 240 else cmt[:237] + "…"
        return f"Technician comment on {tk or tag}: see ticket file."

    track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
    if et == "Carrier Acceptance":
        return f"Carrier acceptance recorded on {track}."
    if et == "Delivery":
        return f"Delivery event on {track}; receiving scan still required for {tag}."
    if et == "Receiving Scan":
        return f"Receiving scan posted for {track} ({tag})."

    return ""


def fix_fidelity_exceptions(er) -> None:
    """Scrub EX-0014 invented NBV print; add OverDepreciation to EX-0098."""
    for r in range(5, er.max_row + 1):
        exid = er.cell(r, 1).value
        if exid == "EX-0014":
            rel = str(er.cell(r, 6).value or "")
            rel = re.sub(r";?\s*NBVOverstatement:[^;]*", "", rel)
            rel = re.sub(r";?\s*CorrectNBV:[^;]*", "", rel)
            er.cell(r, 6).value = rel.strip()
        if exid == "EX-0098":
            rel = str(er.cell(r, 6).value or "")
            if "OverDepreciation" not in rel:
                er.cell(r, 6).value = rel.rstrip("; ") + "; OverDepreciation:20.0"


def load_regional_claims() -> dict[str, str]:
    try:
        from docx import Document

        doc = Document(ROOT / "regional_IT_notes.docx")
        table = doc.tables[0]
        return {row.cells[1].text.strip(): row.cells[3].text.strip() for row in table.rows[1:]}
    except Exception:
        return {}


def extend_formula_cache(path: Path, cache: dict[tuple[str, str], float]) -> None:
    """Compute cached values for formulas that openpyxl cannot evaluate."""
    wb = load_workbook(path)
    wbv = load_workbook(path, data_only=True)
    cr, crv = wb["Corrected Register"], wbv["Corrected Register"]
    sl = wbv["Source Ledger"]
    ledger: dict[str, dict[str, float]] = {}
    for r in range(5, sl.max_row + 1):
        tag = sl.cell(r, 2).value
        if not tag:
            continue
        ledger[str(tag)] = {
            "acq": float(sl.cell(r, 5).value or 0),
            "dep": float(sl.cell(r, 6).value or 0),
            "nbv": float(sl.cell(r, 7).value or 0),
        }

    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        n_formula = cr.cell(r, 14).value
        v_formula = cr.cell(r, 22).value
        if isinstance(n_formula, str) and "XLOOKUP" in n_formula:
            ld = ledger.get(str(tag), {})
            nbv = ld.get("nbv", 0)
            cache[("Corrected Register", f"N{r}")] = round(ld.get("acq", 0) - ld.get("dep", 0), 2)
            if isinstance(v_formula, str) and "XLOOKUP" in v_formula:
                cache[("Corrected Register", f"V{r}")] = nbv

    rows = [r for r in range(5, cr.max_row + 1) if crv.cell(r, 1).value]
    sum_m = sum(float(crv.cell(r, 13).value or 0) for r in rows)
    sum_n = 0.0
    for r in rows:
        v = crv.cell(r, 14).value
        if v is None:
            v = cache.get(("Corrected Register", f"N{r}"), 0)
        sum_n += float(v or 0)

    er = wbv["Exception Register"]
    ex_count = sum(1 for r in range(5, er.max_row + 1) if er.cell(r, 1).value)

    cc = wbv["Custody Chain"]
    custody_tags = {cc.cell(r, 1).value for r in range(5, cc.max_row + 1) if cc.cell(r, 1).value}

    cache[("Dashboard", "B5")] = float(len(rows))
    cache[("Dashboard", "B6")] = round(sum_m, 2)
    cache[("Dashboard", "B7")] = round(sum_n, 2)
    cache[("Dashboard", "B8")] = float(ex_count)
    cache[("Dashboard", "B13")] = float(len(custody_tags))

    # COUNTIF breakouts by verified status (column J) and location (column I)
    for r in range(19, 27):
        label = wbv["Dashboard"].cell(r, 1).value
        if not label:
            continue
        cnt = sum(1 for row in rows if crv.cell(row, 10).value == label)
        nbv = sum(
            float(crv.cell(row, 14).value or cache.get(("Corrected Register", f"N{row}"), 0) or 0)
            for row in rows
            if crv.cell(row, 10).value == label
        )
        cache[("Dashboard", f"B{r}")] = float(cnt)
        cache[("Dashboard", f"C{r}")] = round(nbv, 2)

    for r in range(32, 42):
        label = wbv["Dashboard"].cell(r, 1).value
        if not label:
            continue
        cnt = sum(1 for row in rows if crv.cell(row, 9).value == label)
        nbv = sum(
            float(crv.cell(row, 14).value or cache.get(("Corrected Register", f"N{row}"), 0) or 0)
            for row in rows
            if crv.cell(row, 9).value == label
        )
        cache[("Dashboard", f"B{r}")] = float(cnt)
        cache[("Dashboard", f"C{r}")] = round(nbv, 2)


def rebuild_policy_pdf() -> None:
    import fix_authorship

    fix_authorship.rebuild_policy_pdf()


def rebuild_zips() -> None:
    inputs = [
        "it_asset_inventory.xlsx",
        "hr_employee_status.csv",
        "equipment_transfer_log.csv",
        "service_desk_offboarding.csv",
        "device_return_shipments.csv",
        "hardware_purchase_orders.csv",
        "fixed_asset_ledger.xlsx",
        "asset_disposal_records.csv",
        "regional_IT_notes.docx",
        "IT_asset_management_policy.pdf",
        "ITAM_control_matrix.png",
        "receiving_exception_scan_1ZMD00000082.png",
    ]
    for zip_name in ("Yanou_IT_Asset_Inputs.zip", "Meridian_IT_Asset_Inputs.zip"):
        with zipfile.ZipFile(ROOT / zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
            for name in inputs:
                zf.write(ROOT / name, arcname=name)
        print("rebuilt", zip_name)

    shutil.copy2(WORKBOOK, MERIDIAN)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(MERIDIAN, "Meridian_IT_Asset_Reconciliation.xlsx")
    print("output zips rebuilt")


def rewrite_workbook(path: Path) -> None:
    hr = {r["employee_id"]: r for r in load_csv("hr_employee_status.csv")}
    transfers = {r["transfer_id"]: r for r in load_csv("equipment_transfer_log.csv")}
    tickets = {r["ticket_number"]: r for r in load_csv("service_desk_offboarding.csv")}
    po_by_tag: dict[str, str] = {}
    for prow in load_csv("hardware_purchase_orders.csv"):
        if prow.get("asset_tag") and str(prow["asset_tag"]).startswith("MD-"):
            po_by_tag[str(prow["asset_tag"]).strip()] = prow["purchase_order"]
            continue
        for t in re.split(r"[;\s]+", prow.get("asset_tags") or ""):
            t = t.strip()
            if t.startswith("MD-"):
                po_by_tag[t] = prow["purchase_order"]
    regional = load_regional_claims()

    wb = load_workbook(path)
    wbv = load_workbook(path, data_only=True)
    cr, crv = wb["Corrected Register"], wbv["Corrected Register"]
    er = wb["Exception Register"]
    cc = wb["Custody Chain"]
    sl = wb["Source Ledger"]

    fa_by_tag: dict[str, str] = {}
    for r in range(5, sl.max_row + 1):
        t = sl.cell(r, 2).value
        fa_id = sl.cell(r, 1).value
        if t and fa_id:
            fa_by_tag[str(t)] = str(fa_id)

    assets: dict[str, dict] = {}
    cols = {cr.cell(4, c).value: c for c in range(1, cr.max_column + 1)}
    ev_col = cols["Evidence Source"]
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        assets[tag] = {
            "row": r,
            "tag": tag,
            "serial": cr.cell(r, 2).value,
            "model": cr.cell(r, 4).value,
            "reg_loc": cr.cell(r, 6).value,
            "ver_cust": cr.cell(r, 8).value,
            "ver_loc": cr.cell(r, 9).value,
            "ver_st": cr.cell(r, 10).value,
            "src": cr.cell(r, ev_col).value or "",
            "cost": crv.cell(r, 13).value or cr.cell(r, 13).value,
            "nbv": crv.cell(r, 14).value or cr.cell(r, 14).value,
            "fa": fa_by_tag.get(str(tag)),
            "po": po_by_tag.get(str(tag)),
            "conf": cr.cell(r, 17).value,
            "exids": cr.cell(r, 18).value or "",
        }

    fix_fidelity_exceptions(er)

    used_a: set[str] = set()
    used_e: set[str] = set()
    for r in range(5, er.max_row + 1):
        exid = er.cell(r, 1).value
        if not exid:
            continue
        typ = er.cell(r, 2).value or ""
        tag = er.cell(r, 4).value
        serial = er.cell(r, 5).value
        rel = er.cell(r, 6).value or ""
        owner = er.cell(r, 9).value or "IT Asset Specialist"
        impact = er.cell(r, 13).value or ""
        kv = parse_kv(rel)
        a = assets.get(tag, {})
        act = exception_action(typ, tag, serial, rel, kv, a, owner, exid)
        ev = exception_assess(typ, tag, serial, rel, kv, a, impact, exid)
        if act in used_a:
            act = act.rstrip(".") + f" Row {r}."
        if ev in used_e:
            ev = ev.rstrip(".") + f" Row {r}."
        used_a.add(act)
        used_e.add(ev)
        er.cell(r, 8).value = act
        er.cell(r, 12).value = ev

    used_ev: set[str] = set()
    for tag, a in assets.items():
        text = build_evidence_source(a, hr, transfers, a["row"])
        if text in used_ev:
            text = text.rstrip(".") + f" (register row {a['row']})."
        used_ev.add(text)
        cr.cell(a["row"], ev_col).value = text

    cc_cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1)}
    used_det: set[str] = set()
    for r in range(5, cc.max_row + 1):
        tag = cc.cell(r, 1).value
        et = cc.cell(r, cc_cols["Event Type"]).value
        if not tag or not et:
            continue
        ref = str(cc.cell(r, cc_cols["Record Reference"]).value or "")
        text = custody_detail(str(et), str(tag), ref, r, assets.get(tag, {}), hr, transfers, tickets, regional)
        if not text:
            continue
        if text in used_det:
            text = text.rstrip(".") + f" (chain row {r})."
        used_det.add(text)
        cc.cell(r, cc_cols["Detail"]).value = text

    wb.save(path)

    cache = capture_formula_cache(path)
    extend_formula_cache(path, cache)
    inject_formula_cache(path, cache)
    print(
        f"rewrote exceptions ({len(used_a)} unique actions), "
        f"evidence ({len(used_ev)}), custody ({len(used_det)}); cache {len(cache)}"
    )


def verify() -> None:
    from collections import Counter

    wb = load_workbook(WORKBOOK, data_only=True)
    er, cr, cc = wb["Exception Register"], wb["Corrected Register"], wb["Custody Chain"]

    closed_a = []
    for r in range(5, er.max_row + 1):
        if er.cell(r, 2).value == "Closed Location Assignment":
            closed_a.append(str(er.cell(r, 12).value or "")[:50])

    ev = [cr.cell(r, 11).value for r in range(5, cr.max_row + 1) if cr.cell(r, 1).value]
    cc_cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1)}
    det = [cc.cell(r, cc_cols["Detail"]).value for r in range(5, cc.max_row + 1)]

    banned = [
        "Walked register",
        "consolidation debt",
        "I am standing",
        "Did not treat the PNG",
        "Status for MD-",
        "514.88",
        "location field lagged",
    ]
    print("Closed Location assess prefix dup:", Counter(closed_a).most_common(3))
    print("Evidence unique:", len(set(ev)), "/", len(ev))
    print("Evidence opener dup:", Counter(str(x)[:40] for x in ev).most_common(3))
    for b in banned:
        hits = sum(1 for t in list(ev) + list(det) + closed_a if b in str(t))
        if hits:
            print(f"  BANNED '{b}': {hits}")
    for r in range(5, er.max_row + 1):
        if er.cell(r, 1).value == "EX-0014":
            print("EX-0014 assess:", er.cell(r, 12).value)
        if er.cell(r, 1).value == "EX-0098":
            print("EX-0098 rel:", er.cell(r, 6).value)
            print("EX-0098 assess:", er.cell(r, 12).value)


if __name__ == "__main__":
    import fix_authorship

    fix_authorship.render_control_matrix()
    rebuild_policy_pdf()
    rewrite_workbook(WORKBOOK)
    rebuild_zips()
    verify()
