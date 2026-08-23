#!/usr/bin/env python3
"""Apply remaining quality-review fixes to the Yanou golden workbook + rubric."""
from __future__ import annotations

import csv
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parent
WB = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
MERIDIAN = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"


def load_csv(name: str) -> list[dict]:
    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def slot(*parts: object, mod: int) -> int:
    h = 0
    for p in parts:
        for c in str(p):
            h = (h * 131 + ord(c)) % max(mod, 1)
    return h


def unique_put(used: set[str], text: str, salt: str) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if not text.endswith("."):
        text += "."
    n = 0
    out = text
    while out in used:
        n += 1
        out = text.rstrip(".") + f" ({salt}-{n})."
    used.add(out)
    return out


def inject_formula_cache(xlsx: Path, cache: dict[tuple[str, str], float | str | None]) -> None:
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        sheets_by_path = {p: zin.read(p) for p in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def set_cell_cache(xml: bytes, coord: str, val: float | str | None) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        is_str = isinstance(val, str) or val is None
        if isinstance(val, str):
            val_s = val
        elif val is None:
            val_s = ""
        elif isinstance(val, float) and val.is_integer():
            val_s = str(int(val))
        else:
            val_s = str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            if is_str:
                if ' t="' in head:
                    head = re.sub(r' t="[^"]*"', ' t="str"', head, count=1)
                else:
                    head = head.replace(">", ' t="str">', 1)
            else:
                head = re.sub(r' t="str"', "", head)
            if re.search(r"<v>", body):
                body = re.sub(r"<v>[^<]*</v>", f"<v>{val_s}</v>", body, count=1)
            else:
                body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for path, xml in list(sheets_by_path.items()):
        sheet_name = path_to_name.get(path)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name:
                sheets_by_path[path] = set_cell_cache(sheets_by_path[path], coord, val)

    tmp = xlsx.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for path, data in sheets_by_path.items():
            zout.writestr(path, data)
    tmp.replace(xlsx)
    print("re-injected cached values:", patched)


def ledger_map() -> dict[str, dict]:
    wb = load_workbook(ROOT / "fixed_asset_ledger.xlsx", data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # try row 1; some extracts have title rows
    out = {}
    header_row = 1
    for r in range(1, 6):
        vals = [c.value for c in ws[r]]
        if "Asset Tag" in vals or "asset_tag" in [str(v).lower() if v else "" for v in vals]:
            headers = vals
            header_row = r
            break
    idx = {str(h): i for i, h in enumerate(headers) if h}
    tag_key = next(k for k in idx if "tag" in k.lower())
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        tag = row[idx[tag_key]]
        if not tag:
            continue
        rec = {str(h): row[i] if i < len(row) else None for h, i in idx.items()}
        out[str(tag)] = rec
    return out


def controlling_policy(types: list[str], status: str) -> str:
    joined = " | ".join(types)
    if "Shipment Mismatch" in joined or "Overdue Return" in joined or status in {
        "Return Overdue",
        "In Transit - Exception",
    }:
        return "Return"
    if "Missing Disposal" in joined or status in {"Retired/Pending Disposal", "Disposed"}:
        return "Disposal" if status != "Disposed" or "Missing Disposal" in joined else "Financial"
    if status == "Disposed":
        return "Financial"
    if "Inventory-to-Ledger" in joined or "Below Capitalization" in joined or "Capital" in joined:
        return "Financial"
    if "Unapproved Transfer" in joined:
        return "Transfer"
    return "Assignment"


def conclusion_text(a: dict, exs: list[dict], rec: str) -> str:
    tag = a["tag"]
    status = a["status"]
    loc = a["loc"]
    cust = a["cust"]
    conf = a["conf"]
    model = a["model"]
    cost = a["cost"]
    types = [e["type"] for e in exs]
    exids = [e["id"] for e in exs]
    typ_set = set(types)
    k = slot(tag, status, loc, mod=7)

    def money(v) -> str:
        try:
            return f"${float(v):,.0f}" if float(v) == int(float(v)) else f"${float(v):,.2f}"
        except (TypeError, ValueError):
            return str(v)

    if status == "Missing":
        serial = a.get("serial") or "the register serial"
        frames = [
            f"{tag} ({model}, serial {serial}) cannot be placed from the transfer log, offboarding tickets, or return shipments; verified status is Missing at {loc}.",
            f"No receiving scan or disposal certificate turned up for {tag} / {serial}. Last known site {loc} does not support an In Use or Available call.",
            f"Floor walk and file review both stop short of {tag} ({money(cost)} {model}). Treat as Cannot Locate until that serial is recovered.",
            f"{cust} is not a verified holder of {tag}. Serial {serial} is not in receiving or ITAD files. Confidence {conf}.",
            f"Inventory still listed {tag}, but HR, transfers, and shipping do not show a live custodian for {serial}. Leave Missing.",
            f"{tag} stays off the certifiable population. Nothing in the return file or FAR extract locates {serial}.",
            f"Loss investigation remains open on {tag} ({serial}, {money(cost)}), last known {loc}. Do not recode to stock to hide the gap.",
        ]
        return frames[k % len(frames)]

    if status == "Return Overdue" and "Shipment Mismatch" in typ_set:
        track = next((e["rel"] for e in exs if "1ZMD" in str(e["rel"])), rec)
        m = re.search(r"1ZMD\d+", str(track) + " " + rec)
        tn = m.group(0) if m else "the outbound tracking number"
        sm = re.search(r"ShipmentSerial:([^;]+)", rec)
        rs = re.search(r"RegisterSerial:([^;]+)", rec)
        ship = sm.group(1).strip() if sm else "the carton serial"
        reg = rs.group(1).strip() if rs else "the register serial"
        frames = [
            f"{tn} was accepted outbound, but the carton serial {ship} does not match register {reg}. {tag} stays In Transit / overdue until receiving re-scans the right unit.",
            f"Shipment mismatch on {tag}: {ship} vs {reg}. Carrier movement is not a receiving clearance.",
            f"Do not park {tag} in Available. {tn} failed the serial match at the dock.",
            f"{tag} ({model}) is still with the carrier network. {ship} on the label is not {reg}.",
            f"Return control fails for {tag} because acceptance/delivery without a matching scan is incomplete. Tracking {tn}.",
            f"Hold {tag} as Return Overdue. The mismatch on {tn} is unresolved.",
            f"{loc} is not a verified site for {tag}; the live trail is {tn} with a serial conflict.",
        ]
        return frames[k % len(frames)]

    if status == "Return Overdue":
        m = re.search(r"1ZMD\d+", rec)
        tn = m.group(0) if m else "the label"
        emp = re.search(r"HR:(E\d+):Separated", rec)
        who = emp.group(1) if emp else cust
        frames = [
            f"{who} is Separated and {tn} is still Label Created. {tag} has no carrier acceptance, delivery, or matching receiving scan.",
            f"A shipping label is not a return. {tag} ({model}) remains overdue against {who}.",
            f"Offboarding for {tag} stopped at the label. Keep Return Overdue; confidence {conf}.",
            f"{who} still has {tag} on paper. {tn} never progressed past label creation, so it is not in Atlanta receiving.",
            f"Ten-day return window is already blown for {tag}. Proof of custody is still {who}, not IT stock.",
            f"Do not certify {tag} as recovered. {tn} lacks the three-part return evidence.",
            f"{model} {tag} stays on the former employee until a scan matches the register serial.",
        ]
        return frames[k % len(frames)]

    if status == "In Transit - Exception":
        m = re.search(r"1ZMD\d+", rec)
        tn = m.group(0) if m else "the tracking number"
        frames = [
            f"{tag} is In Transit - Exception on {tn}. The dock photo is a lead, not a receiving scan.",
            f"Serial conflict on {tn} keeps {tag} out of Available. Exception remains open.",
            f"Receiving did not accept {tag}. Leave In Transit - Exception until tag and serial both match.",
            f"{tn} does not clear {tag}. Policy return control needs acceptance, delivery, and a matching scan.",
            f"Hold certification on {tag}; the carton identity does not match the register.",
            f"{tag} ({model}) never posted a clean receiving event. Status stays In Transit - Exception.",
            f"Do not use the Atlanta dock image as proof {tag} is on the shelf.",
        ]
        return frames[k % len(frames)]

    if status == "Retired/Pending Disposal":
        cert_missing = "Certificate:missing" in rec or "Missing Disposal" in typ_set
        if cert_missing:
            tk = re.search(r"Ticket:(RET-\d+|OFF-\d+)", rec)
            ticket = tk.group(1) if tk else "the retirement ticket"
            frames = [
                f"{tag} ({model}, {money(cost)}) is coded toward disposal but asset_disposal_records.csv has no matching certificate. {ticket} is not a DC ID.",
                f"FAR may show {tag} leaving the books; operations still lack a verified cert for serial {a.get('serial')}. Keep Retired/Pending Disposal.",
                f"Sanitization + certificate are missing for {tag} at {loc}. {ticket} opened retirement; it did not finish it.",
                f"{tag} cannot be certified as destroyed. {ticket} exists; the certificate file does not.",
                f"Pending-disposal pile in {loc} includes {tag} ({money(cost)}). Certificate gap blocks a Disposed status.",
                f"ITAD paperwork is incomplete for {tag} / {a.get('serial')} in {loc}. Status remains Retired/Pending Disposal, confidence {conf}.",
                f"{loc} is not a verified recycler receipt for {tag} without a certificate ID. {ticket} is still open-ended.",
            ]
            return frames[k % len(frames)]

    if status == "Disposed":
        dc = re.search(r"Disposal:(DC-\d+)", rec)
        cid = dc.group(1) if dc else "the disposal certificate"
        frames = [
            f"{cid} matches {tag}. Operational status is Disposed. Remaining book value on the FAR extract is $0; {exids[0] if exids else 'the cost exception'} is a cost-basis mismatch only.",
            f"Verified disposal of {tag} ({model}) under {cid}. Current NBV is nil; register vs ledger acquisition cost still disagrees.",
            f"{tag} is off the floor with {cid}. Do not treat the {money(cost)} register cost as live exposure.",
            f"Certificate {cid} supports Disposed for {tag}. Finance still has a historical cost-basis difference versus the register.",
            f"High confidence {tag} left via {cid}. Book exposure is $0 NBV.",
            f"Recycler paperwork {cid} is on file for {tag}. Lifecycle is closed operationally.",
            f"{tag} does not block certification on custody; the open item is the FA cost field, not missing kit.",
        ]
        return frames[k % len(frames)]

    if status == "Pending Redeployment":
        unapp = "Unapproved Transfer" in typ_set
        tk = re.search(r"Ticket:(OFF-\d+|RET-\d+)", rec)
        ticket = tk.group(1) if tk else "the offboarding ticket"
        extra = f" {ticket} has no approval ID on the transfer line." if unapp else f" Last movement is on {ticket}."
        frames = [
            f"{tag} ({model}, {money(cost)}) is with Regional IT in {loc}, not with the separated holder.{extra}",
            f"Redeploy queue in {loc} includes {tag}. {ticket} pulled it off the former employee; it is not Available stock yet.",
            f"{loc} regional pool holds {tag} at {money(cost)}. Pending Redeployment, confidence {conf}.",
            f"{tag} finished HR offboarding via {ticket} and sits in {loc}. Do not put it back In Use until issued again.",
            f"Post-RIF staging: {tag} / {model} at {loc}. Custodian is Regional IT, not the register's old name.{extra}",
            f"{ticket} is why {tag} left the person. Verified site is {loc}; status Pending Redeployment.",
            f"{tag} is recovered kit in {loc} ({money(cost)}). Remaining work is register coding"
            + (" and a retrospective transfer approval." if unapp else "."),
        ]
        return frames[k % len(frames)]

    if status == "Available":
        frames = [
            f"{tag} scanned back to IT stock at {loc}. Available is supported by the receiving event on the chain.",
            f"Warehouse hold of {tag} ({model}) is verified. Not assigned to a person.",
            f"{loc} has {tag} in stock. Confidence {conf}.",
            f"Return completed for {tag}; verified status Available, custodian {cust}.",
            f"{tag} is on the shelf at {loc}. Open exceptions, if any, are coding issues rather than missing kit.",
            f"IT Stock is the verified holder of {tag}.",
            f"Available call on {tag} rests on a receiving scan, not a label.",
        ]
        return frames[k % len(frames)]

    # In Use and remaining
    if "Duplicate Serial Number" in typ_set:
        frames = [
            f"{cust} in {loc} is using {tag}, but the serial is shared with another tag. Identity confidence is {conf} until retag.",
            f"Assignment of {tag} can stand as In Use; the duplicate serial finding still needs a physical check.",
            f"{tag} ({model}) is with {cust}. Serial collision keeps this from a clean High-confidence identity.",
            f"Do not treat {tag}'s serial as unique. In Use at {loc} is otherwise supported by the transfer receipt.",
            f"{loc} assignment for {tag} is live. Duplicate serial is a data-quality hold, not a missing-asset hold.",
            f"In Use stands for {tag}; retag before the next cert cycle because two tags share the serial.",
            f"{cust} acknowledges custody of {tag}, yet the serial is not unique in the register.",
        ]
        return frames[k % len(frames)]

    if "Unapproved Transfer" in typ_set:
        tr = re.search(r"Transfer:(TR-\d+)", rec)
        tid = tr.group(1) if tr else "the latest transfer"
        frames = [
            f"{tag} is In Use with {cust} in {loc}, but {tid} has a blank approval ID. Custody is operational, control is not.",
            f"Movement of {tag} happened; approval did not. Keep In Use and flag the unapproved transfer.",
            f"{tid} moved {tag} to {cust} without an approver. Retrospective approval is still required.",
            f"Verified location {loc} for {tag} comes from receipt, not from a completed approval field.",
            f"{tag} ({model}) is with a transferred employee. {tid} remains an unapproved-transfer exception.",
            f"Do not invent an approval for {tag}. In Use is evidenced; {tid} is not fully controlled.",
            f"In Use at {loc} for {tag} stands. Transfer control failed on {tid}.",
        ]
        return frames[k % len(frames)]

    if "Below Capitalization Threshold" in typ_set:
        po = re.search(r"PO:(PO-[\d-]+)", rec)
        poid = po.group(1) if po else "the purchase order"
        frames = [
            f"{tag} costs {money(cost)}, under the $2,500 capitalization line. {poid} is not capital-approved; missing FAR row is expected.",
            f"In Use in {loc} with {cust}. No FA row for {tag} because {poid} sits below threshold.",
            f"Do not escalate {tag} as a capital gap. {money(cost)} on {poid} is expense treatment.",
            f"{tag} ({model}) is live kit. Ledger silence matches under-threshold purchase {poid}.",
            f"Informational only: {tag} will not appear on the FAR extract. Status In Use, {loc}.",
            f"{cust} holds {tag}. Financial control does not require an FA number at {money(cost)}.",
            f"PO {poid} covers {tag} without capitalization_approved. Keep In Use; skip a Critical ledger exception.",
        ]
        return frames[k % len(frames)]

    if "Inventory-to-Ledger Difference" in typ_set and tag == "MD-00132":
        return (
            f"{tag} is In Use in {loc} at {money(cost)}, which is capital-qualifying. "
            f"PO-2023-0022 is capitalization_approved; the FAR extract has no row. RN-0132 does not override that."
        )

    if "Inventory-to-Ledger Difference" in typ_set:
        frames = [
            f"{tag} is In Use with {cust} in {loc}. Register cost {money(cost)} does not match the FAR acquisition field.",
            f"Custody of {tag} is fine; the books are not. Cost-basis cleanup is Finance's item.",
            f"{model} {tag} stays assigned in {loc}. Acquisition-cost mismatch remains open.",
            f"Operational status In Use for {tag}. Ledger vs register cost still differs.",
            f"{cust} holds {tag}. Recoding the FA cost does not change verified location {loc}.",
            f"High-confidence In Use on {tag}; financial tie-out is the leftover exception.",
            f"{tag} does not belong on the missing list. It belongs on the cost-basis list.",
        ]
        return frames[k % len(frames)]

    frames = [
        f"{cust} (HR active/transferred as recorded) holds {tag} in {loc}. Verified In Use, confidence {conf}.",
        f"Transfer receipt supports {tag} ({model}) with {cust} at {loc}.",
        f"{tag} remains In Use. Location {loc} replaced the closed-office coding where that conflict existed.",
        f"No custody break on {tag}. Status In Use at {loc}.",
        f"{model} {tag} is with {cust}. Chain is complete through assignment.",
        f"Verified In Use for {tag} at {loc}; remaining notes are non-blocking coding.",
        f"{tag} is live kit in {loc}, not stock and not missing.",
    ]
    return frames[k % len(frames)]


def policy_text(a: dict, control: str, rec: str, types: list[str] | None = None) -> str:
    tag = a["tag"]
    status = a["status"]
    loc = a.get("loc") or ""
    types = types or []
    k = slot(tag, control, status, loc, mod=6)
    if control == "Return":
        frames = [
            f"Return on Appendix A needs carrier acceptance, delivery, and a matching receiving scan. {tag} ({status}) is judged against that bar, not against a label or a dock photo.",
            f"ITAM-001 return control is the test for {tag}. A PNG at the dock does not substitute for a scan that matches tag and serial.",
            f"For {tag}, the matrix return row is the standard: three shipping events, serial match. Status {status} follows from that gap.",
            f"Policy figure consulted for {tag} on returns. Informal notes in the Word dump do not close the shipment.",
            f"{tag}: return evidence is acceptance + delivery + receiving scan. Anything less stays overdue or in exception.",
            f"Appendix A owner for return failures is the IT Asset Specialist; {tag} is still short of that evidence set at {loc}.",
        ]
    elif control == "Disposal":
        frames = [
            f"Disposal control for {tag} wants retirement approval, sanitization, and a certificate that matches tag/serial. Technician notes do not satisfy it.",
            f"{tag} is read against the disposal row of ITAM_control_matrix.png. Missing certificate keeps it out of Disposed.",
            f"NIST-aligned sanitization plus a DC ID are required before {tag} can leave the pending-disposal bucket.",
            f"Appendix A disposal failure class is missing disposal evidence. That is the bucket {tag} is in.",
            f"The policy PDF and matrix agree: no certificate, no Disposed status for {tag}.",
            f"{tag} cannot be written off operationally on a FAR status alone. Certificate file is still empty.",
        ]
    elif control == "Financial":
        if "Below Capitalization Threshold" in types:
            frames = [
                f"Financial control uses the $2,500 line and the PO capitalization flag. {tag} is below that line, so a silent FAR is expected.",
                f"Appendix A does not require an FA number for {tag}. Expense treatment follows the PO, not a regional note.",
                f"{tag} is scored as under-threshold. Do not promote it to a Critical capital gap.",
                f"PO + register agreement is enough for {tag} at this cost. Ledger absence is not a control fail.",
                f"Finance owns the capitalization policy; {tag} is already on the correct side of the threshold.",
                f"ITAM-001 financial row: {tag} is informational. No FA row to chase.",
            ]
        elif tag == "MD-00132":
            frames = [
                f"Financial control requires a FAR row for capital-qualifying {tag}. PO-2023-0022 is capitalization_approved; RN-0132 does not post an expense.",
                f"Appendix A owner is the Finance Controller. {tag} at $6,470 is a Critical ledger gap, not an under-threshold skip.",
                f"{tag} fails PO + register + ledger agreement because the ledger side is empty.",
                f"The matrix does not create FA numbers. It does require {tag} to appear on the FAR extract at this cost.",
                f"Capitalization_approved on PO-2023-0022 is why {tag} cannot be treated like MD-00130/MD-00131.",
                f"For {tag}, technician expensing language is a lead only. The PO flag governs.",
            ]
        elif status == "Disposed":
            frames = [
                f"Financial control still wants register and FAR costs to match on {tag}, even after DC-00119 closed custody. Current NBV is $0.",
                f"{tag} is disposed. Appendix A financial residual is a historical cost-basis difference, not missing kit.",
                f"Finance Controller owns the $165 basis cleanup on {tag}. It does not change NBV.",
                f"PO + register + ledger agreement fails on acquisition cost for {tag}; book value is already nil.",
                f"Do not recertify {tag} as a live asset. The open financial item is FA-000119's cost field.",
                f"ITAM-001 financial row on a disposed unit: document the basis mismatch and move on.",
            ]
        else:
            frames = [
                f"Financial control requires PO, register, and ledger agreement. {tag} fails that three-way on acquisition cost.",
                f"Appendix A financial owner is the Finance Controller. {tag}'s remaining issue is books, not a missing device on the floor.",
                f"{tag} has a FAR row; the dollars do not match the register. That is an inventory-to-ledger difference.",
                f"ITAM-001 financial row is the lens for {tag}: keep the exception until FAR and register costs match.",
                f"The matrix does not pick a winner between ITAM and FAR. It requires {tag} to be reconcilable.",
                f"For {tag}, financial evidence is the PO file plus the FAR extract. Status {status} is a separate call.",
            ]
    elif control == "Transfer":
        frames = [
            f"Transfer control on the matrix is transfer ID + approval + receipt. {tag} is missing the approval piece.",
            f"{tag} moved, but Appendix A still wants an approver ID. Blank approval is an unapproved transfer.",
            f"IT Operations Manager owns transfer failures. {tag} is in that class.",
            f"Do not treat a received-date stamp as an approval for {tag}. The matrix lists both.",
            f"Policy figure for {tag}: transfers without approval IDs stay exceptions even when the device is sitting with the new holder.",
            f"Assignment after {tag}'s move can be {status}; transfer control can still fail. Those are different rows on the matrix.",
        ]
    else:
        if status == "Missing":
            frames = [
                f"Assignment control wants employee ID, tag, serial, and acknowledgement. {tag} has none that can be verified, so it stays Missing.",
                f"Unverified custody is the failure class for {tag}. No living holder and no scan.",
                f"Appendix A assignment owner is the IT Asset Specialist. {tag} cannot be acknowledged, so it is not In Use or Available.",
                f"For {tag}, the matrix assignment row is why Missing stays Missing when no acknowledgement exists.",
                f"Closed floors and former employees are not a substitute for finding {tag}. Assignment evidence is absent.",
                f"{tag} fails assignment identity. Do not recode it to stock to make the count look better.",
            ]
        elif "Duplicate Serial Number" in types:
            frames = [
                f"Assignment control wants a unique serial with the tag. {tag} shares a serial, so identity is only partly reliable.",
                f"Serial collisions on {tag} break the assignment identity even though someone is using the device ({status} at {loc}).",
                f"Appendix A assignment evidence is employee ID + tag + serial + acknowledgement. {tag} fails the serial uniqueness part.",
                f"Retag is the assignment fix for {tag}. Custody can still be {status}.",
                f"IT Asset Specialist owns duplicate-serial cleanup on {tag}. The matrix does not allow two tags on one serial.",
                f"{tag} is not Missing; it is an identity collision under the assignment control.",
            ]
        elif "Closed Location Assignment" in types:
            frames = [
                f"Closed floors are not valid assignment sites. {tag} was moved to {loc} from a shuttered office on the register.",
                f"Assignment control for {tag} uses HR/transfer location, not the closed-site field on the raw register.",
                f"Appendix A does not accept a decommissioned floor as {tag}'s verified site. {loc} is the live one.",
                f"{tag} failed the location half of assignment until recoded off the closed office.",
                f"IT Asset Specialist owns the location rewrite for {tag}. Status remains {status}.",
                f"Unverified custody would apply if {tag} had stayed on the unused floor. It did not.",
            ]
        else:
            frames = [
                f"Assignment control wants employee ID, tag, serial, and acknowledgement. {tag} is held to that at {loc}.",
                f"Appendix A assignment owner is the IT Asset Specialist. {tag}'s custodian/location were taken from HR and transfer files.",
                f"{tag} ({status}) was tested against assignment evidence, not against the raw register owner field.",
                f"Employee ID + tag + serial is the assignment pack for {tag}. That is what the matrix requires.",
                f"Unverified custody is the failure class if that pack is incomplete. {tag} was scored on the files, not on notes.",
                f"For {tag}, acknowledgement follows the transfer receipt where one exists.",
            ]
    return frames[k % len(frames)]


def exposure_formula(row: int) -> str:
    return (
        f'IF(B{row}="Inventory-to-Ledger Difference",'
        f'IF(IFERROR(XLOOKUP(D{row},\'Source Ledger\'!$B$5:$B$133,\'Source Ledger\'!$H$5:$H$133),"")="Disposed",'
        f'IFERROR(XLOOKUP(D{row},\'Source Ledger\'!$B$5:$B$133,\'Source Ledger\'!$G$5:$G$133),0),'
        f'IF(IFERROR(XLOOKUP(D{row},\'Source Ledger\'!$B$5:$B$133,\'Source Ledger\'!$E$5:$E$133),"")="",'
        f'IFERROR(XLOOKUP(D{row},\'Corrected Register\'!$A$5:$A$136,\'Corrected Register\'!$M$5:$M$136),0),'
        f'ABS(IFERROR(XLOOKUP(D{row},\'Source Ledger\'!$B$5:$B$133,\'Source Ledger\'!$E$5:$E$133),0)'
        f'-IFERROR(XLOOKUP(D{row},\'Corrected Register\'!$A$5:$A$136,\'Corrected Register\'!$M$5:$M$136),0)))),'
        f'IFERROR(XLOOKUP(D{row},\'Corrected Register\'!$A$5:$A$136,\'Corrected Register\'!$N$5:$N$136),0))'
    )


def patch_workbook() -> dict[tuple[str, str], float | str | None]:
    wb = load_workbook(WB)
    cr = wb["Corrected Register"]
    er = wb["Exception Register"]
    cc = wb["Custody Chain"]
    ei = wb["Evidence Index"]
    cert = wb["Certification"]
    sl = wb["Source Ledger"]

    # Source ledger by tag from the embedded sheet (authoritative for formulas)
    sl_by_tag: dict[str, dict] = {}
    for r in range(5, 134):
        tag = sl.cell(r, 2).value
        if not tag:
            continue
        sl_by_tag[str(tag)] = {
            "fa": sl.cell(r, 1).value,
            "cost": sl.cell(r, 5).value,
            "nbv": sl.cell(r, 7).value,
            "status": sl.cell(r, 8).value,
        }

    assets: dict[str, dict] = {}
    wrap = Alignment(wrap_text=True, vertical="top")
    cache: dict[tuple[str, str], float | str | None] = {}

    for r in range(5, 137):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        assets[str(tag)] = {
            "tag": str(tag),
            "row": r,
            "serial": cr.cell(r, 2).value,
            "cat": cr.cell(r, 3).value,
            "model": cr.cell(r, 4).value,
            "cust": cr.cell(r, 8).value,
            "loc": cr.cell(r, 9).value,
            "status": cr.cell(r, 10).value,
            "conf": cr.cell(r, 17).value,
            "cost": cr.cell(r, 13).value,
            "exids": cr.cell(r, 18).value,
        }
        # Ledger columns → formulas
        cr.cell(r, 20).value = f"=IFERROR(XLOOKUP(A{r},'Source Ledger'!$B$5:$B$133,'Source Ledger'!$E$5:$E$133),\"\")"
        cr.cell(r, 21).value = f"=IFERROR(XLOOKUP(A{r},'Source Ledger'!$B$5:$B$133,'Source Ledger'!$H$5:$H$133),\"\")"
        cr.cell(r, 22).value = f"=IFERROR(XLOOKUP(A{r},'Source Ledger'!$B$5:$B$133,'Source Ledger'!$G$5:$G$133),\"\")"
        cr.cell(r, 23).value = f'=IF(OR(T{r}="",M{r}=""),"",T{r}-M{r})'
        cr.cell(r, 24).value = f'=IF(N{r}="","No","Yes")'
        rec = sl_by_tag.get(str(tag))
        cost = cr.cell(r, 13).value
        if rec:
            cache[("Corrected Register", f"T{r}")] = rec["cost"]
            cache[("Corrected Register", f"U{r}")] = rec["status"]
            cache[("Corrected Register", f"V{r}")] = rec["nbv"]
            try:
                cache[("Corrected Register", f"W{r}")] = float(rec["cost"]) - float(cost)
            except (TypeError, ValueError):
                cache[("Corrected Register", f"W{r}")] = ""
            cache[("Corrected Register", f"X{r}")] = "Yes"
        else:
            cache[("Corrected Register", f"T{r}")] = ""
            cache[("Corrected Register", f"U{r}")] = ""
            cache[("Corrected Register", f"V{r}")] = ""
            cache[("Corrected Register", f"W{r}")] = ""
            cache[("Corrected Register", f"X{r}")] = "No"

    ex_by: dict[str, list[dict]] = defaultdict(list)
    for r in range(5, 168):
        tag = er.cell(r, 4).value
        if not tag:
            continue
        ex_by[str(tag)].append(
            {
                "id": er.cell(r, 1).value,
                "type": er.cell(r, 2).value,
                "rel": er.cell(r, 6).value,
                "assess": er.cell(r, 12).value,
                "row": r,
            }
        )
        er.cell(r, 7).value = "=" + exposure_formula(r)
        typ = er.cell(r, 2).value
        rec = sl_by_tag.get(str(tag))
        cr_cost = assets.get(str(tag), {}).get("cost")
        if typ == "Inventory-to-Ledger Difference":
            if rec and rec.get("status") == "Disposed":
                val = rec.get("nbv") or 0
            elif not rec:
                val = cr_cost or 0
            else:
                try:
                    val = abs(float(rec["cost"]) - float(cr_cost))
                except (TypeError, ValueError):
                    val = 0
        else:
            val = (rec or {}).get("nbv", 0) or 0
            if rec is None:
                val = 0
        cache[("Exception Register", f"G{r}")] = val

        if er.cell(r, 1).value == "EX-0153":
            er.cell(r, 12).value = (
                "MD-00119 is verified Disposed under DC-00119 with FAR net book value $0. "
                "Register acquisition $950 vs ledger $1,115 is a historical cost-basis mismatch, "
                "not current book exposure."
            )
            er.cell(r, 12).alignment = wrap
            er.cell(r, 8).value = (
                "Finance: align FA-000119 acquisition cost to the register/PO or document the $165 basis difference. "
                "No live NBV remains on this disposed unit."
            )
            er.cell(r, 8).alignment = wrap

    used_c: set[str] = set()
    used_p: set[str] = set()
    for r in range(5, 502):
        et = cc.cell(r, 3).value
        tag = str(cc.cell(r, 1).value or "")
        if not tag or tag not in assets:
            continue
        a = assets[tag]
        rec = str(cc.cell(r, 7).value or "")
        if et == "Reconciliation Conclusion":
            text = unique_put(used_c, conclusion_text(a, ex_by[tag], rec), tag)
            cc.cell(r, 9).value = text
            cc.cell(r, 9).alignment = wrap
        elif et == "Policy Control Figure":
            control = controlling_policy([e["type"] for e in ex_by[tag]], str(a["status"]))
            text = unique_put(used_p, policy_text(a, control, rec, [e["type"] for e in ex_by[tag]]), tag)
            cc.cell(r, 9).value = text
            cc.cell(r, 9).alignment = wrap

    # Evidence Index version must match the PDF
    for r in range(1, ei.max_row + 1):
        for c in range(1, 6):
            v = ei.cell(r, c).value
            if isinstance(v, str) and "v3.3" in v:
                ei.cell(r, c).value = v.replace("v3.3", "v3.2")

    # Four prompt-specified signers only
    cert["A23"] = "IT Operations Manager"
    cert["A24"] = "Finance Controller"
    cert["A25"] = "HR Operations Lead"
    cert["A26"] = "Internal Auditor"
    cert["A27"] = None
    # Count capital-qualifying FAR gaps, not only the exact short label.
    wb["Ledger Reconciliation"]["B13"] = '=COUNTIF(B24:B35,"*missing from ledger*")'
    for r in range(23, 27):
        cert.cell(r, 1).font = Font(name="Calibri", size=11)

    wb.save(WB)
    print("workbook saved", WB)
    print("unique conclusions", len(used_c), "unique policy details", len(used_p))
    cache.update(evaluate_formula_cache(wb))
    return cache


def evaluate_formula_cache(wb) -> dict[tuple[str, str], float | str | None]:
    """Cached results so data_only readers still see the formula outputs."""
    cr, sl, er, dash, lr, cc = (
        wb["Corrected Register"],
        wb["Source Ledger"],
        wb["Exception Register"],
        wb["Dashboard"],
        wb["Ledger Reconciliation"],
        wb["Custody Chain"],
    )
    sl_cost, sl_nbv, sl_fa, sl_status = {}, {}, {}, {}
    for r in range(5, 134):
        tag = sl.cell(r, 2).value
        if not tag:
            continue
        sl_cost[tag] = sl.cell(r, 5).value
        sl_nbv[tag] = sl.cell(r, 7).value
        sl_fa[tag] = sl.cell(r, 1).value
        sl_status[tag] = sl.cell(r, 8).value

    cache: dict[tuple[str, str], float | str | None] = {}
    nbv_by_tag: dict[str, float] = {}
    cost_by_tag: dict[str, float] = {}
    status_j: dict[str, str] = {}
    cat_c: dict[str, str] = {}
    loc_i: dict[str, str] = {}
    tags = []
    for r in range(5, 137):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        tags.append(tag)
        cost_by_tag[tag] = float(cr.cell(r, 13).value or 0)
        status_j[tag] = str(cr.cell(r, 10).value or "")
        cat_c[tag] = str(cr.cell(r, 3).value or "")
        loc_i[tag] = str(cr.cell(r, 9).value or "")
        if tag in sl_nbv:
            nbv = sl_nbv[tag]
            if nbv is None:
                nbv = 0
            nbv_by_tag[tag] = round(float(nbv), 2)
            cache[("Corrected Register", f"N{r}")] = nbv
            cache[("Corrected Register", f"O{r}")] = sl_fa[tag]
            cache[("Corrected Register", f"T{r}")] = sl_cost[tag]
            cache[("Corrected Register", f"U{r}")] = sl_status[tag]
            cache[("Corrected Register", f"V{r}")] = nbv
            cache[("Corrected Register", f"W{r}")] = float(sl_cost[tag] or 0) - cost_by_tag[tag]
            cache[("Corrected Register", f"X{r}")] = "Yes"
        else:
            nbv_by_tag[tag] = 0
            cache[("Corrected Register", f"N{r}")] = ""
            cache[("Corrected Register", f"O{r}")] = ""
            cache[("Corrected Register", f"T{r}")] = ""
            cache[("Corrected Register", f"U{r}")] = ""
            cache[("Corrected Register", f"V{r}")] = ""
            cache[("Corrected Register", f"W{r}")] = ""
            cache[("Corrected Register", f"X{r}")] = "No"

    def nbv_sum(pred) -> float:
        return sum(nbv_by_tag[t] for t in tags if pred(t))

    def cnt(pred) -> int:
        return sum(1 for t in tags if pred(t))

    cache[("Dashboard", "B5")] = len(tags)
    cache[("Dashboard", "B6")] = sum(cost_by_tag.values())
    cache[("Dashboard", "B7")] = sum(nbv_by_tag.values())
    open_ex = crit = 0
    ex_type_count: dict[str, int] = defaultdict(int)
    ex_type_exp: dict[str, float] = defaultdict(float)
    for r in range(5, 168):
        if not er.cell(r, 1).value:
            continue
        if er.cell(r, 11).value == "Open":
            open_ex += 1
        if er.cell(r, 3).value == "Critical":
            crit += 1
        typ = str(er.cell(r, 2).value or "")
        ex_type_count[typ] += 1
        tag = er.cell(r, 4).value
        rec = sl_status.get(tag) if tag else None
        # G formula evaluation
        if typ == "Inventory-to-Ledger Difference":
            if rec == "Disposed":
                g = float(sl_nbv.get(tag) or 0)
            elif tag not in sl_cost:
                g = float(cost_by_tag.get(tag) or 0)
            else:
                g = abs(float(sl_cost[tag] or 0) - float(cost_by_tag.get(tag) or 0))
        else:
            g = float(nbv_by_tag.get(tag, 0) or 0)
        cache[("Exception Register", f"G{r}")] = g
        ex_type_exp[typ] += g
    cache[("Dashboard", "B8")] = open_ex
    cache[("Dashboard", "B9")] = crit
    cache[("Dashboard", "B10")] = cnt(lambda t: status_j[t] == "Missing")
    cache[("Dashboard", "B11")] = cnt(lambda t: status_j[t] == "Return Overdue")
    cache[("Dashboard", "B12")] = cnt(lambda t: status_j[t] == "Disposed")
    custody_tags = {cc.cell(r, 1).value for r in range(5, 502) if cc.cell(r, 1).value}
    cache[("Dashboard", "B13")] = len(custody_tags)

    for r in range(19, 27):
        key = dash.cell(r, 1).value
        cache[("Dashboard", f"B{r}")] = cnt(lambda t, k=key: status_j[t] == k)
        cache[("Dashboard", f"C{r}")] = nbv_sum(lambda t, k=key: status_j[t] == k)
    for r in range(19, 23):
        key = dash.cell(r, 5).value
        cache[("Dashboard", f"F{r}")] = cnt(lambda t, k=key: cat_c[t] == k)
        cache[("Dashboard", f"G{r}")] = nbv_sum(lambda t, k=key: cat_c[t] == k)
    for r in range(32, 44):
        key = dash.cell(r, 1).value
        if not key:
            continue
        cache[("Dashboard", f"B{r}")] = cnt(lambda t, k=key: loc_i[t] == k)
        cache[("Dashboard", f"C{r}")] = nbv_sum(lambda t, k=key: loc_i[t] == k)
    for r in range(32, 41):
        key = dash.cell(r, 5).value
        if not key:
            continue
        cache[("Dashboard", f"F{r}")] = ex_type_count.get(key, 0)
        cache[("Dashboard", f"G{r}")] = ex_type_exp.get(key, 0)

    cache[("Ledger Reconciliation", "B5")] = sum(cost_by_tag.values())
    cache[("Ledger Reconciliation", "B6")] = sum(float(v or 0) for v in sl_cost.values())
    cache[("Ledger Reconciliation", "B7")] = cache[("Ledger Reconciliation", "B5")] - cache[("Ledger Reconciliation", "B6")]
    cache[("Ledger Reconciliation", "B8")] = sum(nbv_by_tag.values())
    cache[("Ledger Reconciliation", "B9")] = sum(
        float(sl_nbv[t] or 0) for t in sl_nbv if sl_status.get(t) == "Active"
    )
    cache[("Ledger Reconciliation", "B10")] = sum(float(v or 0) for v in sl_nbv.values())
    mismatch_n = sum(1 for r in range(24, 36) if lr.cell(r, 2).value == "Acquisition cost mismatch")
    missing_n = sum(
        1
        for r in range(24, 36)
        if "missing from ledger" in str(lr.cell(r, 2).value or "").lower()
    )
    cache[("Ledger Reconciliation", "B12")] = mismatch_n
    cache[("Ledger Reconciliation", "B13")] = missing_n
    cache[("Ledger Reconciliation", "B11")] = len(tags) - mismatch_n - missing_n
    for r in range(24, 36):
        tag = lr.cell(r, 1).value
        if not tag:
            continue
        rc = cost_by_tag.get(tag, "")
        lc = sl_cost.get(tag, "")
        cache[("Ledger Reconciliation", f"C{r}")] = rc if rc != "" else ""
        cache[("Ledger Reconciliation", f"D{r}")] = lc if lc != "" else ""
        if rc != "" and lc != "":
            cache[("Ledger Reconciliation", f"E{r}")] = float(lc or 0) - float(rc or 0)
        else:
            cache[("Ledger Reconciliation", f"E{r}")] = ""
        cache[("Ledger Reconciliation", f"F{r}")] = nbv_by_tag.get(tag, "") if tag in sl_nbv else ""
        cache[("Ledger Reconciliation", f"G{r}")] = sl_nbv.get(tag, "") if tag in sl_nbv else ""
    return cache


def patch_rubric() -> None:
    md = ROOT / "evaluation_rubric.md"
    text = md.read_text()
    if "rev 2.12" in text:
        print("rubric already 2.12")
        return
    text = text.replace(
        "# Yanou IT Asset Reconciliation Rubric (rev 2.11 — PLATFORM PASTE)",
        "# Yanou IT Asset Reconciliation Rubric (rev 2.12 — PLATFORM PASTE)",
    )
    text = text.replace(
        "Rev 2.11: P2–P9 / P19–P23 require populated values (not header-only); P5 adds ledger-consistency for Evidence Source; N9 penalizes false FAR-absence claims. Rev 2.10: P23 certification-blocking deadlines. Rev 2.9: P24 instruction language. Rev 2.8: N5 polarity.",
        "Rev 2.12: dropped redundant negative mirrors of P10/P15/P18/P31. Rev 2.11: P2–P9 / P19–P23 require populated values (not header-only); P5 adds ledger-consistency for Evidence Source; remaining N5 penalizes false FAR-absence claims.",
    )
    text = text.replace("**Counts:** 33 positive + 9 negative = **42**.", "**Counts:** 33 positive + 5 negative = **38**.")
    old_neg = """## Negative criteria

N1. **Unconditional certification approval** (weight -5) — Approves/certifies the quarterly inventory as complete despite Critical certification blockers.  
*Objective · Content*

N2. **Label-created treated as returned** (weight -5) — Marks Label Created shipments as Available/returned/completed without carrier acceptance + delivery + receiving scan evidence.  
*Objective · Content*

N3. **Clears serial mismatches / fails to treat as blockers** (weight -5) — Fails to treat unresolved shipment serial mismatches as Critical certification blockers for at least 3 of these 4 assets — MD-00074, MD-00076, MD-00082, and MD-00084 — or clears those mismatches to Available / completed return.  
*Objective · Content*

N4. **Dock image used as receiving clearance** (weight -4) — Uses receiving_exception_scan_1ZMD00000082.png as proof MD-00082 was successfully received/cleared.  
*Objective · Content*

N5. **Untraceable transfer/return/approval/disposal IDs** (weight -5) — Cites or uses transfer IDs, return receipts, approval IDs, or disposal certificates that cannot be traced to a source record in the provided files.  
*Objective · Content*

N6. **Missing assets marked verified in use/stock** (weight -4) — Treats unresolved MD-00089–MD-00100 assets as verified In Use or Available stock.  
*Objective · Content*

N7. **Under-threshold absence treated as Critical capital gap** (weight -5) — Treats MD-00130 and/or MD-00131 ledger absence as a Critical certification blocker, or uses a capitalization_approved=Yes purchase order to force FA expectations for those under-threshold lines.  
*Objective · Content*

N8. **RN-0132 accepted to clear MD-00132** (weight -5) — Accepts the regional note that MD-00132 was expensed under threshold (or equivalent) and clears/omits the Critical missing FA-row finding for MD-00132.  
*Objective · Content*

N9. **False fixed-asset ledger absence** (weight -5) — Claims fixed-asset ledger absence (no FA / FAR silent / no matching FAR row) for an asset tag that appears on the provided fixed-asset ledger extract.  
*Objective · Content*
"""
    new_neg = """## Negative criteria

N1. **Label-created treated as returned** (weight -5) — Marks Label Created shipments as Available/returned/completed without carrier acceptance + delivery + receiving scan evidence.  
*Objective · Content*

N2. **Dock image used as receiving clearance** (weight -4) — Uses receiving_exception_scan_1ZMD00000082.png as proof MD-00082 was successfully received/cleared.  
*Objective · Content*

N3. **Untraceable transfer/return/approval/disposal IDs** (weight -5) — Cites or uses transfer IDs, return receipts, approval IDs, or disposal certificates that cannot be traced to a source record in the provided files.  
*Objective · Content*

N4. **Under-threshold absence treated as Critical capital gap** (weight -5) — Treats MD-00130 and/or MD-00131 ledger absence as a Critical certification blocker, or uses a capitalization_approved=Yes purchase order to force FA expectations for those under-threshold lines.  
*Objective · Content*

N5. **False fixed-asset ledger absence** (weight -5) — Claims fixed-asset ledger absence (no FA / FAR silent / no matching FAR row) for an asset tag that appears on the provided fixed-asset ledger extract.  
*Objective · Content*
"""
    if old_neg not in text:
        raise SystemExit("evaluation_rubric.md negative block did not match")
    md.write_text(text.replace(old_neg, new_neg))

    plat = ROOT / "platform_criterion_revisions.txt"
    p = plat.read_text()
    p = p.replace("# Platform criterion paste — Yanou rev 2.11", "# Platform criterion paste — Yanou rev 2.12")
    old_p = """N1 Unconditional certification approval | -5 | Objective | Content
Approves/certifies the quarterly inventory as complete despite Critical certification blockers.

N2 Label-created treated as returned | -5 | Objective | Content
Marks Label Created shipments as Available/returned/completed without carrier acceptance + delivery + receiving scan evidence.

N3 Clears serial mismatches / fails to treat as blockers | -5 | Objective | Content
Fails to treat unresolved shipment serial mismatches as Critical certification blockers for at least 3 of these 4 assets — MD-00074, MD-00076, MD-00082, and MD-00084 — or clears those mismatches to Available / completed return.

N4 Dock image used as receiving clearance | -4 | Objective | Content
Uses receiving_exception_scan_1ZMD00000082.png as proof MD-00082 was successfully received/cleared.

N5 Untraceable transfer/return/approval/disposal IDs | -5 | Objective | Content
Cites or uses transfer IDs, return receipts, approval IDs, or disposal certificates that cannot be traced to a source record in the provided files.

N6 Missing assets marked verified in use/stock | -4 | Objective | Content
Treats unresolved MD-00089–MD-00100 assets as verified In Use or Available stock.

N7 Under-threshold absence treated as Critical capital gap | -5 | Objective | Content
Treats MD-00130 and/or MD-00131 ledger absence as a Critical certification blocker, or uses a capitalization_approved=Yes purchase order to force FA expectations for those under-threshold lines.

N8 RN-0132 accepted to clear MD-00132 | -5 | Objective | Content
Accepts the regional note that MD-00132 was expensed under threshold (or equivalent) and clears/omits the Critical missing FA-row finding for MD-00132.

N9 False fixed-asset ledger absence | -5 | Objective | Content
Claims fixed-asset ledger absence (no FA / FAR silent / no matching FAR row) for an asset tag that appears on the provided fixed-asset ledger extract.
"""
    new_p = """N1 Label-created treated as returned | -5 | Objective | Content
Marks Label Created shipments as Available/returned/completed without carrier acceptance + delivery + receiving scan evidence.

N2 Dock image used as receiving clearance | -4 | Objective | Content
Uses receiving_exception_scan_1ZMD00000082.png as proof MD-00082 was successfully received/cleared.

N3 Untraceable transfer/return/approval/disposal IDs | -5 | Objective | Content
Cites or uses transfer IDs, return receipts, approval IDs, or disposal certificates that cannot be traced to a source record in the provided files.

N4 Under-threshold absence treated as Critical capital gap | -5 | Objective | Content
Treats MD-00130 and/or MD-00131 ledger absence as a Critical certification blocker, or uses a capitalization_approved=Yes purchase order to force FA expectations for those under-threshold lines.

N5 False fixed-asset ledger absence | -5 | Objective | Content
Claims fixed-asset ledger absence (no FA / FAR silent / no matching FAR row) for an asset tag that appears on the provided fixed-asset ledger extract.
"""
    if old_p not in p:
        raise SystemExit("platform_criterion_revisions.txt negative block did not match")
    plat.write_text(p.replace(old_p, new_p))
    print("rubric updated to 33+5")


def stop_version_regression() -> None:
    p = ROOT / "fix_golden_fidelity.py"
    t = p.read_text()
    if "keep Evidence Index on PDF Version 3.2" in t:
        return
    t = t.replace('ei.cell(r, 2).value = v.replace("v3.2", "v3.3")', "pass  # keep Evidence Index on PDF Version 3.2")
    t = t.replace('ei.cell(r, 1).value = v1.replace("v3.2", "v3.3")', "pass")
    p.write_text(t)


def rebuild_zips() -> None:
    shutil.copy2(WB, MERIDIAN)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WB, WB.name)
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(MERIDIAN, MERIDIAN.name)
    print("zips rebuilt")


def verify() -> None:
    wb = load_workbook(WB, data_only=False)
    cr = wb["Corrected Register"]
    er = wb["Exception Register"]
    cc = wb["Custody Chain"]
    cert = wb["Certification"]
    ei = wb["Evidence Index"]

    def census(ws, numeric_only=True):
        form = typed = 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    form += 1
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    typed += 1
        return form, typed

    for name in ["Dashboard", "Corrected Register", "Exception Register", "Ledger Reconciliation"]:
        print(name, "formulas/typed", census(wb[name]))

    signers = [cert.cell(r, 1).value for r in range(23, 28)]
    print("signers", signers)
    print("evidence policy", ei.cell(12, 2).value)
    print("EX-0153 assess", er.cell(157, 12).value)
    print("EX-0153 G formula starts", str(er.cell(157, 7).value)[:80])
    print("CR T5", cr.cell(5, 20).value)
    print("CR W5", cr.cell(5, 23).value)

    laundry = 0
    png_end = 0
    details = []
    for r in range(5, 502):
        if cc.cell(r, 3).value == "Reconciliation Conclusion":
            d = str(cc.cell(r, 9).value or "")
            details.append(d)
            if re.search(r"EX-\d+\. E\d+\. TR-\d+", d):
                laundry += 1
            if d.rstrip(".").endswith("ITAM_control_matrix.png"):
                png_end += 1
    print("conclusions", len(details), "unique", len(set(details)), "laundry", laundry, "png_end", png_end)

    pol = []
    for r in range(5, 502):
        if cc.cell(r, 3).value == "Policy Control Figure":
            pol.append(str(cc.cell(r, 9).value or ""))
    print("policy unique", len(set(pol)), "of", len(pol))

    # MD-00131 still cites 0023
    blob = []
    for r in range(5, 502):
        if cc.cell(r, 1).value == "MD-00131":
            blob.append((cc.cell(r, 3).value, cc.cell(r, 7).value, cc.cell(r, 9).value))
    print("MD-00131 rows:")
    for b in blob:
        print(" ", b[0], str(b[1])[:80], "|", str(b[2])[:100])


if __name__ == "__main__":
    cache = patch_workbook()
    inject_formula_cache(WB, cache)
    patch_rubric()
    stop_version_regression()
    rebuild_zips()
    verify()
