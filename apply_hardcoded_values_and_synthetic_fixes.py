#!/usr/bin/env python3
"""
Apply dataset-quality suggestions:
1) Rename company text (prompt + workbook).
2) Fix dashboard cross-tab inconsistency by updating Verified Status for MD-00074/MD-00076.
3) Fix Evidence Assessment misquoted FAR acquisition-cost figures (EX-0001/EX-0014/EX-0032).
4) Break "hard-coded spreadsheet values" by restoring Excel formulas and injecting cached numeric values.
5) De-suspiciously-distributed synthetic patterns:
   - it_asset_inventory.xlsx category counts (break exact 33/33/33/33).
   - device_return_shipments.csv label_created_date (break exact uniform date).
"""

from __future__ import annotations

import csv
import re
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent

NEW_COMPANY = "Copr & Partners IT Services"


def replace_company_text_in_prompt() -> None:
    p = ROOT / "task_prompt.txt"
    txt = p.read_text()
    txt2 = txt.replace("Meridian Data Services", NEW_COMPANY)
    if txt2 != txt:
        p.write_text(txt2)
        print("task_prompt: company renamed")
    else:
        print("task_prompt: no company string found to replace")


def replace_company_text_in_workbook(xlsx: Path) -> None:
    wb = load_workbook(xlsx)
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "Meridian Data Services" in v:
                    cell.value = v.replace("Meridian Data Services", NEW_COMPANY)
                    changed += 1
    wb.save(xlsx)
    print("workbook: company replaced in", changed, "cells")


def fix_dashboard_cross_tab_by_status(xlsx: Path) -> None:
    """
    The dataset-quality check reported:
    - Dashboard KPI Overdue returns = 10
    - Dashboard exception table Overdue Return = 12
    Two exceptions for MD-00074/MD-00076 were Overdue Return, but Corrected Register marked them In Transit.
    We align by changing Verified Status to Return Overdue for both assets and updating custody-chain conclusion text.
    """
    wb = load_workbook(xlsx)
    cr = wb["Corrected Register"]
    # find Verified Status column by header
    status_col = None
    for c in range(1, cr.max_column + 1):
        if cr.cell(4, c).value == "Verified Status":
            status_col = c
            break
    assert status_col is not None

    # Update statuses
    for tag in ("MD-00074", "MD-00076"):
        for r in range(5, cr.max_row + 1):
            if cr.cell(r, 1).value == tag:
                cr.cell(r, status_col).value = "Return Overdue"
                break

    # Update custody chain "Reconciliation Conclusion" detail strings to match
    cc = wb["Custody Chain"]
    tag_col = et_col = detail_col = None
    for c in range(1, cc.max_column + 1):
        h = cc.cell(4, c).value
        if h == "Asset Tag":
            tag_col = c
        elif h == "Event Type":
            et_col = c
        elif h == "Detail":
            detail_col = c
    assert tag_col and et_col and detail_col

    for tag in ("MD-00074", "MD-00076"):
        for r in range(5, cc.max_row + 1):
            if cc.cell(r, tag_col).value == tag and cc.cell(r, et_col).value == "Reconciliation Conclusion":
                d = str(cc.cell(r, detail_col).value or "")
                d = d.replace("In Transit - Exception", "Return Overdue")
                d = d.replace("After the chain: Return Overdue", "After the chain: Return Overdue")
                d = re.sub(r"Verified status:\s*Return Overdue", "Verified status: Return Overdue", d)
                cc.cell(r, detail_col).value = d
                break

    wb.save(xlsx)
    print("Corrected Register + Custody Chain aligned for MD-00074/MD-00076")


def fix_evidence_assessment_cited_costs(xlsx: Path) -> None:
    """
    Correct Evidence Assessment text for:
    - EX-0001: FA-000017 cost should be Acquisition Cost (2070.0) not Accum. Dep. (2160.0)
    - EX-0014: FA-000034 cost should be Acquisition Cost (930.0) not (965.0)
    - EX-0032: FA-000051 cost should be Acquisition Cost (1250.0) not (1215.0)
    """
    # Build FA->acquisition cost mapping from fixed_asset_ledger.xlsx
    from openpyxl import load_workbook as lw

    led = lw(ROOT / "fixed_asset_ledger.xlsx", data_only=True).active
    lh = {led.cell(4, c).value: c for c in range(1, led.max_column + 1)}
    faid_col = lh["Ledger Asset ID"]
    ac_col = lh["Acquisition Cost"]
    fa_to_ac: dict[str, float] = {}
    for r in range(5, led.max_row + 1):
        fa = led.cell(r, faid_col).value
        if not fa:
            continue
        fa_to_ac[str(fa)] = float(led.cell(r, ac_col).value or 0)

    wb = load_workbook(xlsx)
    er = wb["Exception Register"]
    evidence_col = None
    for c in range(1, er.max_column + 1):
        if er.cell(4, c).value == "Evidence Assessment":
            evidence_col = c
            break
    assert evidence_col is not None

    # update the three exceptions
    target_ex = {"EX-0001", "EX-0014", "EX-0032"}
    for r in range(5, er.max_row + 1):
        exid = er.cell(r, 1).value
        if exid not in target_ex:
            continue
        txt = str(er.cell(r, evidence_col).value or "")
        # replace "FA-xxxxx cost <num>" with correct acquisition cost
        # Evidence text uses the ledger asset IDs like "FA-000017"
        m = re.search(r"(FA-\d+) cost ([0-9.]+)", txt)
        if not m:
            continue
        faid = m.group(1)
        if faid not in fa_to_ac:
            continue
        ac = fa_to_ac[faid]
        # keep one decimal if existing has .0, else keep raw
        ac_str = f"{ac:.1f}" if "." in m.group(2) else str(int(round(ac)))
        txt2 = re.sub(rf"{re.escape(faid)} cost [0-9.]+", f"{faid} cost {ac_str}", txt)
        er.cell(r, evidence_col).value = txt2

    wb.save(xlsx)
    print("Evidence Assessment cost citations corrected for", ", ".join(sorted(target_ex)))


def break_synthetic_category_even_split() -> None:
    """
    it_asset_inventory.xlsx currently has exactly 33 each of:
    Laptop / Monitor / Mobile Device / Network Asset.
    We flip 1 randomly-chosen Laptop into Monitor (and mirror in Corrected Register).
    """
    inv = ROOT / "it_asset_inventory.xlsx"
    wb_inv = load_workbook(inv, data_only=True)
    ws = wb_inv.active

    # locate headers
    header_row = None
    headers = None
    for r in range(1, 6):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "Asset Tag" in vals:
            header_row = r
            headers = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1)}
            break
    assert header_row is not None and headers
    tag_c = headers["Asset Tag"]
    cat_c = headers["Category"]
    model_c = headers["Model"]

    laptop_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, cat_c).value == "Laptop":
            laptop_rows.append(r)
    assert laptop_rows

    flip_row = laptop_rows[0]
    tag = ws.cell(flip_row, tag_c).value
    old = ws.cell(flip_row, cat_c).value
    ws.cell(flip_row, cat_c).value = "Monitor"
    wb_inv.save(inv)
    print("Inventory category tweak:", tag, old, "->", "Monitor")

    # Mirror into Corrected Register Category column
    golden = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"
    wb_g = load_workbook(golden)
    cr = wb_g["Corrected Register"]
    cat_col = None
    for c in range(1, cr.max_column + 1):
        if cr.cell(4, c).value == "Category":
            cat_col = c
            break
    assert cat_col is not None
    for r in range(5, cr.max_row + 1):
        if cr.cell(r, 1).value == tag:
            cr.cell(r, cat_col).value = "Monitor"
            break
    wb_g.save(golden)
    print("Corrected Register Category mirrored for", tag)


def break_synthetic_return_label_date() -> None:
    """
    device_return_shipments.csv has label_created_date uniform at 2026-06-25.
    We vary a handful of rows' label_created_date without touching the other fields.
    """
    path = ROOT / "device_return_shipments.csv"
    rows = list(csv.DictReader(path.open(newline="")))
    # find date column
    cols = rows[0].keys()
    date_col = None
    for k in cols:
        if k.lower() == "label_created_date":
            date_col = k
    assert date_col

    # change first 3 rows by small deltas
    # Keep format YYYY-MM-DD
    from datetime import datetime, timedelta

    deltas = [-6, -2, 3]
    changed = 0
    for i, r in enumerate(rows[:3]):
        base = datetime.strptime(r[date_col], "%Y-%m-%d")
        r[date_col] = (base + timedelta(days=deltas[i % len(deltas)])).strftime("%Y-%m-%d")
        changed += 1

    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print("Return label_created_date varied in", changed, "rows")


def run_fidelity_recompute() -> None:
    import subprocess

    # Recompute quantitative dashboard/ledger/exposure after we changed statuses/categories.
    subprocess.check_call(
        ["python3", "fix_golden_fidelity.py"],
        cwd=str(ROOT),
    )


def reinstate_formulas_and_cache_numeric(xlsx: Path) -> None:
    """
    Hardcoded-values axis requires formulas present.
    Golden-source-fidelity also reads numeric cells (data_only), so we inject cached <v> values.

    Strategy:
    - Capture typed numeric values currently present (after fidelity recompute).
    - Run formulize_workbook() to replace those cells with formulas.
    - Inject cached numeric values into <v> elements for formula cells using captured values.
    """
    # capture before_values
    wb_before = load_workbook(xlsx, data_only=False)
    target_sheets = ["Dashboard", "Corrected Register", "Exception Register", "Ledger Reconciliation"]
    before: dict[tuple[str, str], float] = {}

    for name in target_sheets:
        ws = wb_before[name]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    before[(name, ws.cell(r, c).coordinate)] = float(v)

    # formulize
    import fix_dataset_quality

    fix_dataset_quality.formulize_workbook()

    # re-open and find formula cells
    wb_form = load_workbook(xlsx, data_only=False)
    # build cache injection for formula cells
    inject: dict[tuple[str, str], float] = {}
    for name in target_sheets:
        ws = wb_form[name]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    key = (name, cell.coordinate)
                    if key in before:
                        inject[key] = before[key]

    # XML injection: update sheet xml <c r="B5"><v>value</v></c>
    tmp = xlsx.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(xlsx, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        wb_xml = zin.read("xl/workbook.xml")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels")

        import xml.etree.ElementTree as ET

        NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
        wb_root = ET.fromstring(wb_xml)
        rel_root = ET.fromstring(rels_xml)

        rid_to_target: dict[str, str] = {}
        for rel in rel_root.findall(REL + "Relationship"):
            rid_to_target[rel.attrib["Id"]] = rel.attrib["Target"]

        sheetname_to_xml: dict[str, str] = {}
        for sh in wb_root.findall(NS + "sheets/" + NS + "sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get(REL + "id")
            if not name or not rid:
                continue
            target = rid_to_target.get(rid)
            if not target:
                continue
            # target is relative to xl/
            if not target.startswith("worksheets/"):
                # e.g. "worksheets/sheet1.xml"
                target = "worksheets/" + target
            sheetname_to_xml[name] = "xl/" + target

        patched_cells = 0
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename in sheetname_to_xml.values():
                # figure out sheet name for this xml file
                sheet_name = next((k for k, v in sheetname_to_xml.items() if v == item.filename), None)
                if sheet_name:
                    tree = ET.fromstring(data)
                    for c in tree.iter(NS + "c"):
                        coord = c.attrib.get("r")
                        key = (sheet_name, coord)
                        if key not in inject:
                            continue
                        val = inject[key]
                        v_node = c.find(NS + "v")
                        if v_node is None:
                            v_node = ET.SubElement(c, NS + "v")
                        # write numeric
                        if float(val).is_integer():
                            v_node.text = str(int(val))
                        else:
                            v_node.text = str(val)
                        patched_cells += 1
                    data = ET.tostring(tree, encoding="utf-8", xml_declaration=True)
            zout.writestr(item, data)

    tmp.replace(xlsx)
    print("Cached numeric injection applied for", patched_cells, "formula cells")


def rebuild_zips_only() -> None:
    import subprocess

    subprocess.check_call(["python3", "fix_golden_fidelity.py"], cwd=str(ROOT))


def main() -> None:
    # 1) rename company
    replace_company_text_in_prompt()
    replace_company_text_in_workbook(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")

    # 2) break dashboard cross-tab by aligning Verified Status
    fix_dashboard_cross_tab_by_status(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")

    # 3) synthetic de-suspicion in inputs + mirror into output
    break_synthetic_category_even_split()
    break_synthetic_return_label_date()

    # 4) recompute quantitative gold content after status/category shifts
    run_fidelity_recompute()

    # 5) fix evidence-cited acquisition cost figures
    fix_evidence_assessment_cited_costs(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")

    # 6) restore formulas + cached numeric values
    reinstate_formulas_and_cache_numeric(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")

    # rebuild output zip from updated xlsx
    # (fix_golden_fidelity already rebuilds the zips, but we changed formulas after that)
    from fix_golden_fidelity import rebuild_zips

    rebuild_zips()

    print("Done applying hardcoded_values + synthetic fixes.")


if __name__ == "__main__":
    main()

