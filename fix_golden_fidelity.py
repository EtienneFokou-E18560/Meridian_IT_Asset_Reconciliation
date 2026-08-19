#!/usr/bin/env python3
"""Populate gold quantitative cells (data_only readers), tie POs to inventory, vary cost deltas."""
from __future__ import annotations

import csv
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent

# Varied ledger-over-register deltas (not a uniform $175).
DELTAS = {
    "MD-00017": 85,
    "MD-00034": 140,
    "MD-00051": 210,
    "MD-00068": 55,
    "MD-00085": 195,
    "MD-00102": 120,
    "MD-00119": 165,
}


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str) and v.startswith("="):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def setf(ws, row, col, formula, value, cache: dict) -> None:
    cell = ws.cell(row, col)
    cell.value = formula
    if value is not None and value != "":
        cache[(ws.title, cell.coordinate)] = value


def apply_formulas_with_cache(wb, assets, er, dash, lr, cache, n_assets, sum_m, sum_n, open_ex, blockers, missing, overdue, disposed, uniq_cc, sum_t, sum_n_active, n_match, n_mismatch, n_missing_fa) -> None:
    cr = wb["Corrected Register"]
    cr_a, cr_m, cr_n, cr_j, cr_i, cr_c, cr_t = (
        "'Corrected Register'!$A$5:$A$136",
        "'Corrected Register'!$M$5:$M$136",
        "'Corrected Register'!$N$5:$N$136",
        "'Corrected Register'!$J$5:$J$136",
        "'Corrected Register'!$I$5:$I$136",
        "'Corrected Register'!$C$5:$C$136",
        "'Corrected Register'!$T$5:$T$136",
    )
    er_k, er_mcol, er_b, er_g = (
        "'Exception Register'!$K$5:$K$167",
        "'Exception Register'!$M$5:$M$167",
        "'Exception Register'!$B$5:$B$167",
        "'Exception Register'!$G$5:$G$167",
    )
    for r, a in ((a["row"], a) for a in assets.values()):
        setf(cr, r, 14, f'=IF(V{r}="","",V{r})', a["n"], cache)
        if a["m"] is not None and a["t"] is not None:
            setf(cr, r, 23, f'=IF(OR(M{r}="",T{r}=""),"",T{r}-M{r})', round(a["t"] - a["m"], 2), cache)
        else:
            setf(cr, r, 23, f'=IF(OR(M{r}="",T{r}=""),"",T{r}-M{r})', None, cache)

    for r in range(5, er.max_row + 1):
        if not er.cell(r, 1).value:
            continue
        exp = er.cell(r, 7).value
        setf(
            er, r, 7,
            (
                f'=IF(B{r}="Inventory-to-Ledger Difference",'
                f'IF(IFERROR(VLOOKUP(D{r},{cr_a.replace("$A$5:$A$136","$A$5:$T$136")},20,FALSE),"")="",'
                f'IFERROR(VLOOKUP(D{r},\'Corrected Register\'!$A$5:$M$136,13,FALSE),0),'
                f'ABS(IFERROR(VLOOKUP(D{r},\'Corrected Register\'!$A$5:$T$136,20,FALSE),0)'
                f'-IFERROR(VLOOKUP(D{r},\'Corrected Register\'!$A$5:$M$136,13,FALSE),0))),'
                f'IFERROR(VLOOKUP(D{r},\'Corrected Register\'!$A$5:$N$136,14,FALSE),0))'
            ),
            exp, cache,
        )

    setf(dash, 5, 2, f"=COUNTA({cr_a})", n_assets, cache)
    setf(dash, 6, 2, f"=SUM({cr_m})", round(sum_m, 2), cache)
    setf(dash, 7, 2, f"=SUM({cr_n})", round(sum_n, 2), cache)
    setf(dash, 8, 2, f'=COUNTIF({er_k},"Open")', open_ex, cache)
    setf(dash, 9, 2, f'=COUNTIF({er_mcol},"Blocks certification")', blockers, cache)
    setf(dash, 10, 2, f'=COUNTIF({cr_j},"Missing")', missing, cache)
    setf(dash, 11, 2, f'=COUNTIF({cr_j},"Return Overdue")', overdue, cache)
    setf(dash, 12, 2, f'=COUNTIF({cr_j},"Disposed")', disposed, cache)
    setf(dash, 13, 2, "=COUNTA(UNIQUE(FILTER('Custody Chain'!A5:A501,'Custody Chain'!A5:A501<>\"\")))", uniq_cc, cache)

    for r in range(19, 27):
        if dash.cell(r, 1).value:
            setf(dash, r, 2, f"=COUNTIF({cr_j},A{r})", dash.cell(r, 2).value, cache)
            setf(dash, r, 3, f"=SUMIF({cr_j},A{r},{cr_n})", dash.cell(r, 3).value, cache)
    for r in range(19, 23):
        if dash.cell(r, 5).value:
            setf(dash, r, 6, f"=COUNTIF({cr_c},E{r})", dash.cell(r, 6).value, cache)
            setf(dash, r, 7, f"=SUMIF({cr_c},E{r},{cr_n})", dash.cell(r, 7).value, cache)
    for r in range(32, 44):
        if dash.cell(r, 1).value:
            setf(dash, r, 2, f"=COUNTIF({cr_i},A{r})", dash.cell(r, 2).value, cache)
            setf(dash, r, 3, f"=SUMIF({cr_i},A{r},{cr_n})", dash.cell(r, 3).value, cache)
    for r in range(32, 41):
        if dash.cell(r, 5).value:
            setf(dash, r, 6, f"=COUNTIF({er_b},E{r})", dash.cell(r, 6).value, cache)
            setf(dash, r, 7, f"=SUMIF({er_b},E{r},{er_g})", dash.cell(r, 7).value, cache)

    setf(lr, 5, 2, f"=SUM({cr_m})", round(sum_m, 2), cache)
    setf(lr, 6, 2, f"=SUM({cr_t})", round(sum_t, 2), cache)
    setf(lr, 7, 2, "=B5-B6", round(sum_m - sum_t, 2), cache)
    setf(lr, 8, 2, f"=SUM({cr_n})", round(sum_n, 2), cache)
    setf(lr, 9, 2, f'=SUMIF(\'Corrected Register\'!$U$5:$U$136,"Active",{cr_n})', round(sum_n_active, 2), cache)
    setf(lr, 10, 2, f"=SUM({cr_n})", round(sum_n, 2), cache)
    setf(lr, 11, 2, f"=COUNTA({cr_a})-B12-B13", n_match, cache)
    setf(lr, 12, 2, n_mismatch, n_mismatch, cache)  # count is a value; keep numeric to avoid COUNTIF string drift
    lr["B12"] = n_mismatch
    lr["B13"] = n_missing_fa
    cache[("Ledger Reconciliation", "B12")] = n_mismatch
    cache[("Ledger Reconciliation", "B13")] = n_missing_fa

    for r in range(24, 36):
        if not lr.cell(r, 1).value:
            continue
        setf(lr, r, 3, f"=IFERROR(VLOOKUP(A{r},'Corrected Register'!$A$5:$M$136,13,FALSE),\"\")", lr.cell(r, 3).value, cache)
        setf(lr, r, 4, f"=IFERROR(VLOOKUP(A{r},'Corrected Register'!$A$5:$T$136,20,FALSE),\"\")", lr.cell(r, 4).value, cache)
        setf(lr, r, 5, f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})', lr.cell(r, 5).value, cache)
        setf(lr, r, 6, f"=IFERROR(VLOOKUP(A{r},'Corrected Register'!$A$5:$N$136,14,FALSE),\"\")", lr.cell(r, 6).value, cache)
        setf(lr, r, 7, f"=IFERROR(VLOOKUP(A{r},'Corrected Register'!$A$5:$V$136,22,FALSE),\"\")", lr.cell(r, 7).value, cache)


def inject_cached_values(path: Path, cache: dict[tuple[str, str], float | int]) -> None:
    import xml.etree.ElementTree as ET

    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("", ns["m"])
    NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    tmp = path.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        wb_xml = ET.fromstring(zin.read("xl/workbook.xml"))
        rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {}
        for rel in rels:
            rid_to_target[rel.get("Id")] = rel.get("Target")
        name_to_xml = {}
        for sh in wb_xml.find(f"{NS}sheets"):
            name = sh.get("name")
            rid = sh.get(f"{REL_NS}id")
            target = rid_to_target[rid]
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            name_to_xml[name] = target

        by_sheet: dict[str, dict[str, float | int]] = {}
        for (sheet, coord), val in cache.items():
            by_sheet.setdefault(sheet, {})[coord] = val

        patched = set()
        for item in zin.infolist():
            data = zin.read(item.filename)
            xml_name = item.filename
            sheet_name = next((n for n, t in name_to_xml.items() if t == xml_name), None)
            if sheet_name and sheet_name in by_sheet:
                root = ET.fromstring(data)
                wanted = by_sheet[sheet_name]
                for c in root.iter(f"{NS}c"):
                    ref = c.get("r")
                    if ref not in wanted:
                        continue
                    fnode = c.find(f"{NS}f")
                    if fnode is None:
                        continue
                    val = wanted[ref]
                    vnode = c.find(f"{NS}v")
                    if vnode is None:
                        vnode = ET.SubElement(c, f"{NS}v")
                    if isinstance(val, float) and val.is_integer():
                        vnode.text = str(int(val))
                    else:
                        vnode.text = str(val)
                    if c.get("t") == "str":
                        del c.attrib["t"]
                    patched.add((sheet_name, ref))
                data = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
            zout.writestr(item, data)
    tmp.replace(path)
    print("cached values injected", len(patched), "of", len(cache))


def main() -> None:
    vary_ledger_costs()
    tie_po_amounts()
    populate_gold()
    rebuild_zips()
    verify()


def vary_ledger_costs() -> None:
    path = ROOT / "fixed_asset_ledger.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    headers = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    tag_c, cost_c = headers["Asset Tag"], headers["Acquisition Cost"]
    n = 0
    for r in range(5, ws.max_row + 1):
        tag = ws.cell(r, tag_c).value
        if tag in DELTAS:
            base = float(ws.cell(r, cost_c).value)
            # current ledger is inventory+175; reset from inventory via gold CR after, here: subtract 175 then add new delta
            # Safer: we'll set from inventory in populate_gold; here set inventory_cost + delta after reading inventory.
            n += 1
    wb.save(path)

    inv = load_workbook(ROOT / "it_asset_inventory.xlsx", data_only=True).active
    ih = {inv.cell(4, c).value: c for c in range(1, inv.max_column + 1)}
    inv_cost = {
        inv.cell(r, 1).value: float(inv.cell(r, ih["Acquisition Cost"]).value)
        for r in range(5, inv.max_row + 1)
        if inv.cell(r, 1).value
    }

    wb = load_workbook(path)
    ws = wb.active
    for r in range(5, ws.max_row + 1):
        tag = ws.cell(r, tag_c).value
        if tag in DELTAS and tag in inv_cost:
            ws.cell(r, cost_c).value = inv_cost[tag] + DELTAS[tag]
    wb.save(path)
    print("ledger deltas", {k: inv_cost[k] + v for k, v in DELTAS.items() if k in inv_cost})


def tie_po_amounts() -> None:
    inv = load_workbook(ROOT / "it_asset_inventory.xlsx", data_only=True).active
    ih = {inv.cell(4, c).value: c for c in range(1, inv.max_column + 1)}
    inv_cost = {
        inv.cell(r, 1).value: float(inv.cell(r, ih["Acquisition Cost"]).value)
        for r in range(5, inv.max_row + 1)
        if inv.cell(r, 1).value
    }
    path = ROOT / "hardware_purchase_orders.csv"
    rows = list(csv.DictReader(path.open(newline="")))
    for row in rows:
        tags = [t.strip() for t in row["asset_tags"].split(";") if t.strip()]
        row["approved_amount"] = str(int(round(sum(inv_cost[t] for t in tags))))
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print("PO amounts tied to inventory", [(r["purchase_order"], r["approved_amount"]) for r in rows[:3]])


def populate_gold() -> None:
    led = load_workbook(ROOT / "fixed_asset_ledger.xlsx", data_only=True).active
    lh = {led.cell(4, c).value: c for c in range(1, led.max_column + 1)}
    ledger = {}
    for r in range(5, led.max_row + 1):
        tag = led.cell(r, lh["Asset Tag"]).value
        if not tag:
            continue
        ledger[tag] = {
            "fa": led.cell(r, lh["Ledger Asset ID"]).value,
            "cost": num(led.cell(r, lh["Acquisition Cost"]).value),
            "nbv": num(led.cell(r, lh["Net Book Value"]).value),
            "status": led.cell(r, lh["Ledger Status"]).value,
        }

    wb = load_workbook(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")
    if "Source Ledger" in wb.sheetnames:
        src = wb["Source Ledger"]
        sh = {src.cell(4, c).value: c for c in range(1, src.max_column + 1)}
        for r in range(5, src.max_row + 1):
            tag = src.cell(r, sh.get("Asset Tag", 2)).value
            if tag in ledger:
                src.cell(r, sh["Acquisition Cost"]).value = ledger[tag]["cost"]
                src.cell(r, sh["Net Book Value"]).value = ledger[tag]["nbv"]

    cr = wb["Corrected Register"]
    assets = {}
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        m = num(cr.cell(r, 13).value)
        t = ledger[tag]["cost"] if tag in ledger else None
        v = ledger[tag]["nbv"] if tag in ledger else None
        cr.cell(r, 14).value = v  # Remaining Book Value
        if tag in ledger:
            cr.cell(r, 15).value = ledger[tag]["fa"]
            cr.cell(r, 20).value = t
            cr.cell(r, 21).value = ledger[tag]["status"]
            cr.cell(r, 22).value = v
        else:
            cr.cell(r, 15).value = None
            cr.cell(r, 20).value = None
            cr.cell(r, 22).value = None
        if m is not None and t is not None:
            cr.cell(r, 23).value = round(t - m, 2)
        else:
            cr.cell(r, 23).value = None
        cr.cell(r, 24).value = "Yes" if v is not None else "No"
        j = cr.cell(r, 10).value or ""
        exids = str(cr.cell(r, 18).value or "")
        cr.cell(r, 19).value = (
            "Yes"
            if j in ("Missing", "Return Overdue", "In Transit - Exception", "Retired/Pending Disposal")
            or "Critical" in exids
            else "No"
        )
        assets[tag] = {
            "row": r,
            "m": m,
            "t": t,
            "n": v,
            "status": cr.cell(r, 10).value,
            "loc": cr.cell(r, 9).value,
            "cat": cr.cell(r, 3).value,
        }

    er = wb["Exception Register"]
    for r in range(5, er.max_row + 1):
        tag = er.cell(r, 4).value
        typ = er.cell(r, 2).value
        rel = str(er.cell(r, 6).value or "")
        a = assets.get(tag, {})
        if typ == "Inventory-to-Ledger Difference":
            if a.get("t") is not None and a.get("m") is not None:
                exp = abs(a["t"] - a["m"])
            else:
                exp = a.get("m") or 0
            # keep related-record costs in sync
            if a.get("m") is not None:
                rel = re.sub(r"RegisterCost:[^;]*", f"RegisterCost:{a['m']}", rel)
            if a.get("t") is not None:
                rel = re.sub(r"LedgerCost:[^;]*", f"LedgerCost:{a['t']}", rel)
            elif "LedgerCost:" in rel:
                rel = re.sub(r"LedgerCost:[^;]*", "LedgerCost:missing", rel)
            er.cell(r, 6).value = rel
        else:
            exp = a.get("n") if a.get("n") is not None else 0
        er.cell(r, 7).value = round(float(exp), 2) if exp is not None else 0

        # refresh I2L action/assessment numbers if they still mention old 175-spread costs
        if typ == "Inventory-to-Ledger Difference" and tag in DELTAS:
            rc, lc = a.get("m"), a.get("t")
            act = str(er.cell(r, 8).value or "")
            act = re.sub(r"register [\d.]+ vs ledger [\d.]+", f"register {rc} vs ledger {lc}", act)
            act = re.sub(r"\([\d.]+ vs [\d.]+\)", f"({rc} vs {lc})", act)
            er.cell(r, 8).value = act

    dash = wb["Dashboard"]
    n_assets = len(assets)
    sum_m = sum(a["m"] or 0 for a in assets.values())
    sum_n = sum(a["n"] or 0 for a in assets.values() if a["n"] is not None)
    open_ex = sum(1 for r in range(5, er.max_row + 1) if er.cell(r, 11).value == "Open")
    blockers = sum(1 for r in range(5, er.max_row + 1) if er.cell(r, 13).value == "Blocks certification")
    missing = sum(1 for a in assets.values() if a["status"] == "Missing")
    overdue = sum(1 for a in assets.values() if a["status"] == "Return Overdue")
    disposed = sum(1 for a in assets.values() if a["status"] == "Disposed")
    cc = wb["Custody Chain"]
    uniq_cc = len({cc.cell(r, 1).value for r in range(5, cc.max_row + 1) if cc.cell(r, 1).value})

    dash["B5"] = n_assets
    dash["B6"] = round(sum_m, 2)
    dash["B7"] = round(sum_n, 2)
    dash["B8"] = open_ex
    dash["B9"] = blockers
    dash["B10"] = missing
    dash["B11"] = overdue
    dash["B12"] = disposed
    dash["B13"] = uniq_cc

    for r in range(19, 27):
        st = dash.cell(r, 1).value
        if not st:
            continue
        rows = [a for a in assets.values() if a["status"] == st]
        dash.cell(r, 2).value = len(rows)
        dash.cell(r, 3).value = round(sum(a["n"] or 0 for a in rows), 2)
    for r in range(19, 23):
        cat = dash.cell(r, 5).value
        if not cat:
            continue
        rows = [a for a in assets.values() if a["cat"] == cat]
        dash.cell(r, 6).value = len(rows)
        dash.cell(r, 7).value = round(sum(a["n"] or 0 for a in rows), 2)
    for r in range(32, 44):
        loc = dash.cell(r, 1).value
        if not loc:
            continue
        rows = [a for a in assets.values() if a["loc"] == loc]
        dash.cell(r, 2).value = len(rows)
        dash.cell(r, 3).value = round(sum(a["n"] or 0 for a in rows), 2)
    for r in range(32, 41):
        typ = dash.cell(r, 5).value
        if not typ:
            continue
        rs = [rr for rr in range(5, er.max_row + 1) if er.cell(rr, 2).value == typ]
        dash.cell(r, 6).value = len(rs)
        dash.cell(r, 7).value = round(sum(num(er.cell(rr, 7).value) or 0 for rr in rs), 2)

    lr = wb["Ledger Reconciliation"]
    sum_t = sum(a["t"] or 0 for a in assets.values() if a["t"] is not None)
    sum_n_active = sum(
        a["n"] or 0
        for tag, a in assets.items()
        if a["n"] is not None and ledger.get(tag, {}).get("status") == "Active"
    )
    n_mismatch = sum(1 for a in assets.values() if a["m"] is not None and a["t"] is not None and abs(a["t"] - a["m"]) > 0.01)
    n_missing_fa = sum(1 for a in assets.values() if a["t"] is None)
    n_match = n_assets - n_mismatch - n_missing_fa

    lr["B5"] = round(sum_m, 2)
    lr["B6"] = round(sum_t, 2)
    lr["B7"] = round(sum_m - sum_t, 2)
    lr["B8"] = round(sum_n, 2)
    lr["B9"] = round(sum_n_active, 2)
    lr["B10"] = round(sum_n, 2)
    lr["B11"] = n_match
    lr["B12"] = n_mismatch
    lr["B13"] = n_missing_fa

    combo = sum(DELTAS.values())
    delta_txt = ", ".join(f"{k} ${v}" for k, v in DELTAS.items())
    lr["A19"] = (
        f"3) Seven assets have ledger-over-register acquisition cost variances "
        f"({delta_txt}). Combined ledger cost excess = ${combo:,.0f}. "
        f"PO approved amounts tie to the inventory lines; Finance should align the FAR. "
        f"Per §9 these are High remediation items, not sole Critical blockers."
    )

    for r in range(24, 36):
        tag = lr.cell(r, 1).value
        if not tag:
            continue
        a = assets[tag]
        lr.cell(r, 3).value = a["m"]
        lr.cell(r, 4).value = a["t"] if a["t"] is not None else ""
        if a["m"] is not None and a["t"] is not None:
            lr.cell(r, 5).value = round(a["t"] - a["m"], 2)
        else:
            lr.cell(r, 5).value = ""
        lr.cell(r, 6).value = a["n"] if a["n"] is not None else ""
        lr.cell(r, 7).value = a["n"] if a["n"] is not None else ""
        note = lr.cell(r, 8).value
        if isinstance(note, str) and "$175" in note and tag in DELTAS:
            fa = ledger.get(tag, {}).get("fa") or ""
            po_m = re.search(r"(PO-[\d-]+)", note)
            po = po_m.group(1) if po_m else "the PO"
            lr.cell(r, 8).value = (
                f"Ledger {fa} cost exceeds register by ${DELTAS[tag]}. Validate against {po}."
            )

    cache: dict[tuple[str, str], float | int] = {}
    # Keep computed numbers (not unevaluated formulas). The fidelity checker
    # reads data_only/pandas and treats formula cells as blank.
    _ = cache

    # Evidence index policy version
    ei = wb["Evidence Index"]
    for r in range(1, ei.max_row + 1):
        v = ei.cell(r, 2).value
        if isinstance(v, str) and "v3.2" in v:
            ei.cell(r, 2).value = v.replace("v3.2", "v3.3")
        v1 = ei.cell(r, 1).value
        if isinstance(v1, str) and "v3.2" in v1:
            ei.cell(r, 1).value = v1.replace("v3.2", "v3.3")

    cert = wb["Certification"]
    for r in range(1, cert.max_row + 1):
        for c in range(1, 8):
            v = cert.cell(r, c).value
            if isinstance(v, str) and "$175" in v:
                cert.cell(r, c).value = v.replace("$175 cost-basis variances", "cost-basis variances").replace(
                    "seven $175 acquisition-cost variances",
                    "seven acquisition-cost variances (amounts differ by asset)",
                )

    # Methodology duplicate $175 if present
    if "Methodology" in wb.sheetnames:
        meth = wb["Methodology"]
        for r in range(1, meth.max_row + 1):
            for c in range(1, 6):
                v = meth.cell(r, c).value
                if isinstance(v, str) and "$175" in v:
                    meth.cell(r, c).value = v.replace("systematic $175 ", "").replace("$175 ", "")

    wb.save(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")
    print("gold quantitative cells populated")


def rebuild_zips() -> None:
    inputs = [
        "it_asset_inventory.xlsx",
        "hr_employee_status.csv",
        "equipment_transfer_log.csv",
        "service_desk_offboarding.csv",
        "device_return_shipments.csv",
        "hardware_purchase_orders.csv",
        "fixed_asset_ledger.xlsx",
        "asset_disposal_records.csv",
        "regional_IT_notes.docx",
        "IT_asset_management_policy.pdf",
        "ITAM_control_matrix.png",
        "receiving_exception_scan_1ZMD00000082.png",
    ]
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Inputs.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        for n in inputs:
            zf.write(ROOT / n, n)
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx", "Meridian_IT_Asset_Reconciliation.xlsx")
    print("zips rebuilt")


def verify() -> None:
    wb = load_workbook(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx", data_only=True)
    dash, cr, er, lr = wb["Dashboard"], wb["Corrected Register"], wb["Exception Register"], wb["Ledger Reconciliation"]

    def blanks(ws, col, start=5):
        vals = [ws.cell(r, col).value for r in range(start, ws.max_row + 1) if ws.cell(r, 1).value]
        none = sum(1 for v in vals if v is None or (isinstance(v, str) and v.startswith("=")))
        return len(vals), none, vals[:3]

    print("Dashboard B5-B13", [dash.cell(r, 2).value for r in range(5, 14)])
    print("CR NBV", blanks(cr, 14))
    print("CR cost diff", blanks(cr, 23))
    print("ER exposure", blanks(er, 7))
    print("LR amounts B5-B13", [lr.cell(r, 2).value for r in range(5, 14)])
    print("LR table C24", lr.cell(24, 3).value, "E24", lr.cell(24, 5).value)
    ei = wb["Evidence Index"]
    print("policy row", ei.cell(12, 2).value)

    # PO vs inventory vs ledger
    import csv as csvmod
    from openpyxl import load_workbook as lw

    inv = lw(ROOT / "it_asset_inventory.xlsx", data_only=True).active
    ih = {inv.cell(4, c).value: c for c in range(1, inv.max_column + 1)}
    inv_cost = {inv.cell(r, 1).value: float(inv.cell(r, ih["Acquisition Cost"]).value) for r in range(5, inv.max_row + 1)}
    led = lw(ROOT / "fixed_asset_ledger.xlsx", data_only=True).active
    lh = {led.cell(4, c).value: c for c in range(1, led.max_column + 1)}
    led_cost = {led.cell(r, 2).value: float(led.cell(r, lh["Acquisition Cost"]).value) for r in range(5, led.max_row + 1)}
    rows = list(csvmod.DictReader((ROOT / "hardware_purchase_orders.csv").open()))
    mismatches = 0
    for row in rows:
        tags = [t.strip() for t in row["asset_tags"].split(";") if t.strip()]
        inv_s = sum(inv_cost[t] for t in tags)
        led_s = sum(led_cost[t] for t in tags if t in led_cost)
        po = float(row["approved_amount"])
        if abs(po - inv_s) > 0.01:
            mismatches += 1
            print("PO!=inv", row["purchase_order"], po, inv_s)
        print(row["purchase_order"], "po", int(po), "inv", int(inv_s), "led", int(led_s), "po==inv", abs(po - inv_s) < 0.01)
    print("PO vs inventory mismatches", mismatches)

    # formula leftover census on key cols
    wb2 = load_workbook(ROOT / "Meridian_IT_Asset_Reconciliation.xlsx")
    fcount = 0
    for ws in (wb2["Dashboard"], wb2["Exception Register"], wb2["Ledger Reconciliation"], wb2["Corrected Register"]):
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    fcount += 1
    print("remaining formulas", fcount)


if __name__ == "__main__":
    main()
