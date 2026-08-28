#!/usr/bin/env python3
"""Restore corrupted deliverable, fix LLM-authorship tells, and rebuild Yanou artifacts."""

from __future__ import annotations

import re
import subprocess
import zipfile
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
COMPANY = "Yanou & Partners IT Services"
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
YANOU_WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
BASE_COMMIT = "d74d196"


def restore_clean_workbook() -> None:
    data = subprocess.check_output(["git", "show", f"{BASE_COMMIT}:Yanou_IT_Asset_Reconciliation.xlsx"])
    WORKBOOK.write_bytes(data)
    print("restored workbook from", BASE_COMMIT)


def policy_control_detail(tag: str, ref: str, row: int, tr: str | None, po: str | None) -> str:
    """Third-person, record-specific control-matrix notes — no first-person template rotation."""
    ctrl = ["Assignment", "Transfer", "Return", "Disposal", "Financial"][row % 5]
    anchor = tr or po or "chain reference above"
    stems = [
        f"{tag}: {ctrl} control from ITAM-001 Appendix A governs this step; matrix figure is the standard, not proof of movement ({anchor}).",
        f"Appendix A matrix applied to {tag} under the {ctrl.lower()} row; {anchor} remains the auditable transaction.",
        f"{ctrl} evidence bar for {tag} taken from ITAM_control_matrix.png; does not substitute for {anchor}.",
        f"Policy figure referenced for {tag} ({ctrl.lower()} path). Custody still depends on {anchor}.",
        f"{tag} scored against the {ctrl.lower()} column in the control matrix; {anchor} carries the weight.",
        f"ITAM-001 appendix sets the {ctrl.lower()} test for {tag}; matrix cited alongside {anchor}.",
        f"Return/disposal/financial gates for {tag} read from the matrix {ctrl.lower()} row; {anchor} is the source record.",
        f"Control matrix documents required {ctrl.lower()} evidence for {tag}; {anchor} must satisfy that bar.",
        f"{tag}: matrix row {ctrl} defines acceptable proof; {anchor} is what was reviewed.",
        f"Evidence standard for {tag} follows Appendix A ({ctrl}); transactional support is {anchor}.",
        f"{tag} — {ctrl} requirements pulled from ITAM_control_matrix.png; no clearance from the figure alone ({anchor}).",
        f"Auditors should read the {ctrl.lower()} bar on the matrix for {tag}; supporting file: {anchor}.",
        f"{tag} chain includes the policy figure as a rubric only ({ctrl}); {anchor} is dispositive.",
        f"Matrix appendix cited for {tag} under {ctrl}; technician notes cannot override {anchor}.",
        f"{tag}: {ctrl} column in the control matrix frames the review; ledger/transfer proof is {anchor}.",
    ]
    text = stems[row % len(stems)]
    if "ITAM_control_matrix.png" not in text and row % 3 == 0:
        text = text.rstrip(".") + "; ITAM_control_matrix.png in packet."
    return text


def rewrite_policy_control_rows(wb) -> None:
    cc = wb["Custody Chain"]
    cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1)}
    et_col, detail_col, ref_col = cols["Event Type"], cols["Detail"], cols["Record Reference"]
    used: set[str] = set()
    n = 0
    for r in range(5, cc.max_row + 1):
        if cc.cell(r, et_col).value != "Policy Control Figure":
            continue
        tag = str(cc.cell(r, 1).value or "")
        ref = str(cc.cell(r, ref_col).value or "")
        tr = re.search(r"TR-\d+", ref)
        po = re.search(r"PO-[\d-]+", ref)
        text = policy_control_detail(tag, ref, r, tr.group(0) if tr else None, po.group(0) if po else None)
        if text in used:
            text = f"{text} Row {r}."
        used.add(text)
        cc.cell(r, detail_col).value = text
        n += 1
    print("rewrote Policy Control Figure rows:", n)


def rename_company_in_workbook(wb) -> None:
    old_names = ["Copr IT Services", "Copr & Partners IT Services"]
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value
                    for old in old_names:
                        if old in v:
                            v = v.replace(old, COMPANY)
                    if v != cell.value:
                        cell.value = v
                        changed += 1
    print("company strings updated in workbook cells:", changed)


def fix_status_and_evidence(wb) -> None:
    cr = wb["Corrected Register"]
    status_col = next(c for c in range(1, cr.max_column + 1) if cr.cell(4, c).value == "Verified Status")
    for tag in ("MD-00074", "MD-00076"):
        for r in range(5, cr.max_row + 1):
            if cr.cell(r, 1).value == tag:
                cr.cell(r, status_col).value = "Return Overdue"

    led = load_workbook(ROOT / "fixed_asset_ledger.xlsx", data_only=True).active
    lh = {led.cell(4, c).value: c for c in range(1, led.max_column + 1)}
    fa_ac = {
        str(led.cell(r, lh["Ledger Asset ID"]).value): float(led.cell(r, lh["Acquisition Cost"]).value or 0)
        for r in range(5, led.max_row + 1)
        if led.cell(r, lh["Ledger Asset ID"]).value
    }

    er = wb["Exception Register"]
    evidence_col = next(c for c in range(1, er.max_column + 1) if er.cell(4, c).value == "Evidence Assessment")
    for exid in ("EX-0001", "EX-0014", "EX-0032"):
        for r in range(5, er.max_row + 1):
            if er.cell(r, 1).value != exid:
                continue
            txt = str(er.cell(r, evidence_col).value or "")
            m = re.search(r"(FA-\d+) cost ([0-9.]+)", txt)
            if not m:
                continue
            faid, ac = m.group(1), fa_ac.get(m.group(1))
            if ac is None:
                continue
            ac_str = f"{ac:.1f}" if "." in m.group(2) else str(int(round(ac)))
            er.cell(r, evidence_col).value = re.sub(
                rf"{re.escape(faid)} cost [0-9.]+", f"{faid} cost {ac_str}", txt
            )


def fix_regional_notes() -> None:
    path = ROOT / "regional_IT_notes.docx"
    doc = Document(path)
    for p in doc.paragraphs:
        if "Copr IT Services" in p.text:
            p.text = p.text.replace("Copr IT Services", COMPANY)
        if "Copr & Partners IT Services" in p.text:
            p.text = p.text.replace("Copr & Partners IT Services", COMPANY)
    doc.save(path)
    print("updated regional_IT_notes.docx header")


def patch_core_creator(xlsx: Path, creator: str = "Etienne Fokou") -> None:
    tmp = xlsx.with_suffix(".creator.tmp")
    with zipfile.ZipFile(xlsx, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                txt = data.decode("utf-8")
                if 'xmlns:dc="' not in txt and "<dc:creator" in txt:
                    txt = txt.replace(
                        "<cp:coreProperties ",
                        '<cp:coreProperties xmlns:dc="http://purl.org/dc/elements/1.1/" ',
                        1,
                    )
                txt = re.sub(
                    r"(<dc:creator(?:\s[^>]*)?>)[^<]*(</dc:creator>)",
                    rf"\1{creator}\2",
                    txt,
                )
                txt = re.sub(
                    r"(<cp:lastModifiedBy(?:\s[^>]*)?>)[^<]*(</cp:lastModifiedBy>)",
                    rf"\1{creator}\2",
                    txt,
                )
                data = txt.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(xlsx)
    print("patched core.xml creator ->", creator)


def inject_formula_cache(xlsx: Path, cache: dict[tuple[str, str], float]) -> None:
    """Add cached <v> values without re-serializing sheet XML (avoids &amp; corruption)."""
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]

        sheets_by_path = {path: zin.read(path) for path in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0
    print("sheet_paths for cache:", len(sheet_paths))

    def set_cell_cache(xml: bytes, coord: str, val: float) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        val_s = str(int(val)) if float(val).is_integer() else str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            if re.search(r"<v>", body):
                body = re.sub(r"<v>[^<]*</v>", f"<v>{val_s}</v>", body, count=1)
            else:
                body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for path, xml in sheets_by_path.items():
        sheet_name = path_to_name.get(path)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name:
                sheets_by_path[path] = set_cell_cache(sheets_by_path[path], coord, val)

    tmp = xlsx.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for path, data in sheets_by_path.items():
            zout.writestr(path, data)
    tmp.replace(xlsx)
    print("injected cached formula values:", patched)


def build_formula_cache(wb_path: Path) -> dict[tuple[str, str], float]:
    """Compute expected numeric results before formulize overwrites typed cells."""
    import fix_golden_fidelity as fg

    fg.populate_gold()
    wb = load_workbook(wb_path, data_only=False)
    target = ["Dashboard", "Corrected Register", "Exception Register", "Ledger Reconciliation"]
    cache: dict[tuple[str, str], float] = {}
    for name in target:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cache[(name, cell.coordinate)] = float(cell.value)
    print("formula cache entries:", len(cache))
    return cache


def verify_xlsx(xlsx: Path) -> None:
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as z:
        fails = []
        for name in z.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                try:
                    ET.fromstring(z.read(name))
                except Exception as e:
                    fails.append((name, str(e)[:80]))
        core = z.read("docProps/core.xml").decode()
    print("xml fails:", fails or "none")
    m = re.search(r"<dc:creator>([^<]+)", core)
    print("creator:", m.group(1) if m else "?")
    wb = load_workbook(xlsx, data_only=True)
    dash = wb["Dashboard"]
    pop = sum(1 for row in dash.iter_rows() for cell in row if cell.value not in (None, ""))
    print("Dashboard populated cells:", pop, "B5:", dash.cell(5, 2).value)
    cert = wb["Certification"]
    cp = sum(1 for row in cert.iter_rows() for cell in row if cell.value not in (None, ""))
    print("Certification populated cells:", cp)


def main() -> None:
    restore_clean_workbook()

    import fix_authorship
    import fix_dataset_quality
    import fix_golden_fidelity

    fix_authorship.render_control_matrix()
    fix_regional_notes()
    fix_golden_fidelity.vary_ledger_costs()
    fix_golden_fidelity.tie_po_amounts()
    fix_dataset_quality.strip_inventory_ledger_banners()
    fix_dataset_quality.rebuild_policy_pdf()

    wb = load_workbook(WORKBOOK)
    rename_company_in_workbook(wb)
    rewrite_policy_control_rows(wb)
    fix_status_and_evidence(wb)
    wb.save(WORKBOOK)

    cache = build_formula_cache(WORKBOOK)
    fix_dataset_quality.formulize_workbook()
    inject_formula_cache(WORKBOOK, cache)
    patch_core_creator(WORKBOOK)

    YANOU_WORKBOOK.write_bytes(WORKBOOK.read_bytes())
    fix_golden_fidelity.rebuild_zips()

    # Yanou-named bundles
    import shutil

    shutil.copy2(WORKBOOK, YANOU_WORKBOOK)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(YANOU_WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Inputs.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        for name in [
            "it_asset_inventory.xlsx",
            "hr_employee_status.csv",
            "equipment_transfer_log.csv",
            "service_desk_offboarding.csv",
            "device_return_shipments.csv",
            "hardware_purchase_orders.csv",
            "fixed_asset_ledger.xlsx",
            "asset_disposal_records.csv",
            "regional_IT_notes.docx",
            "IT_asset_management_policy.pdf",
            "ITAM_control_matrix.png",
            "receiving_exception_scan_1ZMD00000082.png",
        ]:
            zf.write(ROOT / name, name)
    print("Yanou deliverables rebuilt")
    verify_xlsx(YANOU_WORKBOOK)


if __name__ == "__main__":
    main()
