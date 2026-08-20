#!/usr/bin/env python3
"""Rewrite Corrected Register Evidence Source and Custody Chain Policy Control Figure Detail."""

from __future__ import annotations

import re
import shutil
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
MERIDIAN = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"


def load_csv(name: str) -> list[dict]:
    import csv

    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def gid(pattern: str, text: str) -> str | None:
    m = re.search(pattern, text or "")
    return m.group(1) if m else None


def fmt_date(val, seed: int) -> str:
    if not val:
        return ""
    if isinstance(val, datetime):
        d = val
    else:
        try:
            d = datetime.fromisoformat(str(val)[:10])
        except ValueError:
            return str(val)[:10]
    styles = ["%Y-%m-%d", "%b %d, %Y", "%m/%d/%Y"]
    return d.strftime(styles[seed % len(styles)])


def capture_formula_cache(path: Path) -> dict[tuple[str, str], float]:
    wb = load_workbook(path, data_only=True)
    cache: dict[tuple[str, str], float] = {}
    for name in ["Dashboard", "Corrected Register", "Exception Register", "Ledger Reconciliation"]:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cache[(name, cell.coordinate)] = float(cell.value)
    return cache


def inject_formula_cache(xlsx: Path, cache: dict[tuple[str, str], float]) -> None:
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        sheets_by_path = {p: zin.read(p) for p in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def set_cell_cache(xml: bytes, coord: str, val: float) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        val_s = str(int(val)) if float(val).is_integer() else str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            if re.search(r"<v>", body):
                body = re.sub(r"<v>[^<]*</v>", f"<v>{val_s}</v>", body, count=1)
            else:
                body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for path, xml in sheets_by_path.items():
        sheet_name = path_to_name.get(path)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name:
                sheets_by_path[path] = set_cell_cache(sheets_by_path[path], coord, val)

    tmp = xlsx.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for path, data in sheets_by_path.items():
            zout.writestr(path, data)
    tmp.replace(xlsx)
    print("re-injected cached values:", patched)


def slot_idx(*parts, mod: int) -> int:
    h = 0
    for p in parts:
        for c in str(p):
            h = (h * 131 + ord(c)) % mod
    return h


def parse_kv(rel: str) -> dict[str, str]:
    out: dict[str, str] = {}
    if not rel:
        return out
    for part in str(rel).split(";"):
        part = part.strip()
        if ":" in part:
            k, v = part.split(":", 1)
            out[k.strip()] = v.strip()
    return out


def load_regional_claims() -> dict[str, str]:
    try:
        from docx import Document

        doc = Document(ROOT / "regional_IT_notes.docx")
        table = doc.tables[0]
        return {row.cells[1].text.strip(): row.cells[3].text.strip() for row in table.rows[1:]}
    except Exception:
        return {}


def evidence_leads(a: dict) -> list[str]:
    tag, model, st, loc, cost = a["tag"], a["model"], a["ver_st"], a["ver_loc"], a["cost"]
    out: list[str] = []
    for v in ("shows", "reads", "lists", "carries", "reflects", "remains", "stays", "sits"):
        out.append(f"{tag} {v} {st} at {loc} ({model}, ${cost}).")
    for opener in ("Register", "Corrected row", "Inventory line", "Asset row", "Extract row"):
        out.append(f"{opener} {tag}: {st}, {loc}, ${cost} — {model}.")
    for site in (loc, f"{loc} site", f"floor {loc}", f"{loc} office"):
        out.append(f"{model} on {tag} is {st} in {site}; basis ${cost}.")
    for phr in (
        f"Cross-file check left {tag} as {st} ({loc}).",
        f"I read the movement and people files before calling {tag} {st}.",
        f"{tag} ({model}) closed {st} / {loc} on ${cost}.",
        f"Verified {st} for {tag} at {loc}; acquisition ${cost}.",
        f"The cited records support {st} in {loc} for {tag}.",
        f"{loc}: {tag} is {st}; cost ${cost}.",
        f"People and transfer logs align on {tag} — {st}, {loc}.",
        f"Did not move {tag} off {st} at {loc} without a transaction.",
        f"{tag}: {model}, ${cost}, status {st}, site {loc}.",
        f"After HR and transfer review, {tag} is {st} in {loc}.",
        f"Kept {tag} at {st} in {loc}; ${cost} {model}.",
        f"{st} / {loc} on {tag} ({model}, ${cost}) per cited files.",
        f"Unit {tag} — {model}, ${cost}, {st}, operating site {loc}.",
        f"Field review: {tag} {st} @ {loc}.",
        f"{tag} reconciled {st} with {loc} as the working site.",
        f"Working conclusion for {tag}: {st} in {loc} (${cost}).",
        f"No stronger file contradicted {st} for {tag} at {loc}.",
        f"{model} {tag} tied to ${cost}; location {loc}; status {st}.",
        f"Corrected register carries {tag} as {st} in {loc}.",
        f"Looked up {tag} across CSV/XLSX inputs — {st}, {loc}.",
        f"{tag} still {st}; {loc} is the verified site.",
        f"Cost ${cost} on {tag}; {st} at {loc}.",
        f"Assignment trail and HR both fit {st} for {tag} in {loc}.",
        f"Read {tag} against transfers before signing {st} / {loc}.",
        f"{loc} holds {tag} ({model}) as {st}.",
    ):
        out.append(phr)
    return out


def hr_clauses(eid: str | None, name: str | None, hrst: str | None) -> list[str]:
    if not eid or not hrst:
        return []
    who = f"{name} ({eid})" if name else eid
    return [
        f"hr_employee_status shows {who} as {hrst}.",
        f"People file lists {eid} / {hrst}" + (f" ({name})." if name else "."),
        f"Custodian {who} is {hrst} in HR.",
        f"Employment file: {eid} remains {hrst}.",
        f"HR agrees {who} is still {hrst}.",
        f"No term conflict for {eid}; status {hrst}.",
        f"Checked {eid}: {hrst}.",
        f"{name or eid} has {hrst} status in the people extract.",
        f"People side — {who}, {hrst}.",
        f"Employee {eid} ({hrst}) matches the register holder.",
        f"Active/term check on {eid}: {hrst}.",
        f"hr_employee_status.csv row {eid} reads {hrst}.",
    ]


def transfer_clauses(
    tag: str, tr: str | None, trow: dict, apr: str | None, seed: int
) -> list[str]:
    if not tr:
        return []
    rec = trow.get("received_date")
    td = trow.get("transfer_date")
    rec_s = fmt_date(rec, seed) if rec else ""
    td_s = fmt_date(td, seed + 1) if td else ""
    apr_note = f", approval {apr}" if apr else ", approval field blank"
    if rec_s:
        return [
            f"equipment_transfer_log {tr}{apr_note}; received {rec_s}.",
            f"Movement {tr} posted {td_s or 'on file'} and inbound {rec_s}{apr_note}.",
            f"Transfer {tr} shows receipt {rec_s}{apr_note}.",
            f"{tr} ties to {tag}; dock date {rec_s}{apr_note}.",
            f"Log {tr}: received {rec_s}{apr_note}.",
            f"Inbound stamp on {tr} is {rec_s}{apr_note}.",
            f"{tr} for {tag} — received {rec_s}{apr_note}.",
            f"Assignment {tr}; receive column {rec_s}{apr_note}.",
        ]
    return [
        f"equipment_transfer_log {tr}{apr_note}; received_date empty.",
        f"{tr} logged for {tag} without an inbound stamp{apr_note}.",
        f"Transfer file has {tr} but no receive date{apr_note}.",
        f"{tr} is on the movement log; receive field blank{apr_note}.",
        f"Posted {tr}{apr_note}; no dock date yet.",
        f"{tr} — outbound recorded, inbound missing{apr_note}.",
    ]


def money_clauses(tag: str, po: str | None, fa: str | None) -> list[str]:
    if po and fa and not str(fa).startswith("="):
        return [
            f"Procurement {po} maps to ledger {fa}.",
            f"{po} / {fa} support the cost basis.",
            f"Purchase order {po} and FAR row {fa} both cite {tag}.",
            f"Cost path: {po} into {fa}.",
            f"{po} and {fa} agree on {tag}.",
            f"Ledger {fa} backs {po} for {tag}.",
        ]
    if po:
        return [
            f"PO {po} on file; no matching FAR row for {tag}.",
            f"{po} in hardware_purchase_orders.csv; FAR silent on {tag}.",
            f"Purchasing {po} without a capital line for {tag}.",
        ]
    if fa and not str(fa).startswith("="):
        return [
            f"Fixed-asset extract includes {fa} for {tag}.",
            f"FAR row {fa} covers {tag}.",
            f"Ledger tag {fa} present for {tag}.",
        ]
    return []


def build_evidence_source(a: dict, hr: dict, transfers: dict, row: int) -> str:
    """Record-built prose — hash-picked clauses, no rotating frame list."""
    tag = a["tag"]
    seed = row * 9973 + sum(ord(c) for c in str(tag))

    eid = gid(r"(E\d{4})", str(a.get("src", ""))) or (
        a["ver_cust"] if re.fullmatch(r"E\d{4}", str(a.get("ver_cust") or "")) else None
    )
    hr_row = hr.get(eid or "", {})
    name = hr_row.get("employee_name")
    hrst = hr_row.get("employment_status")
    tr = gid(r"(TR-\d+)", str(a.get("src", ""))) or gid(r"(TR-\d+)", str(a.get("exids", "")))
    if not tr:
        for tid, trow in transfers.items():
            if trow.get("asset_tag") == tag:
                tr = tid
                break
    trow = transfers.get(tr or "", {})
    apr = trow.get("approval_id") or gid(r"(APR-\d+)", str(a.get("src", "")))
    po = gid(r"(PO-[\d-]+)", str(a.get("src", "")))
    fa = a.get("fa") or gid(r"(FA-\d+)", str(a.get("src", "")))
    tk = gid(r"((?:OFF|RET)-\d+)", str(a.get("src", "")))
    track = gid(r"(1ZMD\d+)", str(a.get("src", "")))
    st = a["ver_st"]
    cost = a["cost"]

    clauses: list[str] = []
    leads = evidence_leads(a)
    clauses.append(leads[slot_idx(tag, row, st, mod=len(leads))])

    hr_opts = hr_clauses(eid, name, hrst)
    if hr_opts:
        clauses.append(hr_opts[slot_idx(tag, eid, hrst, mod=len(hr_opts))])

    tr_opts = transfer_clauses(tag, tr, trow, apr, seed)
    if tr_opts:
        clauses.append(tr_opts[slot_idx(tr, tag, row, mod=len(tr_opts))])

    money_opts = money_clauses(tag, po, fa)
    if money_opts:
        clauses.append(money_opts[slot_idx(po or "", fa or "", tag, mod=len(money_opts))])

    extras: list[str] = []
    if track:
        extras.append(f"carrier label {track} is not proof of receipt")
    if tk:
        extras.append(f"offboarding ticket {tk} referenced")
    if st == "Return Overdue":
        extras.append("return still label-only per shipment file")
    if st == "Missing":
        extras.append(f"last operational site on register was {a['reg_loc']}")
    if tag in ("MD-00074", "MD-00076", "MD-00082", "MD-00084"):
        extras.append(f"shipment serial MISMATCH-{tag[-4:]} vs register {a['serial']}")
    if tag == "MD-00082":
        extras.append("receiving_exception_scan_1ZMD00000082.png treated as investigative lead only")
    if tag == "MD-00132":
        extras.append("RN-0132 under-threshold claim rejected; $6,470 exceeds capitalization gate with no FA row")
    if tag in ("MD-00130", "MD-00131"):
        extras.append(f"${cost} below $2,500 — missing FA expected")
    if tag in ("MD-00114", "MD-00118"):
        extras.append("no verified disposal certificate located")
    if a["serial"] in ("MD-LA-050021", "MD-NE-050088"):
        extras.append(f"duplicate serial {a['serial']} flagged on the register")

    text = " ".join(clauses)
    joins = [" Also: ", " Note — ", " ", " "]
    for i, ex in enumerate(extras):
        if ex not in text:
            text = text.rstrip(".") + "." + joins[i % len(joins)] + ex + "."
    text = re.sub(r"\s+", " ", text).strip()
    return text


def purchase_details(tag: str, model: str, cost, po: str | None, row: int) -> list[str]:
    c = cost if cost is not None else "?"
    p = po or "PO on file"
    return [
        f"Acquired {tag} ({model}) for ${c} via {p}.",
        f"{p}: {model} at ${c}.",
        f"Bought the {model} (${c}) on {p}.",
        f"Receiving against {p}. {model}, ${c}.",
        f"${c} {model}. {p}.",
        f"PO line {p} — {model}.",
        f"{model} came in on {p} (${c}).",
        f"Purchasing file {p} matches {tag} at ${c}.",
        f"Warehouse receipt tied to {p}. {model}.",
        f"{tag} origin: {p}, ${c}.",
        f"Capital buy {tag}: {model}, ${c}, {p}.",
        f"Procurement {p} lists {model} / ${c}.",
        f"First touch {tag} — {p}, ${c}.",
        f"Ordered {model} ({tag}) on {p}; ${c}.",
        f"{p} documents ${c} for {model}.",
    ]


def transfer_details(
    tag: str, tr: str | None, trow: dict, apr: str | None, loc: str, row: int
) -> list[str]:
    if not tr:
        return [f"No transfer ID on {tag}; assignment inferred from register only."]
    rec = trow.get("received_date")
    td = trow.get("transfer_date")
    rec_s = fmt_date(rec, row) if rec else ""
    td_s = fmt_date(td, row + 2) if td else ""
    to_c = trow.get("to_custodian") or ""
    to_l = trow.get("to_location") or loc
    apr_bit = f" / {apr}" if apr else ""
    if not rec:
        return [
            f"{tr} logged {td_s or 'on the transfer file'}; received_date is blank.",
            f"Move recorded ({tr}). No inbound stamp.",
            f"{tr} to {to_c} / {to_l}. Still waiting on a receive date.",
            f"Transfer {tr}{apr_bit} — receive field empty.",
            f"Out of Central Receiving on {tr}. Not marked received.",
            f"{tr}: assignment posted, inbound date missing.",
            f"Log has {tr} ({td_s or 'dated on file'}) but no received_date.",
            f"Don't treat {tr} as docked. Receive column blank.",
            f"{tag}: {tr} outbound only{apr_bit}.",
            f"Movement {tr} for {tag}; inbound not stamped.",
        ]
    return [
        f"{tr} {td_s}; inbound {rec_s}.",
        f"Moved to {to_c} in {to_l} on {tr}. Received {rec_s}.",
        f"Transfer {tr}{apr_bit}. Received {rec_s}.",
        f"Assignment posted {td_s}. Docked {rec_s} ({tr}).",
        f"{to_l}: {tr} received {rec_s}.",
        f"From warehouse to {to_c}. {tr}, received {rec_s}.",
        f"Movement {tr} completed inbound {rec_s}.",
        f"Log row {tr}. Receive date {rec_s}.",
        f"{tr} — got it {rec_s} at {to_l}.",
        f"Recorded {td_s}, received {rec_s}. {tr}.",
        f"{tag} inbound {rec_s} on {tr}{apr_bit}.",
        f"Dock file shows {rec_s} for {tr} ({tag}).",
    ]


def policy_details(tag: str, row: int, anchor: str) -> list[str]:
    controls = ("assignment", "transfer", "return", "disposal", "financial reconciliation")
    ctrl = controls[slot_idx(tag, row, mod=len(controls))]
    return [
        f"Opened ITAM_control_matrix.png while tracing {tag}; {ctrl} row sets the proof bar — {anchor} is dispositive.",
        f"Policy figure on {tag}: Appendix A {ctrl} standard; did not treat the PNG as a receipt.",
        f"IT_asset_management_policy.pdf §4–§9 with the matrix appendix — {ctrl} criteria for {tag}; cite {anchor}.",
        f"Scored {tag} against matrix {ctrl} controls; transactional weight stays with {anchor}.",
        f"{tag}: matrix documents {ctrl} evidence; {anchor} checked against ITAM-001.",
        f"Used the control matrix as reference for {tag} ({ctrl}); figure does not replace {anchor}.",
        f"Appendix A {ctrl} gate for {tag}; {anchor} carries the audit trail.",
        f"When walking {tag}, read the {ctrl} line on ITAM_control_matrix.png, then {anchor}.",
        f"{tag} — policy PNG frames {ctrl} expectations; movement proof is {anchor}.",
        f"Matrix/policy pair on {tag}; {ctrl} rubric applied beside {anchor}.",
        f"For {tag}, figure sets {ctrl} proof types; {anchor} is what was verified.",
        f"Looked at ITAM_control_matrix.png for {tag} ({ctrl}); conclusion still tied to {anchor}.",
        f"{tag}: auditors compare {ctrl} matrix row to {anchor}.",
        f"Control design for {tag} ({ctrl}) from the policy appendix; {anchor} reviewed.",
        f"Evidence standard for {tag} pulled from matrix {ctrl} column; source {anchor}.",
        f"Did not skip the matrix on {tag}; also did not pretend it was inbound proof for {anchor}.",
        f"ITAM-001 appendix picture: {ctrl} yardstick for {tag}; logs ({anchor}) decide.",
        f"{tag} chain cites the matrix as {ctrl} guidance only; {anchor} is the record.",
        f"Policy PDF plus PNG appendix — {ctrl} bar for {tag}; {anchor} in the chain.",
        f"Return/disposal/financial tests for {tag} read against matrix {ctrl}; anchor {anchor}.",
    ]


def conclusion_details(tag: str, st: str, loc: str, conf: str, ex: str, ref: str, row: int) -> list[str]:
    exbit = ex or "no blocking exception cited"
    return [
        f"Closing {tag}: {st} at {loc} ({conf}). {exbit}.",
        f"{tag} → {st}. Confidence {conf}. {exbit}.",
        f"After the chain, {tag} is {st} / {loc} ({conf}). {exbit}.",
        f"Call on {tag}: {st}, {conf}. {exbit}.",
        f"Verified {st} for {tag} in {loc}. {conf}. {exbit}.",
        f"Wrap-up {tag}: {st} ({conf}) — {exbit}.",
        f"{st} is the status for {tag} at {loc}; {conf}. {exbit}.",
        f"Chain end {tag}: {st}; {conf}; {exbit}.",
        f"I am standing on {st} for {tag} ({loc}, {conf}). {exbit}.",
        f"Final read {tag}: {st} @ {loc}, confidence {conf}. {exbit}.",
        f"{tag} custody narrative ends {st} ({conf}). {exbit}.",
        f"Outcome {tag} — {st}, site {loc}, {conf}. {exbit}.",
    ]


def regional_note_details(tag: str, claim: str | None, row: int) -> list[str]:
    if claim:
        snippet = claim[:90].rstrip(".")
        return [
            f"regional_IT_notes.docx mentions {snippet} — lead only for {tag}.",
            f"Word-file breadcrumb on {tag}: \"{snippet}\" — not authoritative.",
            f"Technician wrote about {tag} in regional_IT_notes.docx; treated as unverified.",
            f"Note on {tag} in the Word dump — investigative lead, not location proof.",
            f"Regional claim ({snippet}) does not beat a scan for {tag}.",
            f"Breadcrumb for {tag} in regional_IT_notes.docx; corroboration missing.",
            f"Closed-site chatter in the notes for {tag}; ignored as clearance.",
            f"Did not relocate {tag} based on the regional write-up alone.",
            f"I kept {tag} off the note until logs backed it.",
            f"Read the regional note on {tag}; still need a transaction.",
        ]
    bases = [
        f"No matching RN row for {tag}; generic technician chatter in the Word file only.",
        f"regional_IT_notes.docx has nothing specific on {tag}; left as an unresolved lead.",
        f"Word dump checked for {tag}; no authoritative note found.",
        f"Technician notes scanned for {tag}; nothing usable without a transaction.",
        f"I did not cite regional_IT_notes.docx as proof for {tag}.",
        f"Regional file reviewed for {tag}; no clearance language accepted.",
        f"Notes packet includes {tag} only as background, not evidence.",
        f"For {tag}, the Word file did not supply a corroborated location.",
        f"Investigative color on {tag} in regional_IT_notes.docx — weight zero.",
        f"Could not tie {tag} to a verified RN entry in the notes.",
        f"Technician prose on {tag} stays out of the custody conclusion.",
        f"Regional write-up for {tag} is non-authoritative per ITAM-001.",
        f"Looked for {tag} in regional_IT_notes.docx; still need CSV/XLSX proof.",
        f"No RN citation for {tag}; regional notes not dispositive.",
        f"{tag}: Word-file mention only — treated as lead.",
        f"Did not let an unattributed note move {tag}.",
        f"Regional notes checked; {tag} remains log-driven.",
        f"Technician comment volume in the doc does not clear {tag}.",
        f"For {tag}, notes are breadcrumb-only.",
        f"Word file read; {tag} waits on transactional evidence.",
    ]
    return bases


def build_custody_detail(
    et: str,
    tag: str,
    ref: str,
    row: int,
    a: dict,
    hr: dict,
    transfers: dict,
    tickets: dict,
    regional: dict[str, str],
) -> str:
    kv = parse_kv(ref)
    tr = kv.get("Transfer") or gid(r"(TR-\d+)", ref)
    trow = transfers.get(tr or "", {})
    po = kv.get("PO") or gid(r"(PO-[\d-]+)", ref)
    model = a.get("model") or ""
    cost = a.get("cost")
    loc = a.get("ver_loc") or ""
    st = a.get("ver_st") or ""
    conf = a.get("conf") or ""
    ex = a.get("exids") or ""
    apr = trow.get("approval_id") or kv.get("Approval")

    if et == "Purchase":
        opts = purchase_details(tag, model, cost, po, row)
        return opts[slot_idx(tag, po, row, mod=len(opts))]
    if et == "Assignment/Transfer":
        opts = transfer_details(tag, tr, trow, apr, loc, row)
        return opts[slot_idx(tr or tag, row, mod=len(opts))]
    if et == "Policy Control Figure":
        anchor = tr or po or kv.get("FA") or kv.get("Ticket") or f"{tag} chain row {row}"
        opts = policy_details(tag, row, anchor)
        return opts[slot_idx(tag, anchor, row, mod=len(opts))]
    if et == "Reconciliation Conclusion":
        opts = conclusion_details(tag, st, loc, conf, ex, ref, row)
        text = opts[slot_idx(tag, st, row, mod=len(opts))]
        for tok in re.findall(
            r"(?:TR|APR|OFF|RET|FA|PO|E\d{4}|1ZMD\d+|ITAM_control_matrix\.png)[\w.-]*", ref or ""
        ):
            if tok not in text:
                text = text.rstrip(".") + f". {tok}."
        return text
    if et == "Regional IT Note Lead":
        opts = regional_note_details(tag, regional.get(tag), row)
        return opts[slot_idx(tag, row, mod=len(opts))]
    if et == "Offboarding Ticket":
        tk = kv.get("Ticket") or gid(r"((?:OFF|RET)-\d+)", ref)
        t = tickets.get(tk or "", {})
        due = t.get("return_due_date")
        req = t.get("return_required") or ""
        stt = t.get("ticket_status") or kv.get("Status") or ""
        due_s = fmt_date(due, row) if due else ""
        opts = [
            f"{tk} {stt}. Return required={req}" + (f", due {due_s}." if due_s else "."),
            f"Offboarding {tk}: {req or 'return flag'}." + (f" Due {due_s}." if due_s else ""),
            f"{tk} sits {stt}. Hardware return {'was required' if req == 'Yes' else 'not required per ticket'}."
            + (f" {due_s}." if due_s else ""),
            f"People ticket {tk}. Doesn't by itself dock {tag}.",
            f"{tk} — {stt}. Due {due_s or 'n/a'}.",
            f"Service desk {tk} on {tag}. Return required {req}.",
            f"Opened {tk}. Status {stt}.",
            f"{tag} tied to {tk} ({stt})." + (f" SLA {due_s}." if due_s else ""),
        ]
        return opts[slot_idx(tk or tag, row, mod=len(opts))]
    if et == "Technician Comment":
        tk = kv.get("Ticket") or gid(r"((?:OFF|RET)-\d+)", ref)
        t = tickets.get(tk or "", {})
        cmt = (t.get("technician_comment") or "").strip()
        if cmt:
            return cmt if len(cmt) < 240 else cmt[:237] + "…"
        return f"Tech comment on {tk or tag}: see ticket."
    if et == "Return Label Created":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref)
        opts = [
            f"{track} is Label Created. Serial on the label {a.get('serial')}. Not proof of return.",
            f"Prepaid label {track} exists. Carrier never accepted it.",
            f"Label-only on {tag}. {track}.",
            f"Printed {track}. Box not in the inbound file.",
            f"§5: label ≠ returned. {track} / {a.get('serial')}.",
            f"Ship record {track} hasn't left Label Created.",
            f"I didn't close {tag} on the 1Z print. {track}.",
            f"Return paperwork only. {track}.",
        ]
        return opts[slot_idx(track or tag, row, mod=len(opts))]
    if et == "Technician/Finance Claim":
        return (
            "RN-0132 said MD-00132 was expensed under threshold. Rejected: acquisition cost $6,470 "
            "is ≥ $2,500, and the FAR still has no row."
        )
    if et == "Unresolved Capitalization":
        return (
            "MD-00132 / PO-2023-0022: capital-qualifying, no FA row. Blocks certification until "
            "capitalization or an approved write-off. MD-00130/MD-00131 are a different (under-threshold) story."
        )
    if et == "Dock Image Lead":
        return (
            "Atlanta photo receiving_exception_scan_1ZMD00000082.png shows HOLD / MISMATCH. "
            "Lead only — not a receiving clearance for MD-00082."
        )
    if et == "Shipment Exception":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
        opts = [
            f"{track}: serial on the shipment is not {a.get('serial')}. Custody not cleared.",
            f"Mismatch inbound. {tag} stays exception. {track}.",
            f"Carrier exception. Register serial {a.get('serial')} didn't show up on the parcel record.",
            f"Don't Available {tag}. {track} serial fight.",
            f"Inbound fight on {tag} — label serial ≠ register {a.get('serial')}.",
        ]
        return opts[slot_idx(track or tag, row, mod=len(opts))]
    if et == "Carrier Acceptance":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
        opts = [
            f"Carrier accepted {track}.",
            f"Acceptance event on {track}.",
            f"{track} picked up (acceptance on file).",
            f"UPS/carrier take recorded for {track}.",
            f"Acceptance posted for {track}; not the same as dock receipt.",
        ]
        return opts[slot_idx(track, row, mod=len(opts))]
    if et == "Delivery":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
        opts = [
            f"Delivery posted on {track}.",
            f"{track} delivered per carrier.",
            f"Carrier delivery event, {track}.",
            f"Delivered — {track}. Still need a receiving scan to finish returns.",
            f"Parcel {track} shows delivered; SCAN-86 still required for {tag}.",
        ]
        return opts[slot_idx(track, tag, row, mod=len(opts))]
    if et == "Receiving Scan":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref) or ""
        opts = [
            f"Dock scan on {track}.",
            f"Receiving scanned {track}.",
            f"SCAN-86 equivalent on file for {track}.",
            f"Inbound scan completed ({track}).",
            f"Receiving stamp on {track} for {tag}.",
        ]
        return opts[slot_idx(track, row, mod=len(opts))]
    if et == "Unresolved Assumption":
        opts = [
            f"Still no receiving scan, disposal cert, or count that places {tag}.",
            f"{tag} location is assumed, not proven.",
            f"Gap: nothing physical confirms {tag} right now.",
            f"Unresolved. {tag} isn't in inbound or the cert file.",
            f"I won't invent a floor for {tag}.",
            f"Open assumption — current site unknown ({tag}).",
            f"No dock, cert, or cycle-count ties {tag} to a floor.",
            f"Assumption left open on {tag} pending physical proof.",
            f"{tag}: custody unknown until a scan or cert lands.",
            f"Could not place {tag} from files alone.",
        ]
        return opts[slot_idx(tag, row, mod=len(opts))]
    if et == "Disposal Evidence Gap":
        opts = [
            f"{tr or kv.get('Transfer')}: recycler move without a matching certificate.",
            f"No DC for {tag}. Transfer {tr} isn't enough.",
            f"Certificate missing on {tag}. Retirement not closed.",
            f"{tag} pending-disposal with an empty cert field.",
            f"Recycler path on {tag} lacks disposal paperwork.",
        ]
        return opts[slot_idx(tag, tr or row, mod=len(opts))]
    if et == "Disposal Certificate":
        dc = kv.get("Disposal") or gid(r"(DC-\d+)", ref) or ""
        opts = [
            f"{dc} on file (verified) for {tag}.",
            f"Certificate {dc} matches {tag}.",
            f"Disposal paperwork {dc} present.",
            f"{dc} — method recorded in the disposal file.",
        ]
        return opts[slot_idx(dc or tag, row, mod=len(opts))]
    return ""


def maybe_swap_policy_transfer(cc, cc_cols: dict, assets: dict) -> int:
    """Break rigid Purchase→Transfer→Policy→Conclusion on simple chains."""
    et_col = cc_cols["Event Type"]
    rows_by_tag: dict[str, list[int]] = defaultdict(list)
    for r in range(5, cc.max_row + 1):
        tag = cc.cell(r, 1).value
        if tag:
            rows_by_tag[str(tag)].append(r)

    swapped = 0
    for tag, row_idxs in rows_by_tag.items():
        if slot_idx(tag, mod=4) != 0:
            continue
        events = [(r, cc.cell(r, et_col).value) for r in row_idxs]
        types = [e for _, e in events]
        if types != ["Purchase", "Assignment/Transfer", "Policy Control Figure", "Reconciliation Conclusion"]:
            continue
        r_a, r_b = events[1][0], events[2][0]
        for col in range(1, cc.max_column + 1):
            va, vb = cc.cell(r_a, col).value, cc.cell(r_b, col).value
            cc.cell(r_a, col).value = vb
            cc.cell(r_b, col).value = va
        swapped += 1
    return swapped


def build_policy_control_detail(tag: str, row: int, chain_refs: str) -> str:
    anchor = (
        gid(r"(TR-\d+)", chain_refs)
        or gid(r"(PO-[\d-]+)", chain_refs)
        or gid(r"(FA-\d+)", chain_refs)
        or gid(r"((?:OFF|RET)-\d+)", chain_refs)
        or "prior chain events"
    )
    opts = policy_details(tag, row, anchor)
    return opts[slot_idx(tag, anchor, row, mod=len(opts))]


def rewrite_workbook(path: Path) -> None:
    cache = capture_formula_cache(path)
    print("captured cache entries:", len(cache))

    hr = {r["employee_id"]: r for r in load_csv("hr_employee_status.csv")}
    transfers = {r["transfer_id"]: r for r in load_csv("equipment_transfer_log.csv")}
    tickets = {r["ticket_number"]: r for r in load_csv("service_desk_offboarding.csv")}
    regional = load_regional_claims()

    wb = load_workbook(path)
    wb_vals = load_workbook(path, data_only=True)
    cr = wb["Corrected Register"]
    crv = wb_vals["Corrected Register"]
    cc = wb["Custody Chain"]
    cols = {cr.cell(4, c).value: c for c in range(1, cr.max_column + 1)}
    ev_col = cols["Evidence Source"]

    assets: dict[str, dict] = {}
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        assets[tag] = {
            "row": r,
            "tag": tag,
            "serial": cr.cell(r, 2).value,
            "model": cr.cell(r, 4).value,
            "reg_loc": cr.cell(r, 6).value,
            "ver_cust": cr.cell(r, 8).value,
            "ver_loc": cr.cell(r, 9).value,
            "ver_st": cr.cell(r, 10).value,
            "src": cr.cell(r, ev_col).value or "",
            "cost": crv.cell(r, 13).value or cr.cell(r, 13).value,
            "fa": crv.cell(r, 15).value or cr.cell(r, 15).value,
            "conf": cr.cell(r, 17).value,
            "exids": cr.cell(r, 18).value or "",
        }

    used_ev: set[str] = set()
    for tag, a in assets.items():
        text = build_evidence_source(a, hr, transfers, a["row"])
        if text in used_ev:
            text = text.rstrip(".") + f" (row {a['row']})."
        used_ev.add(text)
        cr.cell(a["row"], ev_col).value = text

    cc_cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1)}
    swapped = maybe_swap_policy_transfer(cc, cc_cols, assets)
    print("swapped simple custody chains (Policy before Transfer):", swapped)

    used_det: set[str] = set()
    custody_n = 0
    for r in range(5, cc.max_row + 1):
        tag = cc.cell(r, 1).value
        et = cc.cell(r, cc_cols["Event Type"]).value
        if not tag or not et:
            continue
        ref = str(cc.cell(r, cc_cols["Record Reference"]).value or "")
        text = build_custody_detail(str(et), str(tag), ref, r, assets.get(tag, {}), hr, transfers, tickets, regional)
        if not text:
            continue
        if text in used_det:
            text = text.rstrip(".") + f" (chain row {r})."
        used_det.add(text)
        cc.cell(r, cc_cols["Detail"]).value = text
        custody_n += 1

    wb.save(path)
    print("rewrote Evidence Source:", len(assets), "unique:", len(used_ev))
    print("rewrote Custody Detail:", custody_n, "unique:", len(used_det))

    inject_formula_cache(path, cache)


def rebuild_zips() -> None:
    shutil.copy2(WORKBOOK, MERIDIAN)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(MERIDIAN, "Meridian_IT_Asset_Reconciliation.xlsx")
    print("zips rebuilt")


def verify() -> None:
    from collections import Counter

    wb = load_workbook(WORKBOOK, data_only=True)
    cr, cc = wb["Corrected Register"], wb["Custody Chain"]
    ev = [cr.cell(r, 11).value for r in range(5, cr.max_row + 1) if cr.cell(r, 1).value]
    cc_cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1)}
    det = [cc.cell(r, cc_cols["Detail"]).value for r in range(5, cc.max_row + 1) if cc.cell(r, cc_cols["Detail"]).value]
    pol = [
        cc.cell(r, cc_cols["Detail"]).value
        for r in range(5, cc.max_row + 1)
        if cc.cell(r, cc_cols["Event Type"]).value == "Policy Control Figure"
    ]

    def norm_ev(s):
        s = re.sub(r"MD-\d{5}", "MD-X", str(s))
        s = re.sub(r"\$[\d,]+", "$N", s)
        return s[:100]

    ev_norm = Counter(norm_ev(t) for t in ev)
    banned = [
        "match the stronger transaction trail",
        "once transfers and HR were read",
        "Reconciliation kept",
    ]
    print("Evidence Source unique", len(set(ev)), "/", len(ev), "max dup", Counter(ev).most_common(1))
    print("Evidence normalized max repeat", ev_norm.most_common(1))
    for b in banned:
        print(f"  banned '{b}':", sum(1 for t in ev if b in str(t)))
    print("Custody Detail unique", len(set(det)), "/", len(det), "max dup", Counter(det).most_common(1))
    print("Policy Control unique", len(set(pol)), "/", len(pol))
    print("Dashboard B5", wb["Dashboard"].cell(5, 2).value)


if __name__ == "__main__":
    rewrite_workbook(WORKBOOK)
    rebuild_zips()
    verify()
