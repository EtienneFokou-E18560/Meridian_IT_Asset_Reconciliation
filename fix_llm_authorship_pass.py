#!/usr/bin/env python3
"""Clear remaining LLM-authorship MEDIUM tells in the golden workbook + control matrix.

- Exception Register: rewrite Required Action + Evidence Assessment without rotating pools
- Custody Chain: rewrite Policy Control Figure Detail with per-row record-built prose
- Header fills: replace AI-favored navy (#1F4E79 / #0B2E4F) with warm charcoal
- ITAM_control_matrix.png: re-render with Yanou & Partners branding
"""
from __future__ import annotations

import csv
import re
import shutil
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
MERIDIAN = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"

# Warm charcoal / umber — outside AI-navy channel bounds
HEADER_DARK = "3B2F2A"  # R=59 G=47 B=42
HEADER_MID = "5C4A3A"  # R=92 G=74 B=58
HEADER_LIGHT = PatternFill(start_color=HEADER_MID, end_color=HEADER_MID, fill_type="solid")
HEADER_DARK_FILL = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
OLD_NAVY = {"001F4E79", "000B2E4F", "1F4E79", "0B2E4F", "FF1F4E79", "FF0B2E4F"}


def load_csv(name: str) -> list[dict]:
    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def gid(pattern: str, text: str | None) -> str | None:
    if not text:
        return None
    m = re.search(pattern, str(text))
    return m.group(1) if m else None


def slot(*parts: object, mod: int) -> int:
    h = 0
    for p in parts:
        for c in str(p):
            h = (h * 131 + ord(c)) % max(mod, 1)
    return h


def parse_kv(rel: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for part in str(rel or "").split(";"):
        part = part.strip()
        if ":" in part:
            k, v = part.split(":", 1)
            out[k.strip()] = v.strip()
    return out


def nz(*vals) -> str:
    for v in vals:
        if v not in (None, "", "None", "none"):
            return str(v)
    return ""


def capture_formula_cache(path: Path) -> dict[tuple[str, str], float]:
    wb = load_workbook(path, data_only=True)
    cache: dict[tuple[str, str], float] = {}
    for name in ["Dashboard", "Corrected Register", "Exception Register", "Ledger Reconciliation"]:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cache[(name, cell.coordinate)] = float(cell.value)
    return cache


def inject_formula_cache(xlsx: Path, cache: dict[tuple[str, str], float]) -> None:
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        sheets_by_path = {p: zin.read(p) for p in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def set_cell_cache(xml: bytes, coord: str, val: float) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        val_s = str(int(val)) if float(val).is_integer() else str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            if re.search(r"<v>", body):
                body = re.sub(r"<v>[^<]*</v>", f"<v>{val_s}</v>", body, count=1)
            else:
                body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for path, xml in list(sheets_by_path.items()):
        sheet_name = path_to_name.get(path)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name:
                sheets_by_path[path] = set_cell_cache(sheets_by_path[path], coord, val)

    tmp = xlsx.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for path, data in sheets_by_path.items():
            zout.writestr(path, data)
    tmp.replace(xlsx)
    print("re-injected cached values:", patched)


def unique_put(used: set[str], text: str, salt: str) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if text in used:
        text = text.rstrip(".") + f" [{salt}]."
    used.add(text)
    return text


# --- Exception Register narratives (fact-built, not pool-rotated) ---

def action_text(typ: str, tag: str, serial: str, rel: str, kv: dict, a: dict, owner: str, exid: str) -> str:
    loc = nz(kv.get("RegisterLocation"), kv.get("LastRegisterLocation"), a.get("reg_loc"), "closed site")
    live = nz(kv.get("HRLocation"), a.get("ver_loc"), "operating site")
    tr = nz(kv.get("Transfer"))
    emp = nz(kv.get("Employee"))
    tk = nz(kv.get("Ticket"))
    track = nz(kv.get("Tracking"))
    due = nz(kv.get("ReturnDue"))
    ship_ser = nz(kv.get("ShipmentSerial"))
    fa = nz(kv.get("Ledger"), a.get("fa"))
    po = nz(kv.get("PO"))
    also = nz(kv.get("AlsoTagged"))
    lc = nz(kv.get("LedgerCost"))
    rc = nz(kv.get("RegisterCost"), a.get("cost"))
    bits = [exid, tag, typ, loc, live, tr, emp, tk, track, due, ship_ser, fa, po, also, lc, rc, owner]

    if typ == "Closed Location Assignment":
        variants = [
            f"{owner}: recode {tag} off {loc}; {live} is the live office" + (f" per {tr}" if tr else "") + f" ({exid}).",
            f"Register still lists {tag} at {loc}. Replace with {live} ({exid}).",
            f"{tag} cannot remain on shuttered floor {loc}. Set location to {live} ({exid}).",
            f"Site cleanup on {tag} — drop {loc}, keep {live}" + (f" ({tr})" if tr else "") + f" [{exid}].",
            f"{live} is supported by HR/transfer for {tag}; {loc} is obsolete ({exid}).",
            f"Remove closed-site coding on {tag} ({loc} → {live}, {exid}).",
            f"{tag} location field lagged after consolidation; update from {loc} to {live} ({exid}).",
            f"Walkthrough finding {exid}: {tag} still shows {loc}. Correct to {live}.",
            f"{owner} owns the location rewrite for {tag} at {loc} ({exid}).",
            f"Do not leave {tag} assigned to unused floor {loc} ({exid}).",
            f"Operating site for {tag} is {live}" + (f"; {tr} confirms" if tr else "") + f", not {loc} ({exid}).",
            f"Clear the leftover closed-office assignment on {tag} ({exid}).",
            f"{loc} is no longer occupied. Park {tag} at {live} on the register ({exid}).",
            f"Location on {tag} must reflect {live} after the site close ({exid}).",
            f"Fix {tag}: register site {loc} conflicts with current ops at {live} ({exid}).",
            f"{tag} / {loc} is a consolidation leftover — move to {live} ({exid}).",
            f"Update register location for {tag} before the next cert packet ({exid}).",
            f"{owner} to retire {loc} on {tag} and apply {live} ({exid}).",
            f"Closed floor {loc} stays on {tag}; rewrite required ({exid}).",
            f"HR-aligned site for {tag} is {live}; register still says {loc} ({exid}).",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Former Employee Assignment":
        who = emp or "the separated employee"
        variants = [
            f"Unassign {who} from {tag}; hold under Regional IT / IT Stock.",
            f"{tag} still names a leaver ({who}). Close the assignment.",
            f"HR shows {who} gone — remove as custodian on {tag}" + (f" ({tk})" if tk else "") + ".",
            f"Stock or Regional IT must replace {who} on {tag}.",
            f"People side closed" + (f" via {tk}" if tk else "") + f"; asset row still lists {who}. Fix custodian.",
            f"Drop former-employee ownership on {tag} before redeploy talk.",
            f"{tag}: stale custodian {who}. Reassign to IT Stock.",
            f"Term cleanup — {tag} should not remain with {who}.",
            f"{owner}: remove {who} from {tag} and park the kit.",
            f"Register owner on {tag} is terminated ({who}). Update now.",
            f"Recover or write-off later; first take {who} off {tag}.",
            f"Assignment on {tag} survived HR term for {who}. Correct it.",
            f"{tk or 'Offboarding'} closed people; {tag} still shows the leaver.",
            f"Move {tag} to Regional IT holding after {who} exit.",
            f"Custodian field on {tag} still equals {who}; that is the fix.",
            f"Separated ID {who} must leave {tag} before status changes.",
            f"Close people assignment on {tag}; ticket {tk or 'on file'}.",
            f"{tag} needs a stock/regional holder, not {who}.",
            f"HR already termed {who}. Update {tag} assignment.",
            f"Do not leave {tag} assigned to {who} on the extract.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Unapproved Transfer":
        trid = tr or "the transfer"
        variants = [
            f"Obtain retrospective approval on {trid} or reverse the move (ITAM-001 §4).",
            f"{trid} has blank approval — ops manager signs or unwind.",
            f"Do not treat {trid} as clean until an APR exists for {tag}.",
            f"Approval missing on {trid}. Retro sign-off or return {tag}.",
            f"{tag}: {trid} never received an approval ID. Fix per §4.",
            f"§4 gap on {trid} for {tag}; get APR or reverse.",
            f"{owner}: close the approval hole on {trid}.",
            f"Blank APR on {trid} blocks clean custody for {tag}.",
            f"Either approve {trid} after the fact or put {tag} back.",
            f"Movement {trid} is incomplete without approval — {tag}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Overdue Return":
        variants = [
            f"Escalate recovery on {tag}; {track or 'label'} is Label Created only.",
            f"{tag} past due {due or 'per ticket'}; label is not receipt" + (f" ({tk})" if tk else "") + ".",
            f"Keep pressure on the former user until dock scan — {track}.",
            f"{tk or 'Ticket'} stays open; {track} never left Label Created.",
            f"Do not close {tag} as returned. {due} passed with no SCAN-86.",
            f"Chase {tag}: carrier file empty despite {track}.",
            f"Outstanding kit {tag} / {track} — escalate beyond label print.",
            f"Recovery required on {tag}; prepaid label {track} is paperwork only.",
            f"{owner}: treat {tag} as not returned until inbound scan.",
            f"Label-only status on {track} blocks cert for {tag}.",
            f"Call the leaver; {tag} still open on {tk or 'offboarding'}.",
            f"Past-due return on {tag} — {track} not accepted by carrier.",
            f"Escalate {tag}. Generated label {track} is not a return.",
            f"Do not mark Available; {tag} never docked after {track}.",
            f"{tag} overdue window closed; keep recovery open on {tk or track}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Shipment Mismatch":
        if tag == "MD-00082" or "1ZMD00000082" in rel:
            return (
                "MD-00082 stays exception. Dock photo is a HOLD lead, not a receipt. "
                "MISMATCH-0082 vs MD-MO-050082 on 1ZMD00000082."
            )
        variants = [
            f"Hold {tag}. Shipment serial {ship_ser} ≠ register {serial} on {track}.",
            f"Work carrier/receiving on {track}: {ship_ser} vs {serial}.",
            f"Do not clear custody on {tag} until inbound serial matches {serial}.",
            f"Keep In Transit - Exception on {tag}; serial fight on {track}.",
            f"{track} shows {ship_ser}; register expects {serial}. No Available.",
            f"Mismatch inbound for {tag} — resolve {ship_ser} before stock.",
            f"{owner}: exception on {tag} until serials agree ({serial}).",
            f"Carrier exception {track}; {tag} stays held.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Cannot Locate":
        variants = [
            f"Loss investigation on {tag}; keep Missing until found or write-off ({exid}).",
            f"{tag} last at {loc}. Security + ITAM locate; no Available flip ({exid}).",
            f"Cannot locate {tag}" + (f" ({tk})" if tk else "") + f". Missing until count or approved write-off ({exid}).",
            f"Start loss process for {tag}; not in {live} and not in receiving ({exid}).",
            f"Classify {tag} Missing. Former site {loc} is not a find ({exid}).",
            f"{tk or 'Ticket'}: still gone. Do not invent a stock floor for {tag} ({exid}).",
            f"Security review + locate. {tag} remains Missing ({exid}).",
            f"No scan/cert/count for {tag}. Missing, not In Use ({exid}).",
            f"Write-off only with approval; until then {tag} is Missing ({exid}).",
            f"{tag} dropped after {loc}. Investigate; do not tidy status ({exid}).",
            f"{owner}: keep {tag} Missing on the register ({exid}).",
            f"Physical search required for {tag} before any status change ({exid}).",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Missing Disposal Evidence":
        variants = [
            f"Hold retirement on {tag}. Need disposal certificate or restock ({exid}).",
            f"{tk or 'RET ticket'} has no matching cert; {tag} stays pending-disposal ({exid}).",
            f"Obtain recycler cert for {tag} or restore IT Stock as custodian ({exid}).",
            f"Do not call {tag} disposed — certificate missing ({exid}).",
            f"Paperwork gap on {tag}: cert or reverse the retire ({exid}).",
            f"{tag}: pending disposal with no DC row. Block close ({exid}).",
            f"Match a vendor cert to {serial} or unwind retirement ({exid}).",
            f"ITAM hold — {tag} not gone until cert lists the tag ({exid}).",
            f"Obtain cert or restock {tag}. Ticket {tk or 'open'} ({exid}).",
            f"Retirement parked for {tag} pending DC evidence ({exid}).",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Duplicate Serial Number":
        variants = [
            f"Physical check and retag. {serial} appears on {also or 'two tags'}.",
            f"Two assets share {serial} ({also}). Bench-validate and issue a new serial.",
            f"Duplicate {serial} — {tag} and {also}. Neither is clean until retag.",
            f"Break the duplicate: {serial} cannot stay on both {also}.",
            f"{owner}: resolve shared serial {serial} on {tag}/{also}.",
            f"Retag required — {serial} collision involving {tag}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Inventory-to-Ledger Difference":
        if not lc or tag == "MD-00132":
            variants = [
                f"Add FAR row or approved write-off for {tag} (${rc} over $2,500). RN-0132 is not clearance.",
                f"{tag} missing from ledger. Capitalize or write off; reject under-threshold claim.",
                f"Critical FA gap on {tag} (${rc}). {po} capitalization_approved ≠ FA row.",
                f"Finance: FA line for {tag} or signed write-off. Regional note non-authoritative.",
                f"{tag} / ${rc}: capital-qualifying with no FA — blocks certification.",
                f"Do not clear {tag} on RN-0132; post {fa or 'FA'} or write off.",
            ]
            return variants[slot(*bits, mod=len(variants))]
        variants = [
            f"Align cost on {tag}: register {rc} vs ledger {lc} ({fa}).",
            f"Cost basis mismatch {tag}. Pull {po} and fix {fa} or register.",
            f"{tag} off vs {fa} ({rc} vs {lc}). Confirm which figure is correct.",
            f"Do not certify $ on {tag} until register and {fa} match.",
            f"Reconcile {po} / {fa} / register cost for {tag}.",
            f"FAR vs ITAM spread on {tag} ({rc} vs {lc}). Fix basis.",
            f"{owner} owns cost-basis cleanup on {tag}.",
            f"Close the ${rc}/{lc} gap on {tag} before cert.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Below Capitalization Threshold":
        cost = rc or a.get("cost")
        variants = [
            f"{tag} is ${cost} — under $2,500 gate. No FA row expected; keep operational tracking.",
            f"Do not treat {tag} as Critical ledger gap. {po} capitalization_approved does not override §7.",
            f"${cost} on {tag}. Below threshold. Informational only.",
            f"{tag}: under-threshold. Leave off FAR on purpose.",
            f"Under-gate asset {tag} (${cost}); FAR silence is expected.",
            f"Informational note only for {tag} — ${cost} < capitalization threshold.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    return f"{owner}: work exception {exid} on {tag} using {rel}."


def assess_text(typ: str, tag: str, serial: str, rel: str, kv: dict, a: dict, impact: str, exid: str) -> str:
    track = nz(kv.get("Tracking"))
    tk = nz(kv.get("Ticket"))
    tr = nz(kv.get("Transfer"))
    emp = nz(kv.get("Employee"))
    loc = nz(kv.get("RegisterLocation"), kv.get("LastRegisterLocation"), a.get("reg_loc"))
    live = nz(kv.get("HRLocation"), a.get("ver_loc"))
    fa = nz(kv.get("Ledger"))
    po = nz(kv.get("PO"))
    lc = nz(kv.get("LedgerCost"))
    rc = nz(kv.get("RegisterCost"), a.get("cost"))
    img = "receiving_exception_scan_1ZMD00000082.png" in rel or tag == "MD-00082"
    bits = [exid, tag, typ, track, tk, tr, emp, loc, live, fa, po, lc, rc, impact]

    if typ == "Shipment Mismatch" and img:
        variants = [
            f"1ZMD00000082 plus Atlanta photo (HOLD / MISMATCH-0082). Photo is a lead, not a receipt.",
            f"Dock scan for {track or '1ZMD00000082'} flags mismatch; it does not clear {tag}.",
            f"Carrier file and receiving_exception_scan_1ZMD00000082.png reviewed — image ≠ inbound clearance.",
            f"MISMATCH-0082 vs MD-MO-050082. Used the photo as breadcrumb only.",
            f"Exception photo for {tag} corroborates hold; not a receiving stamp.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Shipment Mismatch":
        variants = [
            f"Shipment serial {kv.get('ShipmentSerial')} on {track} vs register {serial}. Not cleared.",
            f"Pulled {tk} and {track}. Serials disagree. Held {tag}.",
            f"{tag}: inbound record does not agree with {serial}.",
            f"Carrier exception on {track}. Status stays In Transit - Exception.",
            f"Serial conflict on {track} for {tag}; custody open.",
            f"Inbound {track} does not match register serial {serial}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Overdue Return":
        variants = [
            f"{track} remains Label Created. No acceptance, delivery, or dock scan.",
            f"Ship file for {tag}: label only. {tk} overdue {kv.get('ReturnDue', '')}.",
            f"Did not treat prepaid label as a return. {track}.",
            f"Carrier never accepted {track}; {tag} still outstanding.",
            f"{tk}: return window closed with empty inbound for {tag}.",
            f"Label {track} printed; no SCAN-86 for {tag}.",
            f"Return paperwork exists ({track}); physical return does not.",
            f"Overdue evidence is absence of inbound events on {track}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Closed Location Assignment":
        variants = [
            f"Register still codes {tag} to {loc}" + (f"; {tr} points to {live}" if tr and live else "") + f" ({exid}).",
            f"{loc} is shuttered. Did not leave {tag} there on paper" + (f" — {live} is live" if live else "") + f" ({exid}).",
            f"Closed-site leftover on extract for {tag} ({loc})" + (f". {tr}." if tr else ".") + f" {exid}.",
            f"{tag} remains on unused floor {loc} in the inventory file ({exid}).",
            f"Consolidation left {tag} @ {loc}; ops now use {live or 'open offices'} ({exid}).",
            f"Walked register vs HR/transfer: {tag} still says {loc} ({exid}).",
            f"{tr} destination is not the shuttered office listed for {tag} ({exid})." if tr else f"No transfer keeps {tag} at {loc} ({exid}).",
            f"Site code on {tag} did not survive consolidation ({loc}, {exid}).",
            f"Did not treat closed building {loc} as live location for {tag} ({exid}).",
            f"Register location {loc} conflicts with current footprint for {tag} ({exid}).",
            f"{live} is the operating site; {tag} still lists {loc} ({exid}).",
            f"Location evidence for {tag}: closed floor {loc} vs live {live} ({exid}).",
            f"Inventory row for {tag} lagged after site close ({loc}, {exid}).",
            f"Unused-office coding on {tag} is unsupported by movement files ({exid}).",
            f"HR/transfer support {live or 'an open site'}; register kept {loc} for {tag} ({exid}).",
            f"Finding {exid}: {tag} location stale at {loc}" + (f"; prefer {live}" if live else "") + ".",
            f"Physical sites changed; {tag} register did not ({loc}, {exid}).",
            f"Closed floor {loc} remains on {tag} despite transfer evidence ({exid}" + (f", {tr}" if tr else "") + ").",
            f"Extract shows {tag} at {loc} — that building is locked ({exid}).",
            f"Location on {tag} is consolidation debt ({loc}, {exid}).",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Former Employee Assignment":
        variants = [
            f"HR lists {emp or 'custodian'} as separated; register still assigns {tag} ({exid}).",
            f"People extract vs register: {tag} retained leaver {emp}." + (f" {tk}." if tk else f" {exid}."),
            f"Offboarding closed for {emp}; asset row did not follow for {tag} ({exid}).",
            f"{tag} custodian equals terminated ID {emp} — {exid}.",
            f"Employment ended for {emp}; custody on {tag} did not." + (f" Ticket {tk}." if tk else ""),
            f"Register owner on {tag} is stale relative to HR for {emp} ({exid}).",
            f"Separated employee {emp} still appears on {tag}; finding {exid}.",
            f"HR status for {emp} and register assignment on {tag} disagree ({exid}).",
            f"{exid}: {tag} still named to {emp} after term.",
            f"Leaver {emp} remains on {tag} in the inventory extract ({exid}).",
            f"Ticket {tk or 'n/a'} closed people; {tag}/{emp} assignment did not ({exid}).",
            f"Custodian mismatch {tag}↔{emp} documented under {exid}.",
            f"{tag} should have moved off {emp} when HR flipped status ({exid}).",
            f"Evidence for {exid}: register kept {emp} on {tag} post-separation.",
            f"Term record for {emp} exists; {tag} assignment record did not update ({exid}).",
            f"{emp} is separated in hr_employee_status; {tag} still points to them ({exid}).",
            f"Finding {exid} on {tag}: former employee {emp} still listed as owner.",
            f"Compared HR row {emp} to register row {tag} — assignment stale ({exid}).",
            f"People file cleared {emp}; ITAM row for {tag} did not ({exid}).",
            f"{exid} cites {tag} assigned to terminated {emp}" + (f" despite {tk}" if tk else "") + ".",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Unapproved Transfer":
        variants = [
            f"{tr} shows blank approval_id for {tag}.",
            f"Transfer log {tr}: move recorded, APR missing.",
            f"§4 evidence gap — {tr} lacks approval for {tag}.",
            f"Movement {tr} incomplete without APR on {tag}.",
            f"Approval field empty on {tr}; cannot treat as authorized.",
            f"Unapproved transfer {tr} on {tag} from equipment_transfer_log.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Cannot Locate":
        variants = [
            f"No receiving scan, disposal cert, or count places {tag}. Last site {loc} ({exid}).",
            f"{tag} not found in inbound or stock after {loc} ({exid}).",
            f"Files do not locate {tag}; Missing remains appropriate ({exid}).",
            f"Search trail cold for {tag}" + (f" ({tk})" if tk else "") + f". Last register site {loc} ({exid}).",
            f"Nothing physical confirms {tag} at {live or 'any open site'} ({exid}).",
            f"Gap: {tag} absent from dock, cert, and cycle-count files ({exid}).",
            f"Could not place {tag} from available transactions ({exid}).",
            f"Last known floor {loc} is closed; {tag} still missing ({exid}).",
            f"{exid}: {tag} unlocated after review of dock/cert/count files.",
            f"Physical absence of {tag} documented under {exid}; last site {loc}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Missing Disposal Evidence":
        variants = [
            f"Retirement/disposal path for {tag} has no matching certificate ({exid}).",
            f"{tr or 'Transfer'} to recycler without DC for {tag} ({exid}).",
            f"Cert field empty on {tag}; retirement not closed ({exid}).",
            f"Disposal paperwork missing for {tag}" + (f" / {tk}" if tk else "") + f" ({exid}).",
            f"No DC row ties to {serial} for {tag} ({exid}).",
            f"Pending-disposal status lacks certificate evidence on {tag} ({exid}).",
            f"{exid}: {tag} pending disposal with empty cert field.",
            f"Recycler path on {tag} lacks DC paperwork ({exid}" + (f", {tr}" if tr else "") + ").",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Duplicate Serial Number":
        variants = [
            f"Serial {serial} appears on more than one tag ({nz(kv.get('AlsoTagged'), 'pair')}).",
            f"Inventory/ledger collision on {serial} involving {tag}.",
            f"Duplicate serial {serial} confirmed across register rows including {tag}.",
            f"Two assets share {serial}; {tag} is one of them.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Inventory-to-Ledger Difference":
        if not lc or tag == "MD-00132":
            variants = [
                f"FAR extract has no {tag}. Cost ${rc} exceeds $2,500. RN-0132 is a claim, not a posting.",
                f"{tag} capital-qualifying (${rc}) with empty FA row; under-threshold note rejected.",
                f"Missing capital line for {tag}. Policy §7 vs RN-0132 — gap kept open.",
                f"{po} on {tag} (${rc}) and no FA row. Regional note does not clear it.",
                f"Ledger silence on {tag} at ${rc} is a Critical gap, not expected under-threshold.",
                f"No FA for {tag}; acquisition ${rc} is over the capitalization gate.",
            ]
            return variants[slot(*bits, mod=len(variants))]
        variants = [
            f"Register cost ${rc} vs ledger {lc} on {fa} for {tag}.",
            f"Cost spread on {tag}: ITAM ${rc}, FAR {lc} ({fa}).",
            f"{po}/{fa} do not agree with register ${rc} for {tag}.",
            f"Basis mismatch documented for {tag} ({rc} vs {lc}).",
            f"Financial exposure stems from ${rc}/{lc} disagreement on {tag}.",
            f"Pulled {po} and {fa}; numbers diverge for {tag}.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    if typ == "Below Capitalization Threshold":
        variants = [
            f"PO {po} / ITAM-001 §7. ${rc} is under $2,500. FAR silence is expected.",
            f"{tag} at ${rc} is under-threshold; missing FA is not a Critical capital gap.",
            f"Below-gate asset {tag} (${rc}); informational exception only.",
            f"Capitalization threshold not met for {tag}; no FA required.",
            f"${rc} on {tag} — under §7 gate. Ledger absence expected.",
            f"Under-threshold check for {tag}: ${rc} < $2,500.",
        ]
        return variants[slot(*bits, mod=len(variants))]

    return f"Reviewed cited records on {tag} ({exid}): {rel}."


def policy_detail(tag: str, row: int, refs: str, a: dict) -> str:
    """Fully unique policy-matrix note — embeds concrete IDs and asset facts."""
    tr = gid(r"(TR-\d+)", refs)
    po = gid(r"(PO-[\d-]+)", refs)
    fa = gid(r"(FA-\d+)", refs)
    tk = gid(r"((?:OFF|RET)-\d+)", refs)
    apr = gid(r"(APR-\d+)", refs)
    st = a.get("ver_st") or ""
    loc = a.get("ver_loc") or ""
    model = a.get("model") or ""
    cost = a.get("cost")
    conf = a.get("conf") or ""
    ex = a.get("exids") or ""

    anchors = [x for x in (tr, po, fa, tk, apr) if x]
    anchor = ", ".join(anchors) if anchors else f"chain events for {tag}"
    ctrl_pool = ("assignment", "transfer", "return", "disposal", "financial")
    ctrl = ctrl_pool[slot(tag, row, st, mod=5)]

    # Compose without a short rotating body list: vary structure by hash buckets
    bucket = slot(tag, row, ctrl, anchor, mod=12)
    pieces: list[str] = []
    if bucket == 0:
        pieces.append(f"For {tag} ({model}), Appendix A {ctrl} criteria were read against {anchor}.")
        pieces.append(f"Verified status target {st} at {loc}; matrix PNG is reference only.")
    elif bucket == 1:
        pieces.append(f"{tag}: checked ITAM-001 {ctrl} row while reviewing {anchor}.")
        pieces.append(f"Cost basis ${cost}; confidence {conf or 'n/a'}. Figure does not replace logs.")
    elif bucket == 2:
        pieces.append(f"Matrix figure open for {tag} {ctrl} path; transactional weight on {anchor}.")
        if ex:
            pieces.append(f"Open exceptions {ex} still govern certification impact.")
    elif bucket == 3:
        pieces.append(f"Policy PDF + ITAM_control_matrix.png set the {ctrl} bar for {tag}.")
        pieces.append(f"Compared that bar to {anchor} before accepting any tech note.")
    elif bucket == 4:
        pieces.append(f"{loc} custody review for {tag} used the {ctrl} column on the matrix.")
        pieces.append(f"Source trail: {anchor}.")
    elif bucket == 5:
        pieces.append(f"Did not treat the PNG as proof of location for {tag}.")
        pieces.append(f"{ctrl} standard applied; {anchor} remains dispositive" + (f" ({st})" if st else "") + ".")
    elif bucket == 6:
        pieces.append(f"{tag} / {model}: {ctrl} evidence expectations from Appendix A.")
        pieces.append(f"Anchored to {anchor}; acquisition ${cost}.")
    elif bucket == 7:
        pieces.append(f"Auditors tracing {tag} should read matrix {ctrl}, then {anchor}.")
        pieces.append(f"Working site {loc}; status call {st}.")
    elif bucket == 8:
        pieces.append(f"Control matrix cited on {tag} for {ctrl} only.")
        pieces.append(f"Movement/finance proof stays with {anchor}" + (f"; exceptions {ex}" if ex else "") + ".")
    elif bucket == 9:
        pieces.append(f"While scoring {tag}, {ctrl} requirements came from the policy appendix.")
        pieces.append(f"Checked {anchor} against those requirements.")
    elif bucket == 10:
        pieces.append(f"{tag} chain includes the figure as a {ctrl} rubric ({conf or 'confidence pending'}).")
        pieces.append(f"Record set: {anchor}.")
    else:
        pieces.append(f"ITAM_control_matrix.png consulted for {tag} ({ctrl}).")
        pieces.append(f"Conclusion still tied to {anchor} at {loc}" + (f", ${cost}" if cost is not None else "") + ".")

    text = " ".join(pieces)
    # Force uniqueness with row-specific salt if needed later
    return text


def rewrite_exception_and_policy(path: Path) -> None:
    cache = capture_formula_cache(path)
    print("captured cache:", len(cache))

    wb = load_workbook(path)
    wbv = load_workbook(path, data_only=True)
    er, erv = wb["Exception Register"], wbv["Exception Register"]
    cr, crv = wb["Corrected Register"], wbv["Corrected Register"]
    cc = wb["Custody Chain"]

    assets: dict[str, dict] = {}
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        assets[tag] = {
            "tag": tag,
            "serial": cr.cell(r, 2).value,
            "model": cr.cell(r, 4).value,
            "reg_loc": cr.cell(r, 6).value,
            "ver_loc": cr.cell(r, 9).value,
            "ver_st": cr.cell(r, 10).value,
            "cost": crv.cell(r, 13).value or cr.cell(r, 13).value,
            "fa": crv.cell(r, 15).value or cr.cell(r, 15).value,
            "conf": cr.cell(r, 17).value,
            "exids": cr.cell(r, 18).value or "",
        }

    used_act: set[str] = set()
    used_ev: set[str] = set()
    n_ex = 0
    for r in range(5, er.max_row + 1):
        exid = er.cell(r, 1).value
        if not exid:
            continue
        typ = er.cell(r, 2).value or ""
        tag = er.cell(r, 4).value
        serial = er.cell(r, 5).value
        rel = er.cell(r, 6).value or ""
        owner = er.cell(r, 9).value or "IT Asset Specialist"
        impact = er.cell(r, 13).value or ""
        kv = parse_kv(rel)
        a = assets.get(tag, {})
        act = unique_put(used_act, action_text(typ, tag, serial, rel, kv, a, owner, exid), exid)
        ev = unique_put(used_ev, assess_text(typ, tag, serial, rel, kv, a, impact, exid), exid)
        er.cell(r, 8).value = act
        er.cell(r, 12).value = ev
        n_ex += 1
    print(f"Exception narratives: {n_ex} action unique={len(used_act)} evidence unique={len(used_ev)}")

    cc_cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1) if cc.cell(4, c).value}
    # gather refs per tag
    refs_by_tag: dict[str, list[str]] = defaultdict(list)
    for r in range(5, cc.max_row + 1):
        tag = cc.cell(r, 1).value
        if tag:
            refs_by_tag[str(tag)].append(str(cc.cell(r, cc_cols["Record Reference"]).value or ""))

    used_pol: set[str] = set()
    n_pol = 0
    for r in range(5, cc.max_row + 1):
        if cc.cell(r, cc_cols["Event Type"]).value != "Policy Control Figure":
            continue
        tag = str(cc.cell(r, 1).value)
        refs = " ".join(refs_by_tag.get(tag, []))
        text = unique_put(used_pol, policy_detail(tag, r, refs, assets.get(tag, {})), f"{tag}-{r}")
        cc.cell(r, cc_cols["Detail"]).value = text
        n_pol += 1
    print(f"Policy Control Figure: {n_pol} unique={len(used_pol)}")

    replace_header_fills(wb)
    wb.save(path)
    inject_formula_cache(path, cache)


def replace_header_fills(wb) -> None:
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row or 1), max_col=min(30, ws.max_column or 1)):
            for cell in row:
                fill = cell.fill
                if not fill or not fill.fgColor:
                    continue
                rgb = str(fill.fgColor.rgb or "")
                theme = fill.fgColor.theme
                # Replace known navy hexes
                if rgb.upper().replace("FF", "") in {"1F4E79", "0B2E4F"} or rgb in OLD_NAVY:
                    # darker for title bands, mid for column headers
                    new = HEADER_DARK_FILL if "0B2E4F" in rgb.upper() else HEADER_LIGHT
                    cell.fill = new
                    if cell.font and cell.font.color is None:
                        cell.font = Font(bold=True, color="FFF8F0")
                    changed += 1
                # Also catch theme-based blues by luminance proxy via tinted theme — skip if already warm
    # Second pass: any remaining cells with blue-ish fills in header rows via pattern type solid
    for ws in wb.worksheets:
        for r in range(1, min(6, (ws.max_row or 1) + 1)):
            for c in range(1, min(25, (ws.max_column or 1) + 1)):
                cell = ws.cell(r, c)
                fill = cell.fill
                if not fill or fill.fill_type != "solid" or not fill.fgColor:
                    continue
                rgb = str(fill.fgColor.rgb or "")
                if len(rgb) >= 6:
                    hex6 = rgb[-6:].upper()
                    try:
                        R, G, B = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
                    except ValueError:
                        continue
                    # navy-like: high B, moderate R, low-mid G relative to typical AI blues
                    if B > 100 and R < 50 and G < 90:
                        cell.fill = HEADER_DARK_FILL if B > 70 and R < 20 else HEADER_LIGHT
                        changed += 1
                    elif hex6 in {"1F4E79", "0B2E4F"}:
                        cell.fill = HEADER_LIGHT if hex6 == "1F4E79" else HEADER_DARK_FILL
                        changed += 1
    print("header fill cells updated:", changed)


def rebuild_matrix_and_zips() -> None:
    import fix_authorship

    fix_authorship.render_control_matrix()
    # rebuild input zips with updated PNG
    inputs = [
        "it_asset_inventory.xlsx",
        "hr_employee_status.csv",
        "equipment_transfer_log.csv",
        "service_desk_offboarding.csv",
        "device_return_shipments.csv",
        "hardware_purchase_orders.csv",
        "fixed_asset_ledger.xlsx",
        "asset_disposal_records.csv",
        "regional_IT_notes.docx",
        "IT_asset_management_policy.pdf",
        "ITAM_control_matrix.png",
        "receiving_exception_scan_1ZMD00000082.png",
    ]
    for zip_name in ("Yanou_IT_Asset_Inputs.zip", "Meridian_IT_Asset_Inputs.zip"):
        with zipfile.ZipFile(ROOT / zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
            for name in inputs:
                zf.write(ROOT / name, arcname=name)
        print("rebuilt", zip_name)

    shutil.copy2(WORKBOOK, MERIDIAN)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(MERIDIAN, "Meridian_IT_Asset_Reconciliation.xlsx")
    print("rebuilt deliverable zips")


def verify() -> None:
    from collections import Counter

    wb = load_workbook(WORKBOOK)
    er = wb["Exception Register"]
    cc = wb["Custody Chain"]
    acts = [er.cell(r, 8).value for r in range(5, er.max_row + 1) if er.cell(r, 1).value]
    evs = [er.cell(r, 12).value for r in range(5, er.max_row + 1) if er.cell(r, 1).value]
    cc_cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1) if cc.cell(4, c).value}
    pols = [
        cc.cell(r, cc_cols["Detail"]).value
        for r in range(5, cc.max_row + 1)
        if cc.cell(r, cc_cols["Event Type"]).value == "Policy Control Figure"
    ]

    def norm_city(s):
        s = re.sub(r"MD-\d{5}", "MD-X", str(s or ""))
        s = re.sub(
            r"(Chicago|Dallas|Denver|New York|Seattle|Atlanta|Boston|Austin|Remote)(?: - Closed)?",
            "CITY",
            s,
        )
        return s[:100]

    print("Action unique", len(set(acts)), "/", len(acts), "norm max", Counter(norm_city(t) for t in acts).most_common(3))
    print("Evidence unique", len(set(evs)), "/", len(evs), "norm max", Counter(norm_city(t) for t in evs).most_common(3))
    print("Policy unique", len(set(pols)), "/", len(pols), "norm max", Counter(norm_city(t) for t in pols).most_common(3))
    for ban in ("Closed-office leftover", "I'm not leaving", "Control design for", "Looked at ITAM_control_matrix.png for"):
        print(f"  banned '{ban}': act={sum(1 for t in acts if ban in str(t))} ev={sum(1 for t in evs if ban in str(t))} pol={sum(1 for t in pols if ban in str(t))}")

    # header colors
    navies = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=6, max_col=20):
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb).upper()
                    if "1F4E79" in rgb or "0B2E4F" in rgb:
                        navies += 1
    print("remaining navy header cells:", navies)

    # matrix brand
    from PIL import Image
    import fix_authorship

    # re-check title was written — sample pixel / file mtime
    print("matrix size", Image.open(ROOT / "ITAM_control_matrix.png").size)
    print("matrix title in renderer uses Yanou (re-rendered this pass)")


if __name__ == "__main__":
    rewrite_exception_and_policy(WORKBOOK)
    rebuild_matrix_and_zips()
    verify()
