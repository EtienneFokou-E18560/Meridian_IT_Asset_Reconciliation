#!/usr/bin/env python3
"""Fix golden workbook oracle fidelity: LR PO citations, formula caches, certification blockers."""

from __future__ import annotations

import csv
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Yanou_IT_Asset_Reconciliation.xlsx"
MERIDIAN = ROOT / "Meridian_IT_Asset_Reconciliation.xlsx"

MISMATCH_PO: dict[str, str] = {
    "MD-00017": "PO-2025-0004",
    "MD-00034": "PO-2024-0007",
    "MD-00051": "PO-2023-0010",
    "MD-00068": "PO-2022-0013",
    "MD-00085": "PO-2025-0016",
    "MD-00102": "PO-2024-0019",
    "MD-00119": "PO-2024-0024",
}

SHIPMENT_BLOCKERS = {"MD-00074", "MD-00076", "MD-00082", "MD-00084"}
DISPOSAL_BLOCKERS = {"MD-00114", "MD-00118"}


def load_csv(name: str) -> list[dict]:
    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def po_by_tag() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for row in load_csv("hardware_purchase_orders.csv"):
        po = row["purchase_order"]
        for tag in re.split(r"[;\s]+", row.get("asset_tags") or ""):
            tag = tag.strip()
            if tag.startswith("MD-"):
                mapping[tag] = po
        if row.get("asset_tag", "").startswith("MD-"):
            mapping[str(row["asset_tag"]).strip()] = po
    return mapping


def inject_formula_cache(xlsx: Path, cache: dict[tuple[str, str], float | int | str]) -> None:
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(xlsx, "r") as zin:
        wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb_root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rid and rid in rid_to_target:
                sheet_paths[name] = rid_to_target[rid]
        sheets_by_path = {p: zin.read(p) for p in sheet_paths.values()}
        other = {i.filename: zin.read(i.filename) for i in zin.infolist() if i.filename not in sheet_paths.values()}

    path_to_name = {v: k for k, v in sheet_paths.items()}
    patched = 0

    def set_cell_cache(xml: bytes, coord: str, val: float | int | str) -> bytes:
        nonlocal patched
        s = xml.decode("utf-8")
        if isinstance(val, float):
            val_s = str(int(val)) if float(val).is_integer() else str(round(val, 2))
        else:
            val_s = str(val)
        cell_re = rf'(<c r="{coord}"[^>]*>)(.*?)(</c>)'

        def repl(m: re.Match[str]) -> str:
            nonlocal patched
            head, body, tail = m.group(1), m.group(2), m.group(3)
            body = re.sub(r"<v[^>]*/>", "", body)
            body = re.sub(r"<v>[^<]*</v>", "", body)
            body = body + f"<v>{val_s}</v>"
            patched += 1
            return head + body + tail

        if re.search(cell_re, s, flags=re.DOTALL):
            s = re.sub(cell_re, repl, s, count=1, flags=re.DOTALL)
        return s.encode("utf-8")

    for path, xml in sheets_by_path.items():
        sheet_name = path_to_name.get(path)
        if not sheet_name:
            continue
        for (sn, coord), val in cache.items():
            if sn == sheet_name and val is not None and val != "":
                sheets_by_path[path] = set_cell_cache(sheets_by_path[path], coord, val)

    tmp = xlsx.with_suffix(".cache.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other.items():
            zout.writestr(name, data)
        for path, data in sheets_by_path.items():
            zout.writestr(path, data)
    tmp.replace(xlsx)
    print(f"re-injected {patched} cached values into {xlsx.name}")


def fix_lr_po_explanations(wb) -> int:
    lr = wb["Ledger Reconciliation"]
    n = 0
    for r in range(24, 36):
        tag = lr.cell(r, 1).value
        if not tag or str(tag) not in MISMATCH_PO:
            continue
        po = MISMATCH_PO[str(tag)]
        expl = str(lr.cell(r, 8).value or "")
        expl = re.sub(r"PO-[\d-]+", po, expl, count=1)
        if "validate" in expl.lower() or "Validate" in expl:
            expl = re.sub(
                r"(validate (?:against )?)PO-[\d-]+",
                rf"\1{po}",
                expl,
                flags=re.I,
            )
        if po not in expl:
            expl = re.sub(
                r"\$[\d,]+(?=\.)",
                lambda m: m.group(0),
                expl,
            )
            expl = expl.rstrip(".") + f". Finance should validate {po}."
        lr.cell(r, 8).value = expl
        n += 1
    return n


def fix_dashboard_category_formulas(wb) -> None:
    dash = wb["Dashboard"]
    for r in range(19, 23):
        dash.cell(r, 6).value = (
            f"=COUNTIF('Source Inventory'!$C$5:$C$200,E{r})"
        )
        dash.cell(r, 7).value = (
            f"=SUMPRODUCT(('Source Inventory'!$C$5:$C$200=E{r})*('Corrected Register'!$N$5:$N$200))"
        )
    for r in range(19, 27):
        dash.cell(r, 2).value = f"=COUNTIF('Corrected Register'!$J$5:$J$200,A{r})"
        dash.cell(r, 3).value = f"=SUMIF('Corrected Register'!$J$5:$J$200,A{r},'Corrected Register'!$N$5:$N$200)"
    for r in range(32, 44):
        if dash.cell(r, 1).value:
            dash.cell(r, 2).value = f"=COUNTIF('Corrected Register'!$I$5:$I$200,A{r})"
            dash.cell(r, 3).value = f"=SUMIF('Corrected Register'!$I$5:$I$200,A{r},'Corrected Register'!$N$5:$N$200)"
    for r in range(32, 42):
        if dash.cell(r, 5).value:
            dash.cell(r, 6).value = f"=COUNTIF('Exception Register'!$B$5:$B$200,E{r})"
            dash.cell(r, 7).value = f"=SUMIF('Exception Register'!$B$5:$B$200,E{r},'Exception Register'!$G$5:$G$200)"


def fix_evidence_xlookup(wb) -> int:
    """Replace formula strings leaked into Evidence Source with resolved NBV amounts."""
    cr = wb["Corrected Register"]
    sl = wb["Source Ledger"]
    nbv_by_tag: dict[str, float] = {}
    for r in range(5, sl.max_row + 1):
        tag = sl.cell(r, 2).value
        if tag:
            acq = float(sl.cell(r, 5).value or 0)
            dep = float(sl.cell(r, 6).value or 0)
            nbv_by_tag[str(tag)] = round(acq - dep, 2)

    n = 0
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        ev = str(cr.cell(r, 11).value or "")
        if "=XLOOKUP" not in ev and "XLOOKUP(" not in ev:
            continue
        nbv = nbv_by_tag.get(str(tag))
        if nbv is not None:
            ev = re.sub(
                r"net book value \$=XLOOKUP[^;.\n]+",
                f"net book value ${nbv}",
                ev,
            )
            ev = re.sub(r"\$=XLOOKUP[^;.\n]+", f"${nbv}", ev)
        else:
            ev = re.sub(r"\$=XLOOKUP[^;.\n]+", "", ev)
        cr.cell(r, 11).value = re.sub(r"\s+", " ", ev).strip()
        n += 1
    return n


def fix_critical_blocker_language(wb) -> None:
    er = wb["Exception Register"]
    for r in range(5, er.max_row + 1):
        tag = str(er.cell(r, 4).value or "")
        typ = str(er.cell(r, 2).value or "")
        impact = str(er.cell(r, 3).value or "")
        if impact != "Critical":
            continue
        assess = str(er.cell(r, 12).value or "")
        action = str(er.cell(r, 8).value or "")
        prefix = "Critical certification blocker — "
        if tag in DISPOSAL_BLOCKERS and typ == "Missing Disposal Evidence":
            if prefix not in assess:
                er.cell(r, 12).value = (
                    f"{prefix}{tag} is Retired/Pending Disposal without a verified disposal certificate. "
                    + assess
                )
            if prefix not in action:
                er.cell(r, 8).value = prefix + action
        if tag in SHIPMENT_BLOCKERS and typ == "Shipment Mismatch":
            if prefix not in assess:
                er.cell(r, 12).value = (
                    f"{prefix}unresolved shipment serial mismatch on {tag}. " + assess
                )
            if prefix not in action:
                er.cell(r, 8).value = prefix + action

    cc = wb["Custody Chain"]
    cols = {cc.cell(4, c).value: c for c in range(1, cc.max_column + 1)}
    det_col = cols["Detail"]
    et_col = cols["Event Type"]
    cr = wb["Corrected Register"]
    exids_by_tag = {
        str(cr.cell(r, 1).value): str(cr.cell(r, 18).value or "")
        for r in range(5, cr.max_row + 1)
        if cr.cell(r, 1).value
    }
    for r in range(5, cc.max_row + 1):
        tag = str(cc.cell(r, 1).value or "")
        if cc.cell(r, et_col).value != "Reconciliation Conclusion":
            continue
        det = str(cc.cell(r, det_col).value or "")
        exids = exids_by_tag.get(tag, "")
        if tag in DISPOSAL_BLOCKERS and "Critical certification blocker" not in det:
            cc.cell(r, det_col).value = (
                f"Critical certification blocker — final status on {tag}: "
                f"Retired/Pending Disposal, site Disposed (certificate missing), Low. {exids}."
            )
        elif tag in SHIPMENT_BLOCKERS and "Critical certification blocker" not in det:
            cc.cell(r, det_col).value = (
                f"Critical certification blocker — {tag} held for unresolved shipment serial mismatch "
                f"(In Transit - Exception). {exids}."
            )

    cert = wb["Certification"]
    cert["A30"] = "Critical certification blockers (must clear before sign-off)"
    cert["A31"] = "Asset Tag"
    cert["B31"] = "Blocker basis"
    cert["C31"] = "Verified status / location"
    cert["D31"] = "Exception IDs"
    blockers = [
        ("MD-00074", "Unresolved shipment serial mismatch (1ZMD00000074)", "In Transit - Exception / Carrier Network", "EX-0076"),
        ("MD-00076", "Unresolved shipment serial mismatch (1ZMD00000076)", "In Transit - Exception / Carrier Network", "EX-0081"),
        ("MD-00082", "Unresolved shipment serial mismatch (1ZMD00000082)", "In Transit - Exception / Carrier Network", "EX-0092"),
        ("MD-00084", "Unresolved shipment serial mismatch (1ZMD00000084)", "In Transit - Exception / Carrier Network", "EX-0095"),
        ("MD-00114", "Retired/Pending Disposal without verified disposal certificate", "Retired/Pending Disposal / Disposed (certificate missing)", "EX-0149, EX-0150"),
        ("MD-00118", "Retired/Pending Disposal without verified disposal certificate", "Retired/Pending Disposal / Disposed (certificate missing)", "EX-0151, EX-0152"),
        ("MD-00132", "Capital-qualifying asset missing from fixed-asset ledger", "Pending Redeployment / Atlanta Warehouse", "EX-0163"),
    ]
    for i, (tag, basis, loc, exids) in enumerate(blockers, start=32):
        cert.cell(i, 1).value = tag
        cert.cell(i, 2).value = f"Critical certification blocker — {basis}"
        cert.cell(i, 3).value = loc
        cert.cell(i, 4).value = exids


def evaluate_formula_cache(wb) -> dict[tuple[str, str], float | int]:
    """Compute cached values for formula outputs (data_only readers / oracle tools)."""
    wbv = load_workbook(WORKBOOK, data_only=True)
    cr = wb["Corrected Register"]
    crv = wbv["Corrected Register"]
    sl = wb["Source Ledger"]
    slv = wbv["Source Ledger"]
    er = wb["Exception Register"]
    dash = wb["Dashboard"]
    lr = wb["Ledger Reconciliation"]
    si = wbv["Source Inventory"]
    cc = wb["Custody Chain"]

    sl_cost, sl_nbv, sl_fa, sl_status = {}, {}, {}, {}
    for r in range(5, 200):
        tag = sl.cell(r, 2).value
        if not tag:
            continue
        sl_cost[str(tag)] = float(slv.cell(r, 5).value or sl.cell(r, 5).value or 0)
        nbv_raw = slv.cell(r, 7).value
        if nbv_raw is None or (isinstance(nbv_raw, str) and str(nbv_raw).startswith("=")):
            acq = float(slv.cell(r, 5).value or sl.cell(r, 5).value or 0)
            dep = float(slv.cell(r, 6).value or sl.cell(r, 6).value or 0)
            nbv_raw = acq - dep
        sl_nbv[str(tag)] = float(nbv_raw or 0)
        sl_fa[str(tag)] = sl.cell(r, 1).value
        sl_status[str(tag)] = sl.cell(r, 8).value

    cache: dict[tuple[str, str], float | int] = {}
    nbv_by_tag: dict[str, float] = {}
    cost_by_tag: dict[str, float] = {}
    status_j: dict[str, str] = {}
    cat_c: dict[str, str] = {}
    loc_i: dict[str, str] = {}
    tags: list[str] = []

    for r in range(5, 200):
        tag = cr.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        tags.append(tag)
        cost_raw = crv.cell(r, 13).value
        if cost_raw is None or (isinstance(cost_raw, str) and str(cost_raw).startswith("=")):
            cost_raw = si.cell(r, 8).value
        cost_by_tag[tag] = float(cost_raw or 0)
        status_j[tag] = str(crv.cell(r, 10).value or cr.cell(r, 10).value or "")
        loc_i[tag] = str(crv.cell(r, 9).value or cr.cell(r, 9).value or "")
        cat = si.cell(r, 3).value or cr.cell(r, 3).value
        cat_c[tag] = str(cat or "")

        if tag in sl_nbv:
            nbv = round(sl_nbv[tag], 2)
            nbv_by_tag[tag] = nbv
            cache[("Corrected Register", f"N{r}")] = nbv
        else:
            nbv_by_tag[tag] = 0.0

    def nbv_sum(pred) -> float:
        return round(sum(nbv_by_tag[t] for t in tags if pred(t)), 2)

    def cnt(pred) -> int:
        return sum(1 for t in tags if pred(t))

    cache[("Dashboard", "B5")] = len(tags)
    cache[("Dashboard", "B6")] = round(sum(cost_by_tag.values()), 2)
    cache[("Dashboard", "B7")] = round(sum(nbv_by_tag.values()), 2)

    open_ex = crit = 0
    ex_type_count: dict[str, int] = defaultdict(int)
    ex_type_exp: dict[str, float] = defaultdict(float)
    for r in range(5, 200):
        if not er.cell(r, 1).value:
            continue
        if er.cell(r, 11).value == "Open":
            open_ex += 1
        if er.cell(r, 3).value == "Critical":
            crit += 1
        typ = str(er.cell(r, 2).value or "")
        ex_type_count[typ] += 1
        tag = str(er.cell(r, 4).value or "")
        if typ == "Inventory-to-Ledger Difference":
            if sl_status.get(tag) == "Disposed":
                g = sl_nbv.get(tag, 0)
            elif tag not in sl_cost:
                g = cost_by_tag.get(tag, 0)
            else:
                g = abs(sl_cost[tag] - cost_by_tag.get(tag, 0))
        else:
            g = nbv_by_tag.get(tag, 0)
        cache[("Exception Register", f"G{r}")] = round(float(g), 2)
        ex_type_exp[typ] += float(g)

    cache[("Dashboard", "B8")] = open_ex
    cache[("Dashboard", "B9")] = crit
    cache[("Dashboard", "B10")] = cnt(lambda t: status_j[t] == "Missing")
    cache[("Dashboard", "B11")] = cnt(lambda t: status_j[t] == "Return Overdue")
    cache[("Dashboard", "B12")] = cnt(lambda t: status_j[t] == "Disposed")
    custody_tags = {cc.cell(r, 1).value for r in range(5, 502) if cc.cell(r, 1).value}
    cache[("Dashboard", "B13")] = len(custody_tags)

    for r in range(19, 27):
        key = dash.cell(r, 1).value
        if not key:
            continue
        cache[("Dashboard", f"B{r}")] = cnt(lambda t, k=key: status_j[t] == k)
        cache[("Dashboard", f"C{r}")] = nbv_sum(lambda t, k=key: status_j[t] == k)

    for r in range(19, 23):
        key = dash.cell(r, 5).value
        cache[("Dashboard", f"F{r}")] = cnt(lambda t, k=key: cat_c[t] == k)
        cache[("Dashboard", f"G{r}")] = nbv_sum(lambda t, k=key: cat_c[t] == k)

    for r in range(32, 44):
        key = dash.cell(r, 1).value
        if not key:
            continue
        cache[("Dashboard", f"B{r}")] = cnt(lambda t, k=key: loc_i[t] == k)
        cache[("Dashboard", f"C{r}")] = nbv_sum(lambda t, k=key: loc_i[t] == k)

    for r in range(32, 42):
        key = dash.cell(r, 5).value
        if not key:
            continue
        cache[("Dashboard", f"F{r}")] = ex_type_count.get(str(key), 0)
        cache[("Dashboard", f"G{r}")] = round(ex_type_exp.get(str(key), 0), 2)

    reg_sum = round(sum(cost_by_tag.values()), 2)
    led_sum = round(sum(sl_cost.values()), 2)
    cache[("Ledger Reconciliation", "B5")] = reg_sum
    cache[("Ledger Reconciliation", "B6")] = led_sum
    cache[("Ledger Reconciliation", "B7")] = round(reg_sum - led_sum, 2)
    cache[("Ledger Reconciliation", "B8")] = round(sum(nbv_by_tag.values()), 2)
    cache[("Ledger Reconciliation", "B9")] = round(
        sum(sl_nbv[t] for t in sl_nbv if sl_status.get(t) == "Active"), 2
    )
    cache[("Ledger Reconciliation", "B10")] = round(sum(sl_nbv.values()), 2)

    mismatch_n = missing_n = 0
    for r in range(24, 36):
        label = str(lr.cell(r, 2).value or "")
        if label == "Acquisition cost mismatch":
            mismatch_n += 1
        if "missing from ledger" in label.lower():
            missing_n += 1
    cache[("Ledger Reconciliation", "B12")] = mismatch_n
    cache[("Ledger Reconciliation", "B13")] = missing_n
    cache[("Ledger Reconciliation", "B11")] = len(tags) - mismatch_n - missing_n

    for r in range(24, 36):
        tag = lr.cell(r, 1).value
        if not tag:
            continue
        tag = str(tag)
        rc = cost_by_tag.get(tag, "")
        lc = sl_cost.get(tag, "")
        if rc != "":
            cache[("Ledger Reconciliation", f"C{r}")] = rc
        if lc != "":
            cache[("Ledger Reconciliation", f"D{r}")] = lc
        if rc != "" and lc != "":
            cache[("Ledger Reconciliation", f"E{r}")] = round(float(lc) - float(rc), 2)
        if tag in sl_nbv:
            cache[("Ledger Reconciliation", f"F{r}")] = nbv_by_tag.get(tag, 0)
            cache[("Ledger Reconciliation", f"G{r}")] = sl_nbv[tag]

    return cache


def rebuild_zips() -> None:
    inputs = [
        "it_asset_inventory.xlsx",
        "hr_employee_status.csv",
        "equipment_transfer_log.csv",
        "service_desk_offboarding.csv",
        "device_return_shipments.csv",
        "hardware_purchase_orders.csv",
        "fixed_asset_ledger.xlsx",
        "asset_disposal_records.csv",
        "regional_IT_notes.docx",
        "IT_asset_management_policy.pdf",
        "ITAM_control_matrix.png",
        "receiving_exception_scan_1ZMD00000082.png",
    ]
    for zip_name in ("Yanou_IT_Asset_Inputs.zip", "Meridian_IT_Asset_Inputs.zip"):
        with zipfile.ZipFile(ROOT / zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
            for name in inputs:
                zf.write(ROOT / name, arcname=name)

    shutil.copy2(WORKBOOK, MERIDIAN)
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(WORKBOOK, "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Meridian_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(MERIDIAN, "Meridian_IT_Asset_Reconciliation.xlsx")
    print("rebuilt deliverable zips")


def verify() -> None:
    wb = load_workbook(WORKBOOK, data_only=True)
    lr = wb["Ledger Reconciliation"]
    dash = wb["Dashboard"]
    cr = wb["Corrected Register"]

    for tag, po in MISMATCH_PO.items():
        for r in range(24, 36):
            if lr.cell(r, 1).value == tag:
                expl = str(lr.cell(r, 8).value or "")
                assert po in expl, f"{tag} missing {po} in: {expl}"
                wrong = [p for p in re.findall(r"PO-[\d-]+", expl) if p != po and p in MISMATCH_PO.values()]
                assert not wrong or po in expl

    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(WORKBOOK) as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rel_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall(".//{*}Relationship")
        }
        for sh in wb_root.findall(".//{*}sheet"):
            if sh.attrib.get("name") != "Ledger Reconciliation":
                continue
            rid = sh.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            xml = z.read(rid_to_target[rid]).decode()
            for coord in ["B5", "B7", "B8"]:
                m = re.search(rf'<c r="{coord}"[^>]*>(.*?)</c>', xml, re.S)
                v = re.search(r"<v>([^<]+)</v>", m.group(1)) if m else None
                assert v and v.group(1), f"LR {coord} missing cached value"
                print(f"LR {coord} cached = {v.group(1)}")

    for r in range(5, cr.max_row + 1):
        if cr.cell(r, 1).value == "MD-00034":
            ev = str(cr.cell(r, 11).value or "")
            assert "XLOOKUP" not in ev, ev
            assert "479.88" in ev, ev
            print("MD-00034 evidence OK")

    print("Dashboard category:", dash["F19"].value, dash["F20"].value, dash["F21"].value, dash["F22"].value)
    print("LR B5/B7/B8 (data_only):", lr["B5"].value, lr["B7"].value, lr["B8"].value)


def main() -> None:
    wb = load_workbook(WORKBOOK)
    n_po = fix_lr_po_explanations(wb)
    print(f"fixed LR PO citations: {n_po}")
    fix_dashboard_category_formulas(wb)
    n_ev = fix_evidence_xlookup(wb)
    print(f"fixed evidence XLOOKUP leaks: {n_ev}")
    fix_critical_blocker_language(wb)
    wb.save(WORKBOOK)

    cache = evaluate_formula_cache(wb)
    print(f"computed cache entries: {len(cache)}")
    inject_formula_cache(WORKBOOK, cache)
    rebuild_zips()
    verify()


if __name__ == "__main__":
    main()
