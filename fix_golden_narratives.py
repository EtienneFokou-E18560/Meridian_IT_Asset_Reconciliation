#!/usr/bin/env python3
"""Break uniform-log-template tells in the golden workbook narrative columns.

Rewrites Exception Register (Required Action, Evidence Assessment),
Corrected Register (Evidence Source), and Custody Chain (Detail).
Keeps IDs, statuses, formulas, and certification conclusions intact.
"""
from __future__ import annotations

import csv
import random
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
RNG = random.Random(20260818)

def gid(pattern: str, text: str | None) -> str | None:
    if not text:
        return None
    m = re.search(pattern, str(text))
    return m.group(1) if m else None


ID_RE = re.compile(
    r"(?:TR|APR|OFF|RET|FA|DC|EX)-\d{4,6}"
    r"|PO-\d{4}-\d{4}"
    r"|E\d{4}"
    r"|1ZMD\d+"
    r"|MD-(?:LA|MO|NE)-\d+"
    r"|MD-\d{5}"
    r"|MISMATCH-\d+"
    r"|RN-\d+"
    r"|ITAM_control_matrix\.png"
    r"|receiving_exception_scan_[A-Za-z0-9_.]+"
    r"|ITAM-001"
    r"|FA-000\d+"
)


def tokens(text: str) -> list[str]:
    if not text:
        return []
    return ID_RE.findall(str(text))


def fmt_date(s: str | None, i: int) -> str:
    if not s or s in ("None", "blank", ""):
        return ""
    raw = str(s)[:10]
    try:
        d = datetime.strptime(raw, "%Y-%m-%d")
    except ValueError:
        return raw
    styles = [
        d.strftime("%Y-%m-%d"),
        d.strftime("%-m/%-d"),
        d.strftime("%b %-d"),
        d.strftime("%B %-d, %Y"),
        d.strftime("%-d %b"),
        d.strftime("%Y-%m-%d"),
        d.strftime("%-m/%-d/%y"),
    ]
    return styles[i % len(styles)]


def load_csv(name: str) -> list[dict]:
    with (ROOT / name).open(newline="") as f:
        return list(csv.DictReader(f))


def unique_write(used: dict[str, set], bucket: str, text: str, salt: str) -> str:
    text = re.sub(r"\s+", " ", text).strip().replace(" .", ".").replace(" ,", ",")
    text = text.replace(";;", ";").replace("..", ".")
    if text in used[bucket]:
        text = f"{text} ({salt})"
    used[bucket].add(text)
    return text


def parse_kv(rel: str) -> dict[str, str]:
    out = {}
    if not rel:
        return out
    for part in str(rel).split(";"):
        part = part.strip()
        if ":" in part:
            k, v = part.split(":", 1)
            out[k.strip()] = v.strip()
    return out


def main() -> None:
    hr = {r["employee_id"]: r for r in load_csv("hr_employee_status.csv")}
    transfers = {r["transfer_id"]: r for r in load_csv("equipment_transfer_log.csv")}
    tickets = {r["ticket_number"]: r for r in load_csv("service_desk_offboarding.csv")}
    xfers_by_tag = {r["asset_tag"]: r for r in load_csv("equipment_transfer_log.csv")}

    wb = load_workbook(ROOT / "Yanou_IT_Asset_Reconciliation.xlsx")
    cr = wb["Corrected Register"]
    er = wb["Exception Register"]
    cc = wb["Custody Chain"]
    used: dict[str, set] = defaultdict(set)

    assets: dict[str, dict] = {}
    for r in range(5, cr.max_row + 1):
        tag = cr.cell(r, 1).value
        assets[tag] = {
            "row": r,
            "tag": tag,
            "serial": cr.cell(r, 2).value,
            "cat": cr.cell(r, 3).value,
            "model": cr.cell(r, 4).value,
            "reg_cust": cr.cell(r, 5).value,
            "reg_loc": cr.cell(r, 6).value,
            "reg_st": cr.cell(r, 7).value,
            "ver_cust": cr.cell(r, 8).value,
            "ver_loc": cr.cell(r, 9).value,
            "ver_st": cr.cell(r, 10).value,
            "src": cr.cell(r, 11).value or "",
            "asof": str(cr.cell(r, 12).value or "")[:10],
            "cost": cr.cell(r, 13).value,
            "fa": cr.cell(r, 15).value,
            "conf": cr.cell(r, 17).value,
            "exids": cr.cell(r, 18).value or "",
        }

    rewrite_exceptions(er, assets, used)
    rewrite_evidence(cr, assets, hr, transfers, tickets, used)
    rewrite_custody(cc, assets, hr, transfers, tickets, xfers_by_tag, used)

    wb.save(ROOT / "Yanou_IT_Asset_Reconciliation.xlsx")
    with zipfile.ZipFile(ROOT / "Yanou_IT_Asset_Reconciliation.zip", "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(ROOT / "Yanou_IT_Asset_Reconciliation.xlsx", "Yanou_IT_Asset_Reconciliation.xlsx")
    report(used)


def rewrite_exceptions(er, assets, used) -> None:
    for r in range(5, er.max_row + 1):
        exid = er.cell(r, 1).value
        typ = er.cell(r, 2).value or ""
        tag = er.cell(r, 4).value
        serial = er.cell(r, 5).value
        rel = er.cell(r, 6).value or ""
        owner = er.cell(r, 9).value
        impact = er.cell(r, 13).value
        kv = parse_kv(rel)
        a = assets.get(tag, {})
        i = r + (ord(tag[-1]) if tag else 0)

        er.cell(r, 8).value = unique_write(
            used, "action", action_for(typ, tag, serial, rel, kv, a, owner, i), exid
        )
        er.cell(r, 12).value = unique_write(
            used, "assess", assess_for(typ, tag, serial, rel, kv, a, impact, i), exid
        )


def action_for(typ, tag, serial, rel, kv, a, owner, i) -> str:
    def nz(*vals):
        for v in vals:
            if v and str(v) not in ("None", "none"):
                return str(v)
        return ""

    loc = nz(kv.get("RegisterLocation"), kv.get("LastRegisterLocation"), a.get("reg_loc")) or "a closed site"
    live = nz(kv.get("HRLocation"), a.get("ver_loc")) or "an open office"
    tr = nz(kv.get("Transfer"))
    emp = nz(kv.get("Employee"))
    tk = nz(kv.get("Ticket"))
    track = nz(kv.get("Tracking"))
    due = nz(kv.get("ReturnDue"))
    ship_ser = nz(kv.get("ShipmentSerial"))
    fa = nz(kv.get("Ledger"), a.get("fa"))
    po = nz(kv.get("PO"))
    also = nz(kv.get("AlsoTagged"))
    lc = nz(kv.get("LedgerCost"))
    rc = nz(kv.get("RegisterCost"))

    if typ == "Closed Location Assignment":
        opts = [
            f"Get {tag} off {loc}. Operating location is {live}.",
            f"Register still has the shuttered floor ({loc}). Put {live} on the row.",
            f"{tag} shouldn't sit on a locked site. Move location to {live}.",
            f"Location cleanup for {tag}: {loc} → {live}.",
            f"Stop listing this at {loc}. {tr + ' / ' if tr else ''}HR point to {live}.",
            f"Closed-office leftover. Update {tag} to {live}.",
            f"{live} is current. Register lagged on {loc} for {tag}.",
            f"Don't leave {tag} coded to {loc}.",
            f"Fix the site on {tag}. Not at {loc} anymore.",
            f"Pull {tag} out of {loc} and park it at {live}.",
            f"{owner} to recode location — {tag} is not on the closed floor.",
            f"Clear the closed-site assignment. {tag} belongs at {live}.",
            f"{loc} is shut. Rewrite the register location ({tag}).",
            f"Same finding as the walkthrough: {tag} still tagged {loc}. Change it.",
            f"Update {tag} location. {live} is what transfer/HR support.",
        ]
        return opts[i % len(opts)]

    if typ == "Former Employee Assignment":
        opts = [
            f"Take {emp or 'the separated employee'} off {tag}. Park with Regional IT / IT Stock.",
            f"{tag} is still on a leaver. Close the assignment; stock holds it until recovered or written off.",
            f"HR {emp} is gone. Register custodian on {tag} has to move.",
            f"Unassign {tag} from the former employee. {tk + ' is the offboarding ticket.' if tk else ''}".strip(),
            f"Custodian on {tag} is a terminated ID. Put IT Stock / Regional IT on the row.",
            f"Don't leave {tag} assigned to {emp or 'a separated person'}.",
            f"{tk or 'Offboarding'} closed the people side; asset row still shows the leaver. Fix custodian.",
            f"Recover or lose-process {tag}, but first drop the former employee as owner.",
            f"Assignment cleanup: {tag} → Regional IT (or IT Stock if it's on the shelf).",
            f"{emp} should not still own {tag} on the register.",
            f"Move {tag} off the separated custodian before we talk redeploy.",
            f"Register owner is stale. {tag} needs a stock/regional holder.",
            f"HR already termed {emp or 'this person'}. Update {tag} assignment.",
            f"Close the people assignment on {tag}. Ticket {tk or 'on file'}.",
            f"{tag}: former employee still listed. That's the fix.",
        ]
        return opts[i % len(opts)]

    if typ == "Unapproved Transfer":
        trid = kv.get("Transfer") or tr
        opts = [
            f"Get a retrospective approval on {trid} or reverse it (ITAM-001 §4).",
            f"{trid} has a blank approval. Ops manager signs or we unwind the move.",
            f"Don't treat {trid} as a clean transfer until an APR lands — or reverse it.",
            f"Approval missing on {trid}. Retro sign-off or put the asset back.",
            f"{tag}: movement {trid} never got an approval ID. Fix per §4.",
        ]
        return opts[i % len(opts)]

    if typ == "Overdue Return":
        opts = [
            f"Chase the return. {track or 'The 1Z label'} is Label Created only — not received.",
            f"{tag} past due {due or 'on the ticket'}. Label isn't proof. Escalate recovery.",
            f"Call the former employee; {tk or 'ticket'} stays open until dock scan + delivery.",
            f"Escalate {tag}. Generated label {track or ''} is not a return.".replace("  ", " "),
            f"Recovery, not close-out. {track} never left Label Created.",
            f"Don't mark {tag} returned. {due} came and went with no receiving scan.",
            f"{tk}: overdue. Keep pressure on the person; carrier file is empty.",
            f"Past the window on {tag}. Label-only returns block cert until SCAN-86 / delivery / dock.",
            f"Treat {tag} as outstanding kit. {track} is paperwork, not inbound.",
            f"Escalate to recovery. {tag} / {track} still label-only.",
        ]
        return opts[i % len(opts)]

    if typ == "Shipment Mismatch":
        opts = [
            f"Hold {tag}. Serial on the shipment ({ship_ser}) is not {serial}. No Available until a matching scan.",
            f"Work it with carrier/receiving. {track} shows {ship_ser}; register is {serial}.",
            f"Do not clear custody on {tag}. Mismatch {ship_ser} vs {serial}.",
            f"Exception stays open on {tag} until inbound serial matches {serial}.",
            f"{track}: serial fight. Keep In Transit - Exception.",
        ]
        if "1ZMD00000082" in rel or tag == "MD-00082":
            opts.append(
                "MD-00082 stays exception. Dock photo is a HOLD lead, not a receipt. "
                "MISMATCH-0082 vs MD-MO-050082 on 1ZMD00000082."
            )
        return opts[i % len(opts)]

    if typ == "Cannot Locate":
        opts = [
            f"Loss investigation on {tag}. Keep Missing until we find it or get a write-off.",
            f"{tag} last seen {loc}. Security + ITAM locate; don't flip to Available.",
            f"Cannot locate {tag}. Ticket {tk or 'open'}. Missing until count or approved write-off.",
            f"Start the loss process. {tag} is not in {live} and not in receiving.",
            f"Classify {tag} Missing. Last register site {loc} is a closed floor — not a find.",
            f"{tk}: still gone. Don't invent a stock location for {tag}.",
            f"Security review + locate. {tag} stays Missing.",
            f"No scan, no cert, no count for {tag}. Missing, not In Use.",
            f"Write-off only with approval. Until then {tag} is Missing.",
            f"{tag} dropped off the map after {loc}. Investigate; don't tidy the status.",
        ]
        return opts[i % len(opts)]

    if typ == "Missing Disposal Evidence":
        opts = [
            f"Hold the retirement on {tag}. Need a disposal certificate or put it back in stock.",
            f"{tk or 'RET ticket'} has no matching cert. {tag} stays pending-disposal.",
            f"Either get the recycler cert for {tag} or restore IT Stock as custodian.",
            f"Don't call {tag} disposed. Certificate missing.",
            f"Paperwork gap on {tag}. Cert or reverse the retire.",
            f"{tag}: pending disposal with no DC row. Block retirement close.",
            f"Match a vendor cert to {serial} or unwind.",
            f"ITAM hold. {tag} is not gone until the certificate shows the tag.",
            f"Obtain cert or restock {tag}. Ticket {tk}.",
            f"Retirement parked. {tag} needs DC evidence.",
        ]
        return opts[i % len(opts)]

    if typ == "Duplicate Serial Number":
        opts = [
            f"Physical check and retag. {serial} is on {also or 'two tags'}.",
            f"Two assets share {serial} ({also}). Validate on the bench and issue a new serial.",
            f"Duplicate {serial} — {tag} and the other tag in {also}. Don't treat either as clean until retag.",
            f"Break the duplicate. {serial} cannot stay on both {also}.",
        ]
        return opts[i % len(opts)]

    if typ == "Inventory-to-Ledger Difference":
        if not lc or tag == "MD-00132":
            opts = [
                f"Add a FAR row (or approved write-off) for {tag} — ${rc} is over the $2,500 gate. RN-0132 is not a reason to skip it.",
                f"{tag} is missing from the ledger. Capitalize or write off; do not take RN-0132 as expense-under-threshold.",
                f"Critical FA gap on {tag} (${rc}). {po} capitalization_approved does not replace an FA row.",
                f"Finance: FA line for {tag} or a signed write-off. Regional note RN-0132 is non-authoritative.",
            ]
            return opts[i % len(opts)]
        opts = [
            f"Finance to align cost on {tag}: register {rc} vs ledger {lc} ({fa}).",
            f"Cost basis mismatch {tag}. Pull {po} and fix {fa} or the register.",
            f"{tag} is off vs {fa} ({rc} vs {lc}). Procurement to confirm which number is right.",
            f"Don't certify the $ on {tag} until register and {fa} match.",
            f"Reconcile {po} / {fa} / register cost for {tag}.",
            f"FAR vs ITAM spread on {tag} ({rc} vs {lc}). Fix the basis.",
            f"{owner} owns the cost-basis cleanup on {tag}.",
        ]
        return opts[i % len(opts)]

    if typ == "Below Capitalization Threshold":
        cost = rc or a.get("cost")
        opts = [
            f"{tag} is ${cost} — under the $2,500 per-asset gate. No FA row expected; keep operational tracking.",
            f"Do not treat {tag} as a Critical ledger gap. {po} capitalization_approved doesn't override ITAM-001 §7.",
            f"${cost} on {tag}. Below threshold. Informational only.",
            f"{tag}: under-threshold. Leave it off the FAR on purpose.",
        ]
        return opts[i % len(opts)]

    return f"Work the exception on {tag}. {rel}"


def assess_for(typ, tag, serial, rel, kv, a, impact, i) -> str:
    def nz(*vals):
        for v in vals:
            if v and str(v) not in ("None", "none"):
                return str(v)
        return ""

    track = nz(kv.get("Tracking"))
    tk = nz(kv.get("Ticket"))
    tr = nz(kv.get("Transfer"))
    emp = nz(kv.get("Employee"))
    loc = nz(kv.get("RegisterLocation"), kv.get("LastRegisterLocation"), a.get("reg_loc"))
    live = nz(kv.get("HRLocation"), a.get("ver_loc"))
    fa = nz(kv.get("Ledger"))
    po = nz(kv.get("PO"))
    lc = nz(kv.get("LedgerCost"))
    rc = nz(kv.get("RegisterCost"))
    img = "receiving_exception_scan_1ZMD00000082.png" in rel or tag == "MD-00082"

    if typ == "Shipment Mismatch" and img:
        opts = [
            f"1ZMD00000082 + the Atlanta photo (HOLD / MISMATCH-0082). Photo is a lead, not a receipt.",
            f"Looked at the dock scan for {track or '1ZMD00000082'}. It flags a mismatch; it does not clear {tag}.",
            f"Carrier file and receiving_exception_scan_1ZMD00000082.png. Image ≠ inbound clearance.",
            f"MISMATCH-0082 vs MD-MO-050082. Used the photo as a breadcrumb only.",
        ]
        return opts[i % len(opts)]

    if typ == "Shipment Mismatch":
        opts = [
            f"Shipment serial {kv.get('ShipmentSerial')} on {track} vs register {serial}. Not cleared.",
            f"Pulled {tk} and {track}. Serials don't match. Held.",
            f"{tag}: inbound record doesn't agree with {serial}.",
            f"Carrier exception on {track}. Status stays In Transit - Exception.",
        ]
        return opts[i % len(opts)]

    if typ == "Overdue Return":
        opts = [
            f"{track} is still Label Created. No acceptance / delivery / dock scan.",
            f"Ship file for {tag}: label only. {tk} overdue {kv.get('ReturnDue','')}.",
            f"Didn't treat the prepaid label as a return. {track}.",
            f"Carrier events missing on {track}. That's the assessment.",
            f"{tk} vs device_return_shipments: no SCAN-86.",
            f"Label exists. Receipt doesn't. {tag}.",
            f"Policy §5: label ≠ returned. {track}.",
            f"Looked at tracking {track}. Never left the printer as far as the file shows.",
        ]
        return opts[i % len(opts)]

    if typ == "Cannot Locate":
        opts = [
            f"No receiving scan, cert, or count for {tag}. Last site {kv.get('LastRegisterLocation','closed')}.",
            f"{tk} / {tr}: still gone. Notes don't locate it.",
            f"Walked HR, transfers, tickets. {tag} is not in any inbound.",
            f"Empty trail after {kv.get('LastRegisterLocation')}. Missing is the call.",
            f"{tag}: nothing higher-precedence than a closed-floor register row.",
            f"Couldn't prove a location. {rel}",
        ]
        return opts[i % len(opts)]

    if typ == "Below Capitalization Threshold":
        opts = [
            f"PO {kv.get('PO')} / ITAM-001 §7. ${kv.get('RegisterCost')} is under $2,500. FAR silence is expected.",
            f"Checked the threshold against {tag}. Not a capital gap.",
            f"capitalization_approved on the PO doesn't pull this onto the ledger. Cost {kv.get('RegisterCost')}.",
            f"{tag} under the gate. Informational.",
        ]
        return opts[i % len(opts)]

    if typ == "Unapproved Transfer":
        opts = [
            f"{kv.get('Transfer')} approval field is blank in the transfer log.",
            f"Opened {kv.get('Transfer')}. No APR- id. That's the finding.",
            f"Movement exists; authorization doesn't. {kv.get('Transfer')}.",
            f"ITAM-001 §4: {kv.get('Transfer')} isn't a completed approved transfer.",
        ]
        return opts[i % len(opts)]

    if typ == "Duplicate Serial Number":
        opts = [
            f"Same serial {serial} on {kv.get('AlsoTagged')}. Register dump, not a scan error I could clear.",
            f"Two tags, one serial ({serial}). Flagged both.",
            f"Inventory file lists {serial} twice ({kv.get('AlsoTagged')}).",
            f"Duplicate {serial} is in the extract. Physical next.",
        ]
        return opts[i % len(opts)]

    if typ == "Inventory-to-Ledger Difference":
        if not lc or tag == "MD-00132":
            opts = [
                f"{po} on {tag} (${rc}) and no FA row. RN-0132 does not clear it.",
                f"FAR extract has no {tag}. Cost ${rc} is over $2,500. Note RN-0132 is a claim, not a posting.",
                f"Missing capital line. Policy ITAM-001 §7 vs RN-0132 — I kept the gap.",
                f"{tag} still isn't in fixed_asset_ledger.xlsx. That's the assessment.",
            ]
            return opts[i % len(opts)]
        opts = [
            f"{fa} cost {lc} vs register {rc}. {po}.",
            f"FAR extract and the ITAM row don't tie on {tag}.",
            f"Pulled {po} against {fa}. Spread remains.",
            f"Numbers disagree. Register {rc}, ledger {lc}.",
            f"{tag} cost basis: two systems, two amounts.",
        ]
        return opts[i % len(opts)]

    if typ == "Missing Disposal Evidence":
        opts = [
            f"{tk} / {tr}: certificate field empty." if (tk or tr) else f"Certificate field empty on {tag}.",
            f"Disposal file has no row matching {tag} / {serial}.",
            f"Retirement path without a DC. {tag}.",
            f"Looked in asset_disposal_records. Nothing for {tag}.",
            f"No verified certificate. {rel}",
        ]
        return opts[i % len(opts)]

    if typ == "Former Employee Assignment":
        who = emp or "the assignee"
        opts = [
            f"HR {who} is separated. Register still has them on {tag}." + (f" {tk}." if tk else ""),
            f"{tk or 'Offboarding'} vs hr_employee_status: people side is done, asset isn't.",
            f"Leaver still on {tag}." + (f" {who}." if emp else "") + (f" {tr}." if tr else ""),
            f"Checked the term date against the assignment. Stale owner on {tag}.",
            f"{tag} custodian is a former employee in HR ({who}).",
            f"Offboarding ticket {tk or 'on file'} doesn't put hardware in stock by itself.",
            f"{who} already termed. {tag} assignment wasn't updated.",
            f"People record is closed; {tag} isn't. {tk or tr or who}.",
        ]
        return opts[i % len(opts)]

    if typ == "Closed Location Assignment":
        site = loc or "the closed floor"
        opts = [
            f"Register location {site} is a closed site." + (f" {tr} points elsewhere." if tr else "") + (f" Live office {live}." if live else ""),
            f"{tag} still coded to a floor we don't occupy ({site}).",
            f"Closed-office leftover on the extract. {site}." + (f" {tr}." if tr else ""),
            f"Site code is stale." + (f" {site} vs {live}." if live else f" {site}."),
            f"Didn't treat the closed building as a live location for {tag}.",
            f"{tr} destination isn't the shuttered office on the register." if tr else f"No transfer keeps {tag} at {site}.",
            f"Location on {tag} didn't survive the consolidation ({site}).",
            f"{site} is locked. I'm not leaving {tag} there on paper." + (f" {live} is the operating site." if live else ""),
            f"Walked the register vs HR/transfer. {tag} still says {site}.",
            f"Consolidation leftover: {tag} @ {site}.",
        ]
        return opts[i % len(opts)]

    return f"Used the cited records on {tag}: {rel}"


def rewrite_evidence(cr, assets, hr, transfers, tickets, used) -> None:
    for tag, a in assets.items():
        r = a["row"]
        old = a["src"]
        keep = tokens(old)
        i = a["row"]
        text = evidence_for(a, hr, transfers, tickets, i)
        # guarantee prior IDs still searchable
        missing = [t for t in keep if t not in text]
        # drop boilerplate tokens we are intentionally not repeating
        skip = {"ITAM_control_matrix.png"}
        missing = [t for t in missing if t not in skip]
        if missing:
            text = text.rstrip(".") + ". " + " ".join(missing) + "."
        # force mismatch cites on the four exception in-transit assets
        if tag == "MD-00082":
            for must in ("1ZMD00000082", "MISMATCH-0082", "MD-MO-050082", "receiving_exception_scan_1ZMD00000082.png"):
                if must not in text:
                    text += f" {must}."
            text = text.rstrip(". ") + ". Dock photo is a lead, not receiving clearance."
        cr.cell(r, 11).value = unique_write(used, "evidence", text, tag)


def evidence_for(a, hr, transfers, tickets, i) -> str:
    tag = a["tag"]
    src = a["src"]
    eid = gid(r"(E\d{4})", src)
    if not eid and isinstance(a["ver_cust"], str) and re.fullmatch(r"E\d{4}", a["ver_cust"]):
        eid = a["ver_cust"]
    name = hr.get(eid, {}).get("employee_name") if eid else None
    hrst = hr.get(eid, {}).get("employment_status") if eid else None
    tr = gid(r"(TR-\d+)", src)
    if not tr:
        tr = gid(r"(TR-\d+)", str(a.get("exids")))
    apr = gid(r"(APR-\d+)", src)
    po = gid(r"(PO-[\d-]+)", src)
    fa = a.get("fa") or gid(r"(FA-\d+)", src)
    tk = gid(r"((?:OFF|RET)-\d+)", src)
    track = gid(r"(1ZMD\d+)", src)
    BLANK_APR = {"TR-00058", "TR-00059", "TR-00064", "TR-00065", "TR-00127"}
    trow: dict = {}
    rec = ""
    tdate = ""
    apr_bit = ""
    if not tr:
        for tid, row in transfers.items():
            if row.get("asset_tag") == tag:
                tr = tid
                break
    if tr:
        trow = transfers.get(tr) or {}
        rec = trow.get("received_date") or ""
        tdate = trow.get("transfer_date") or ""
        if tr in BLANK_APR:
            apr_bit = " (approval blank)"
            apr = None
        else:
            apr = apr or trow.get("approval_id")
            apr_bit = f" ({apr})" if apr else ""
    loc = a["ver_loc"]
    st = a["ver_st"]
    model = a["model"]
    cost = a["cost"]

    who = f"{name} ({eid})" if name and eid else (eid or a["ver_cust"] or "stock")
    hr_bits = []
    if eid and hrst:
        hr_bits = [
            f"{who} is {hrst} in HR",
            f"HR file: {eid} {hrst}" + (f" — {name}" if name else ""),
            f"Looked up {eid}; still {hrst}",
            f"{name or eid} hasn't changed status ({hrst})" if hrst == "Active" else f"{name or eid} is {hrst} ({eid})",
            f"Custodian check: {eid} / {hrst}",
            f"hr_employee_status still has {eid} as {hrst}",
        ]
    tr_bits = []
    if tr:
        rec_s = fmt_date(rec, i) if rec else ""
        td_s = fmt_date(tdate, i + 3) if tdate else ""
        tr_bits = [
            f"{tr}{apr_bit} on the movement log",
            f"transfer {tr}{apr_bit}",
            f"pulled {tr}" + (f", inbound {rec_s}" if rec_s else ""),
            f"assignment sits on {tr}{apr_bit}",
            f"{tr} dated {td_s}" if td_s else f"{tr}{apr_bit}",
            f"log row {tr}" + (f"; received {rec_s}" if rec_s else "; no receive date"),
        ]
    money = []
    if po and fa:
        money = [
            f"bought on {po}; FAR {fa}",
            f"{fa} in the ledger extract, {po}",
            f"{po} / {fa}",
            f"cost ${cost} on {po}, asset {fa}",
            f"{po} flowed to {fa}",
        ]
    elif po and not fa:
        money = [
            f"{po} on file; no FA row",
            f"purchasing {po}, nothing in the FAR extract",
            f"{po} only — no ledger tag",
        ]
    elif fa:
        money = [f"ledger {fa}", f"{fa} still on the books"]

    extra = []
    if track:
        extra.append(f"tracking {track}")
    if tk:
        extra.append(f"ticket {tk}")
    if "label-only" in src.lower() or st == "Return Overdue":
        extra.append("return is still label-only")
    if st == "Missing":
        extra.append(f"register last showed {a['reg_loc']}")
    if tag in ("MD-00074", "MD-00076", "MD-00082", "MD-00084"):
        mm = f"MISMATCH-{tag[-4:]}"
        extra.append(f"shipment serial {mm} vs {a['serial']}")
    if tag == "MD-00082":
        extra.append("receiving_exception_scan_1ZMD00000082.png is a HOLD lead")
    if tag == "MD-00132":
        extra.append("RN-0132 claimed under-threshold expense; rejected — $6,470 is over the $2,500 gate and there is still no FA row")
    if tag in ("MD-00130", "MD-00131"):
        extra.append(f"${cost} is under $2,500 so missing FA is expected")
    if tag in ("MD-00114", "MD-00118"):
        extra.append("disposal certificate not in the file")
    if a["serial"] in ("MD-LA-050021", "MD-NE-050088"):
        extra.append(f"duplicate serial {a['serial']}")

    # pick a sentence shape — many structurally different openings
    hi = hr_bits[i % len(hr_bits)] if hr_bits else ""
    ti = tr_bits[(i * 3) % len(tr_bits)] if tr_bits else ""
    mi = money[(i * 5) % len(money)] if money else ""
    ex = extra[(i * 7) % len(extra)] if extra else ""
    rest = [x for x in extra if x != ex]

    shapes = []
    if st == "In Use" and hrst == "Active":
        shapes = [
            f"{hi}. {ti}. {mi}. Left {tag} as In Use in {loc}.".replace(" .", ""),
            f"{model} {tag} — {ti}; {hi}. {mi}.",
            f"Didn't rewrite {tag}. {hi}, and {ti}. {mi}.",
            f"{loc}: {who}. {ti}. {mi}.",
            f"Register owner was {a['reg_cust']}; verified {who}. {ti}. {mi}.",
            f"{mi}. People/move side: {hi}; {ti}.",
            f"Cross-check on {tag} landed on {st} / {loc}. {hi}. {ti}. {mi}.",
            f"{ti}. HR agrees ({hi}). {mi}.",
            f"Keeping {tag} with {who} in {loc}. {ti}. {mi}.",
            f"{hi} so I didn't bounce the assignment. {ti}. {mi}.",
            f"Live kit. {tag} {model}. {hi}. {ti}. {mi}.",
            f"{who}, {loc}. Movement {ti}. {mi}.",
        ]
    elif st == "Available":
        shapes = [
            f"{tag} is on IT Stock in {loc}. {ti}. {mi}.",
            f"Parked as Available — {model}. {ti}. {mi}.",
            f"Not assigned to a person. {tag} / {loc}. {ti}. {mi}.",
            f"Stock row. {ti}. {mi}.",
            f"{mi}. Status Available at {loc} after {ti}.",
            f"Shelf/stock: {tag}. {ti}. {mi}.",
            f"IT Stock holds {tag} ({loc}). {ti}. {mi}.",
        ]
    elif st == "Pending Redeployment":
        shapes = [
            f"{tag} came off a leaver ({hi or who}). Sitting with {a['ver_cust']} in {loc}. {ti}. {tk or ''}. {mi}.",
            f"Pending redeploy, not In Use. {hi}. {ti}. {mi}. {tk or ''}.",
            f"{a['ver_cust']} holding {tag} after term. {ti}. {mi}.",
            f"Don't send this back to a person yet. {tag}. {ti}. {tk or ''}. {mi}.",
        ]
    elif st == "Return Overdue":
        shapes = [
            f"{hi}. {ti}. {track or ''} still Label Created — not a return. {tk or ''}. {mi}.",
            f"Overdue kit {tag}. {ex}. {hi}. {ti}. {mi}.",
            f"{tk or 'Ticket'} past due. {track} never accepted. {hi}. {mi}.",
            f"I left {tag} Return Overdue. Label isn't SCAN-86. {hi}. {ti}. {mi}.",
            f"{who} is {hrst}. Hardware isn't back. {track}. {mi}.",
        ]
    elif st == "In Transit - Exception":
        shapes = [
            f"Held {tag} as In Transit - Exception. {ex}. {ti}. {mi}. {tk or ''}.",
            f"Serial fight on {track}. {a['serial']} vs the shipment record. Not Available. {mi}.",
            f"{tag} inbound exception. {ex}. {ti}. {mi}.",
            f"Did not clear {tag}. {ex}. {tk or ''}. {mi}.",
        ]
    elif st == "Missing":
        shapes = [
            f"{tag} is Missing. Last register site {a['reg_loc']}. {tk or ''}. {ti}. {mi}.",
            f"No scan/cert/count. {tag} stays Missing. {hi or ti}. {mi}.",
            f"Can't locate {tag} ({model}). {tk}. {ti}. {mi}.",
            f"{a['reg_loc']} was the last listed floor — it's closed and the device isn't in receiving. {mi}.",
            f"Loss path. {tag}. {tk or ''}. {ti}. {mi}.",
        ]
    elif st in ("Retired/Pending Disposal", "Disposed"):
        shapes = [
            f"{st} on {tag}. {tk or ''}. {ti}. {mi}. {ex}".strip(),
            f"{tag} retirement/disposal trail: {tk or 'RET ticket'}. {ti}. {mi}.",
            f"{model} {tag} — {st}. {ti}. {mi}. {ex}".strip(),
            f"Not Available. {tag} is {st}. {ti}. {mi}.",
        ]
    else:
        shapes = [
            f"{tag} verified {st} / {loc}. {hi}. {ti}. {mi}.",
            f"{who} is {hrst or 'on file'}. {tag} stays {st} in {loc}. {ti}. {mi}.",
            f"{model} ({tag}): {st}. {hi}. {ti}. {mi}.",
            f"Didn't invent a FAR line. {tag} {st} at {loc}. {ti}. {mi}.",
        ]

    text = shapes[i % len(shapes)]
    joiners = [". ", " — ", "; ", ". Also ", ". Note: "]
    for n, piece in enumerate(extra):
        if piece and piece not in text:
            text = text.rstrip(".") + joiners[(i + n) % len(joiners)] + piece
    text = re.sub(r"\s+", " ", text).strip()
    return text


def rewrite_custody(cc, assets, hr, transfers, tickets, xfers_by_tag, used) -> None:
    for r in range(5, cc.max_row + 1):
        tag = cc.cell(r, 1).value
        et = cc.cell(r, 3).value or ""
        ref = cc.cell(r, 7).value or ""
        detail = cc.cell(r, 9).value or ""
        a = assets.get(tag, {})
        i = r
        new = custody_detail(et, tag, ref, detail, a, hr, transfers, tickets, xfers_by_tag, i)
        cc.cell(r, 9).value = unique_write(used, "detail", new, f"{tag}-{et}-{r}")


def custody_detail(et, tag, ref, old, a, hr, transfers, tickets, xfers_by_tag, i) -> str:
    kv = parse_kv(ref)
    tr = kv.get("Transfer") or gid(r"(TR-\d+)", ref)
    trow = transfers.get(tr or "", {}) or xfers_by_tag.get(tag, {})
    rec = trow.get("received_date") or ""
    tdate = trow.get("transfer_date") or ""
    apr = trow.get("approval_id") or kv.get("Approval")
    po = kv.get("PO") or gid(r"(PO-[\d-]+)", ref)
    model = a.get("model") or ""
    cost = a.get("cost")
    loc = a.get("ver_loc") or cc_loc_from(ref)

    if et == "Purchase":
        opts = [
            f"{po}: {model} at ${cost}.",
            f"Bought the {model} (${cost}) on {po}.",
            f"Receiving against {po}. {model}, ${cost}.",
            f"${cost} {model}. {po}.",
            f"PO line {po} — {model}.",
            f"Acquired {tag} ({model}) for ${cost} via {po}.",
            f"{model} came in on {po} (${cost}).",
            f"Purchasing file {po} matches {tag} at ${cost}.",
            f"Warehouse receipt tied to {po}. {model}.",
            f"{tag} origin: {po}, ${cost}.",
        ]
        return opts[i % len(opts)]

    if et == "Assignment/Transfer":
        rec_s = fmt_date(rec, i) if rec else ""
        td_s = fmt_date(tdate, i + 2) if tdate else ""
        to_c = trow.get("to_custodian") or ""
        to_l = trow.get("to_location") or loc
        if not rec:
            opts = [
                f"{tr} logged {td_s or 'on the transfer file'}; received_date is blank.",
                f"Move recorded ({tr}). No inbound stamp.",
                f"{tr} to {to_c} / {to_l}. Still waiting on a receive date.",
                f"Transfer {tr}" + (f" {apr}" if apr else "") + " — receive field empty.",
                f"Out of Central Receiving on {tr}. Not marked received.",
                f"{tr}: assignment posted, inbound date missing.",
                f"Log has {tr} ({td_s or 'dated on file'}) but no received_date.",
                f"Don't treat {tr} as docked. Receive column blank.",
            ]
        else:
            opts = [
                f"{tr} {td_s}; inbound {rec_s}.",
                f"Moved to {to_c} in {to_l} on {tr}. Received {rec_s}.",
                f"Transfer {tr}" + (f" / {apr}" if apr else "") + f". Received {rec_s}.",
                f"Assignment posted {td_s}. Docked {rec_s} ({tr}).",
                f"{to_l}: {tr} received {rec_s}.",
                f"From warehouse to {to_c}. {tr}, received {rec_s}.",
                f"Movement {tr} completed inbound {rec_s}.",
                f"Log row {tr}. Receive date {rec_s}.",
                f"{tr} — got it {rec_s} at {to_l}.",
                f"Recorded {td_s}, received {rec_s}. {tr}.",
            ]
        return opts[i % len(opts)]

    if et == "Policy Control Figure":
        opts = [
            f"Flipped to ITAM_control_matrix.png for the {tag} path (assignment/return/disposal bars).",
            f"Policy PDF + matrix on this chain. Figure is the standard, not a substitute for {tr or 'the transaction file'}.",
            f"Used the control matrix as the bar for {tag}. Still need the TR/PO/FA/ticket.",
            f"ITAM_control_matrix.png — return row says carrier + dock, not a note.",
            f"Looked at the figure while scoring {tag}. Appendix A, not evidence of a move.",
            f"Matrix/policy pair cited on {tag}. Doesn't prove location by itself.",
            f"Control figure on file (ITAM_control_matrix.png). Applied the return/disposal tests here.",
            f"{tag}: checked the matrix before I accepted a tech note.",
            f"IT_asset_management_policy.pdf with the PNG appendix. That's the yardstick for this chain.",
            f"Evidence standard from the figure — assignment / transfer / return / disposal / financial — applied to {tag}.",
            f"Didn't skip the matrix. Also didn't pretend it was a receipt for {tag}.",
            f"Appendix A open while I walked {tag}.",
            f"PNG control sheet referenced. {tag} still lives or dies on the logs.",
            f"Policy figure is in the packet. I used it as the rubric for this asset.",
            f"ITAM-001 appendix picture: what 'done' looks like for {tag}. Transactions still required.",
        ]
        return opts[i % len(opts)]

    if et == "Reconciliation Conclusion":
        st = a.get("ver_st") or ""
        conf = a.get("conf") or ""
        ex = a.get("exids") or kv.get("exceptions") or ""
        # keep IDs from ref
        bits = tokens(ref)
        opts = [
            f"Landed on {st} ({conf}). {ex or 'no exception line' if st=='In Use' else ex}.",
            f"{tag} → {st}. Confidence {conf}. {ex}".strip(),
            f"Call: {st} at {a.get('ver_loc')}. {conf}. Exceptions {ex or 'none listed'}.",
            f"After the chain: {st}. {conf} confidence. {ex}".strip(),
            f"Verified {st}. {ex}. {conf}.",
            f"Closing note {tag}: {st} / {a.get('ver_loc')} ({conf}). {ex}".strip(),
            f"{st} is the status I'm standing behind. {ex}. {conf}.",
            f"Wrap: {tag} {st}. {ex or 'clean of blocking EX lines' if 'Missing' not in st and 'Overdue' not in st and 'Exception' not in st else ex}.",
        ]
        text = opts[i % len(opts)]
        miss = [t for t in bits if t not in text]
        if miss:
            text = text.rstrip(".") + ". " + " ".join(miss[:8])
        return text

    if et == "Regional IT Note Lead":
        opts = [
            f"regional_IT_notes.docx mentions closed-office leftover color on {tag}. Notes don't beat a scan.",
            f"Tech note in the Word dump. Treated as a lead for {tag}, not location proof.",
            f"Breadcrumb in regional_IT_notes.docx. Still need a transaction for {tag}.",
            f"Someone wrote that it stayed at a closed site. Unverified for {tag}.",
            f"I read the note. I didn't relocate {tag} off of it.",
            f"Word file claim on {tag} — non-authoritative.",
            f"Closed-floor story in the notes. Ignored as clearance.",
            f"regional_IT_notes.docx is in the packet. Weight = lead only ({tag}).",
            f"Note exists. Corroboration doesn't. {tag}.",
            f"Didn't let the regional write-up overrule the logs on {tag}.",
        ]
        return opts[i % len(opts)]

    if et == "Offboarding Ticket":
        tk = kv.get("Ticket") or gid(r"((?:OFF|RET)-\d+)", ref)
        t = tickets.get(tk or "", {})
        due = t.get("return_due_date") or kv.get("due")
        req = t.get("return_required") or ""
        stt = t.get("ticket_status") or kv.get("Status")
        due_s = fmt_date(due, i) if due else ""
        opts = [
            f"{tk} {stt}. Return required={req}" + (f", due {due_s}" if due_s else "") + ".",
            f"Offboarding {tk}: {req or 'return flag'}." + (f" Due {due_s}." if due_s else ""),
            f"{tk} sits {stt}. Hardware return {'was required' if req=='Yes' else 'not required per ticket'}."
            + (f" {due_s}." if due_s else ""),
            f"People ticket {tk}. Doesn't by itself dock {tag}.",
            f"{tk} — {stt}. Due {due_s or 'n/a'}.",
            f"Service desk {tk} on {tag}. Return required {req}.",
            f"Opened {tk}. Status {stt}.",
            f"{tag} tied to {tk} ({stt})." + (f" SLA {due_s}." if due_s else ""),
        ]
        return opts[i % len(opts)]

    if et == "Technician Comment":
        tk = kv.get("Ticket") or gid(r"((?:OFF|RET)-\d+)", ref)
        t = tickets.get(tk or "", {})
        cmt = (t.get("technician_comment") or old or "").strip()
        if cmt:
            # already unique in source; trim if huge
            return cmt if len(cmt) < 240 else cmt[:237] + "…"
        return f"Tech comment on {tk or tag}: see ticket."

    if et == "Return Label Created":
        track = kv.get("Tracking") or gid(r"(1ZMD\d+)", ref)
        opts = [
            f"{track} is Label Created. Serial on the label {a.get('serial')}. Not proof of return.",
            f"Prepaid label {track} exists. Carrier never accepted it.",
            f"Label-only on {tag}. {track}.",
            f"Printed {track}. Box not in the inbound file.",
            f"§5: label ≠ returned. {track} / {a.get('serial')}.",
            f"Ship record {track} hasn't left Label Created.",
            f"I didn't close {tag} on the 1Z print. {track}.",
            f"Return paperwork only. {track}.",
        ]
        return opts[i % len(opts)]

    if et == "Shipment Exception":
        track = kv.get("Tracking") or ""
        opts = [
            f"{track}: serial on the shipment is not {a.get('serial')}. Custody not cleared.",
            f"Mismatch inbound. {tag} stays exception. {track}.",
            f"Carrier exception. Register serial {a.get('serial')} didn't show up on the parcel record.",
            f"Don't Available {tag}. {track} serial fight.",
        ]
        if tag == "MD-00082":
            opts.append("1ZMD00000082 / MISMATCH-0082 vs MD-MO-050082. Held.")
        return opts[i % len(opts)]

    if et == "Carrier Acceptance":
        track = kv.get("Tracking") or ""
        opts = [
            f"Carrier accepted {track}.",
            f"Acceptance event on {track}.",
            f"{track} picked up (acceptance on file).",
            f"UPS/carrier take recorded for {track}.",
        ]
        return opts[i % len(opts)]

    if et == "Delivery":
        track = kv.get("Tracking") or ""
        opts = [
            f"Delivery posted on {track}.",
            f"{track} delivered per carrier.",
            f"Carrier delivery event, {track}.",
            f"Delivered — {track}. Still need a receiving scan to finish returns.",
        ]
        return opts[i % len(opts)]

    if et == "Receiving Scan":
        track = kv.get("Tracking") or ""
        opts = [
            f"Dock scan on {track}.",
            f"Receiving scanned {track}.",
            f"SCAN-86 equivalent on file for {track}.",
            f"Inbound scan completed ({track}).",
        ]
        return opts[i % len(opts)]

    if et == "Dock Image Lead":
        return (
            "Atlanta photo receiving_exception_scan_1ZMD00000082.png shows HOLD / MISMATCH. "
            "Lead only — not a receiving clearance for MD-00082."
        )

    if et == "Unresolved Assumption":
        opts = [
            f"Still no receiving scan, disposal cert, or count that places {tag}.",
            f"{tag} location is assumed, not proven.",
            f"Gap: nothing physical confirms {tag} right now.",
            f"Unresolved. {tag} isn't in inbound or the cert file.",
            f"I won't invent a floor for {tag}.",
            f"Open assumption — current site unknown ({tag}).",
        ]
        return opts[i % len(opts)]

    if et == "Disposal Evidence Gap":
        opts = [
            f"{tr or kv.get('Transfer')}: recycler move without a matching certificate.",
            f"No DC for {tag}. Transfer {tr} isn't enough.",
            f"Certificate missing on {tag}. Retirement not closed.",
            f"{tag} pending-disposal with an empty cert field.",
        ]
        return opts[i % len(opts)]

    if et == "Disposal Certificate":
        dc = kv.get("Disposal") or ""
        return [
            f"{dc} on file (verified). {old}",
            f"Certificate {dc} matches {tag}.",
            f"Disposal paperwork {dc} present.",
            f"{dc} — method recorded in the disposal file.",
        ][i % 4]

    if et == "Technician/Finance Claim":
        return (
            "RN-0132 said MD-00132 was expensed under threshold. Rejected: acquisition cost $6,470 "
            "is ≥ $2,500, and the FAR still has no row."
        )

    if et == "Unresolved Capitalization":
        return (
            "MD-00132 / PO-2023-0022: capital-qualifying, no FA row. Blocks certification until "
            "capitalization or an approved write-off. MD-00130/MD-00131 are a different (under-threshold) story."
        )

    return str(old)


def cc_loc_from(ref: str) -> str:
    return ""


def report(used) -> None:
    def stats(name):
        c = Counter(used[name])
        # used is a set of unique strings we wrote; re-read file for actual dups
        print(f"{name}: unique generated {len(used[name])}")

    wb = load_workbook(ROOT / "Yanou_IT_Asset_Reconciliation.xlsx", data_only=True)
    er, cr, cc = wb["Exception Register"], wb["Corrected Register"], wb["Custody Chain"]

    def col_stats(ws, col, start=5):
        vals = [ws.cell(r, col).value for r in range(start, ws.max_row + 1)]
        c = Counter(vals)
        top = c.most_common(5)
        return len(vals), len(c), top[0][1] if top else 0, top[:3]

    n, u, mx, top = col_stats(er, 12)
    print(f"Evidence Assessment: {u}/{n} unique, max dup {mx}", top)
    n, u, mx, top = col_stats(er, 8)
    print(f"Required Action: {u}/{n} unique, max dup {mx}", top)
    n, u, mx, top = col_stats(cr, 11)
    print(f"Evidence Source: {u}/{n} unique, max dup {mx}", top)
    n, u, mx, top = col_stats(cc, 9)
    print(f"Custody Detail: {u}/{n} unique, max dup {mx}", top)

    banned = [
        "Transaction evidence reviewed",
        "Control matrix figure used as evidence-standard reference (Assignment/Transfer/Return/Disposal/Financial)",
        "Transfer recorded; received",
        "Leave the register row;",
        "Not taking the register row as current.",
        "Matrix figure used as the control reference.",
    ]
    blob = []
    for ws, col in ((er, 8), (er, 12), (cr, 11), (cc, 9)):
        for r in range(5, ws.max_row + 1):
            blob.append(str(ws.cell(r, col).value or ""))
    for b in banned:
        hits = sum(1 for x in blob if b in x)
        print(f"banned '{b[:50]}' hits={hits}")


if __name__ == "__main__":
    main()
